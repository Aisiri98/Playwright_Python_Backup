import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page
from playwright.sync_api import expect

invalid_entries =[]
from decimal import Decimal, ROUND_HALF_UP
def format_number(num_str: str) -> str:
    """Convert number string like '149,760' → '149.76K', '268,000' → '268K'.
       Uses ROUND_HALF_UP and trims unnecessary trailing zeros."""
    num_str = num_str.replace(",", "").strip()

    try:
        num = int(num_str)
    except ValueError:
        raise ValueError(f"Invalid number string: {num_str}")

    def format_decimal(val: Decimal, suffix: str) -> str:
        # Round to 2 decimals
        val = val.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # Convert to string and strip trailing zeros/decimal if not needed
        val_str = f"{val:.2f}".rstrip("0").rstrip(".")
        return f"{val_str}{suffix}"

    if num >= 1_000_000:
        return format_decimal(Decimal(num) / Decimal(1_000_000), "M")
    elif num >= 1_000:
        return format_decimal(Decimal(num) / Decimal(1_000), "K")
    else:
        return str(num)
class FrieghtFilter:
    def __init__(self, page: Page):
        self.C_O_field = None
        self.C_E_field = None
        self.S_JC_field = None
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def check_freight_filter(self):
        # Wait for the mode of transportation label
        Mode_of_Transportation_locator = self.page.locator('[class="_filteritemtext_a6k5f_88"]').first
        expect(Mode_of_Transportation_locator).to_be_visible(timeout=100000)

        # Locate freight filter options
        elements = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count = elements.count()
        print(f"Found {count} Frieght filter options")

        # Loop through each element
        for i in range(count):
            element = elements.nth(i)
            if element.is_visible():
                text = element.inner_text()
                print(f"✅ freight option {i + 1} is visible: {text}")
            else:
                print(f"❌ freight option {i + 1} is not visible")

        # Click on the freight filter again (collapse/close)
        self.page.locator("//span[normalize-space()='Freight']").click()
        self.page.wait_for_timeout(2000)  # give time for collapse

        # Re-locate elements again after DOM change
        elements_after = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count_after = elements_after.count()
        print(f"Checking {count_after} freight filter options")

        # Expect no options should be visible
        for i in range(count_after):
            element = elements_after.nth(i)
            expect(element).not_to_be_visible(timeout=5000)

    def open_freight_dropdown(self):
        frieght_locator = self.page.locator('[class="_filteritemtext_a6k5f_88"]').first
        self.page.pause()

        # Check if "port lading" is visible
        if not frieght_locator.is_visible():
            print("📌 Freight dropdown is closed. Clicking to open...")
            self.page.locator("//span[normalize-space()='Freight']").click()
        else:
            print("✅ freight dropdown is already open.")

    def Click_Mode_of_Transportation(self):
        #Mode of trasportation filter modal opens with sub-filters such as:
        expect(self.page.locator('[class="_filteritemtext_a6k5f_88"]').first).to_be_visible()
        self. page.get_by_text("Mode of Transportation").first.click()
        self.page.wait_for_timeout(2000)
        expect(self.page.get_by_text("Filter by Mode of Transportation").first).to_be_visible()
        #Observe the list of  Mode of Transportation Types


    def Freight_Filter_Search(self, M_T_F: str):
        # Use the search bar in the modal to search for a consignee (e.g., "Walmart")
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").last.fill(M_T_F)
        self.page.wait_for_timeout(2000)
        # Check for the presence of the company name in the results
        locators = self.page.locator('[class="text-no-transform"]')
        count = locators.count()
        print(f"🔍 Found {count} checkbox")

        if count == 0:
            print("❌ FAIL: No elements found for search")
            self.page.locator('input[type="text"]').last.clear()
        else:
            found_match = False

            for i in range(count):
                locator = locators.nth(i)

                if locator.is_visible():
                    actual_text = locator.inner_text().strip()

                    if M_T_F.lower() in actual_text.lower():  # case-insensitive compare
                        print(f"✅ PASS: Search is working → '{M_T_F}' found in element {i + 1}: '{actual_text}'")
                        expect(locator).to_contain_text(M_T_F)
                        found_match = True
                    else:
                        print(f"❌ FAIL: Search text '{M_T_F}' not found in element {i + 1}. Actual: '{actual_text}'")
                else:
                    print(f"⚠️ Element {i + 1} is not visible")

            if not found_match:
                print(f"❌ FAIL: Search text '{M_T_F}' not found in any visible element")
                self.page.locator('input[type="text"]').clear()

    from playwright.sync_api import expect

    def validate_freight_type_checkboxes(self):
        # Locate all checkboxes with their labels
        checkboxes = self.page.locator("//label[input[@type='checkbox']]")

        count = checkboxes.count()
        print(f"Total checkboxes found: {count}")

        for i in range(count):
            checkbox_label = checkboxes.nth(i)

            # Ensure checkbox is visible
            expect(checkbox_label).to_be_visible(timeout=5000)

            # Extract the label text (like "Importer (245,542,386)")
            text_value = checkbox_label.inner_text().strip()
            print(f"Checkbox {i + 1} with label: {text_value} found in Mode of transportation filter")

            # Validate checkbox input exists inside label
            input_box = checkbox_label.locator("input[type='checkbox']")
            expect(input_box).to_be_visible()

            # You can also assert that text is not empty
            assert text_value != "", f"Checkbox {i + 1} has no text!"

    def Apply_Filter_Functionality(self):
        last_input = self.page.locator('input[type="text"]').last
        value = last_input.input_value().strip()

        if value == "":
            print("ℹ️  input box is already empty. Skipping clear()...")
        else:
            print(f"🧹 Clearing input box (current value: '{value}')")
            last_input.clear()
        self.page.wait_for_timeout(2000)
        # Click on the checkbox of the first result
        expect(self.page.locator('[class="text-no-transform"]').nth(1)).to_be_visible()
        self.Mode_of_Transport_selected = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Selected value is : {self.Mode_of_Transport_selected}")

        import re
        # Calculate total from all .grey-text elements
        total = 0

        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        shipper_record_clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        shipper_record_num = float(shipper_record_clean)
        if 0 <= shipper_record_num < 100_000:
            self.total_in_millions = f"{shipper_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{shipper_record_num / 1_000_000:.2f}M"

        print(self.total_in_millions)
        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        return False

    def Validate_ShipmentRecord_Count(self):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=10000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(3000)
        # Validate total shipment count in 'Shipment' tab
        record_count = self.page.locator(
            "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']"
        )

        # Get actual text
        actual_text = record_count.inner_text()

        # If–else validation
        if self.total_in_millions == actual_text:
            print(f"✅ Record count matches: {actual_text}")
        else:
            print(f"❌ Record count mismatch. Expected {self.total_in_millions}, Found {actual_text}")

    def Validate_Total_RecordCount_In_the_Table(self):
        total_record_text = self.page.locator('[class="trademo-table-count-text"]').inner_text()
        total_from_summary = total_record_text.split("of")[1].strip()
        print(f"🔢 Value after 'of': {total_from_summary}")
        print(f"total record count:{self.Total_Record_Count}")
        if total_from_summary == self.Total_Record_Count:
            print(f"✅ Record count matched: {total_from_summary}")
        else:
            print(f"❌ Record count mismatch! Summary: {total_from_summary}, Table: {self.Total_Record_Count}")

    def Validate_Freight_Name_label(self):
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')
        # Click on the consignee filter dropdown
        self.page.locator("//span[normalize-space()='Freight']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        self.page.locator("//span[normalize-space()='Freight']").click()
        self.page.wait_for_timeout(1000)

    def check_mode_of_transport_inthe_grid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.M_T_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "mode of transportation":
                self.M_T_field = i

        # --- Handle missing "Consignee JC" column ---
        if self.M_T_field is None:
            raise Exception("❌ 'port of lading' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Mode of transportation: {self.M_T_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                mode_transport = row_data[self.M_T_field]
                mode_transport_norm = re.sub(r"\s+", " ", mode_transport).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.Mode_of_Transport_selected.lower() in mode_transport_norm:
                        print(f"✅ Grid valid: {mode_transport} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {mode_transport} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": mode_transport, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    expect(self.page.locator("//span[normalize-space()='Shipment Origin Details']")).to_be_visible(
                        timeout=100000)
                    popup_mode_transport = self.page.locator(
                        "//span[normalize-space(text())='Mode of Transportation']/following-sibling::p").inner_text().strip().lower()

                    if self.Mode_of_Transport_selected.lower() in popup_mode_transport:
                        print(f"✅ View details valid: {popup_mode_transport} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ View details invalid: {popup_mode_transport} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in view details"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_mode_transport, "Source": "Popup"})
                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Frieght filter Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.M_T_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.M_T_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/Freight_filter_{self.Mode_of_Transport_selected}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid mode of transport. Excel not generated.")

        return self.invalid_entries

    def Apply_filter_Exclude(self):
        #Again click Freight filter Exclude all options in Mode of Transport and apply the filter
        self.page.locator("//label[normalize-space()='Exclude Selection']").click()
        self.page.wait_for_timeout(2000)
        #check all the checkbox if it is unchecked
        checkboxes = self.page.locator("//label[input[@type='checkbox']]")
        count = checkboxes.count()
        print(f"Total checkboxes found: {count}")

        for i in range(count):
            checkbox_label = checkboxes.nth(i)

            # Get the actual input element
            input_box = checkbox_label.locator("input[type='checkbox']")

            # Check only if visible in the viewport
            if input_box.is_visible():
                text_value = checkbox_label.inner_text().strip()
                print(f"Checkbox {i + 1} with label: {text_value} is visible")

                if not input_box.is_checked():
                    input_box.check()
                    print(f"✅ Checkbox {i + 1} ({text_value}) is now checked")
                else:
                    print(f"ℹ️ Checkbox {i + 1} ({text_value}) was already checked")
            else:
                print(f"⏭️ Checkbox {i + 1} skipped because it is not visible")

        self.page.locator("//button[normalize-space()='Apply Filter']").click()

    def check_mode_of_transport_exclude_inthe_grid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()
        valid_modes = ["air", "maritime", "land", "others"]
        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.M_T_E_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "mode of transportation":
                self.M_T_E_field = i

        # --- Handle missing "Consignee JC" column ---
        if self.M_T_E_field is None:
            raise Exception("❌ 'port of lading' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Port lading: {self.M_T_E_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                mode_transport = row_data[self.M_T_field]
                mode_transport_norm = re.sub(r"\s+", " ", mode_transport).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if mode_transport_norm not in valid_modes:
                        print(f"✅ Grid- Exclude:Shows different option {mode_transport}  at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid-Exclude:selected option : {mode_transport} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": mode_transport, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    expect(self.page.locator("//span[normalize-space()='Shipment Origin Details']")).to_be_visible(
                        timeout=100000)
                    popup_mode_transport = self.page.locator(
                        "//span[normalize-space()='Mode of Transportation']/following-sibling::div[contains(normalize-space(.), 'NA')]").inner_text().strip().lower()

                    if popup_mode_transport not in valid_modes:
                        print(f"✅ View- Exclude:: Shows different option {popup_mode_transport}  for Sl. No: {sl_no}")
                    else:
                        print(f"❌ View -Exclude: Shows selected option: {popup_mode_transport}  for Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in view details"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_mode_transport, "Source": "Popup"})
                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Freight Filter Exclude Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.M_T_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.M_T_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/Freight_filter_Exclude.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid mode of transport. Excel not generated.")

        return self.invalid_entries








