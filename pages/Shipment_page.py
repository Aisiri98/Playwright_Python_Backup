import inspect
import os
import re
import tempfile
from openpyxl.styles import Font
import pypdf
from pypdf import PdfReader
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page
from playwright.sync_api import expect, TimeoutError as PlaywrightTimeoutError
from datetime import datetime
invalid_entries =[]
import difflib
class ShipmentFilter:
    def __init__(self, page: Page):
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def check_India_import(self):
        text_content = self.page.locator(".tw-text-nowrap > span").text_content()

        # Select any country from the Data Source filter and click Apply
        if "India Imports" not in text_content:
            self.page.locator("[class='tw-flex tw-items-center tw-gap-3 tw-border-none tw-text-sm']").click()
            clear_all_btn = self.page.get_by_role("button", name="Clear All")
            # Wait for the element to be attached to DOM
            if clear_all_btn.is_visible() and clear_all_btn.is_enabled():
                clear_all_btn.click()
                print("✅ 'Clear All' button clicked.")
            else:
                print("ℹ️ 'Clear All' button is either not visible or disabled.")
            # Loop through each row
            rows = self.page.locator('[class="tw-flex tw-items-center tw-w-full"]')
            count = rows.count()

            for i in range(count):
                row_text = rows.nth(i).inner_text().strip()
                if "India" in row_text:
                    rows.nth(i).locator('input[type="checkbox"]').first.click()
                    break

            # Click on Apply button
            self.page.locator("//span[normalize-space()='Apply']").click()
            print("India import is selected in filter")

        else:
            print("India import is already selected in filter. No action needed.")

    def Verify_Reset_button(self):
        reset_locator = self.page.get_by_text(" Reset").first
        if reset_locator.is_visible():
            print("🔍 ' Reset' is visible, clicking...")
            reset_locator.click()
        else:
            print("ℹ️ ' Reset' not visible.")

    def Verify_Shipment_Tab(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_contain_text("Shipments")
        expect(self.page.locator("//a[@id='nav-profile-tab']")).to_contain_text("Importers")
        expect(self.page.locator("//a[@id='nav-contact-tab']")).to_contain_text("Exporters")
        expect(self.page.locator("//a[@id='nav-summary-tab']")).to_contain_text("Analytics")
        print("Validated all tabs in Shipment page")

    def Sort_Billing_Entry_Column(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        self.page.locator("//div[contains(text(),'Bill of Entry Date')]").click()
        self.page.wait_for_timeout(2000)
        # Locator for all "Bill of Entry Date" column values
        # Get all dates from the column
        date_locator = self.page.locator("td:nth-child(4)")
        date_count = date_locator.count()

        dates = [
            datetime.strptime(date_locator.nth(i).inner_text().strip(), "%b %d, %Y")
            for i in range(date_count)
        ]

        # ✅ Check ascending row-by-row
        is_ascending = True
        for i in range(len(dates) - 1):
            if dates[i] > dates[i + 1]:
                print(f"❌ Ascending sort issue at row {i + 1}: {dates[i]} comes after {dates[i + 1]}")
                is_ascending = False

        if is_ascending:
            print("✅ Dates are in ascending order.")

        # Click again to sort descending
        self.page.locator("//div[contains(text(),'Bill of Entry Date')]").click()
        self.page.wait_for_timeout(2000)

        dates_desc = [
            datetime.strptime(date_locator.nth(i).inner_text().strip(), "%b %d, %Y")
            for i in range(date_count)
        ]

        # ✅ Check descending row-by-row
        is_descending = True
        for i in range(len(dates_desc) - 1):
            if dates_desc[i] < dates_desc[i + 1]:
                print(f"❌ Descending sort issue at row {i + 1}: {dates_desc[i]} comes before {dates_desc[i + 1]}")
                is_descending = False

        if is_descending:
            print("✅ Dates are in descending order.")

    def Sort_CIF_Column(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)

        # Click once to sort descending first
        self.page.locator("//div[contains(text(),'CIF Value (USD)')]").click()
        self.page.wait_for_timeout(2000)

        # Common locator for CIF column (14th column in table)
        cif_locator = self.page.locator("td:nth-child(14)")
        cif_count = cif_locator.count()

        def extract_numeric_value(text):
            """Return float value including negatives if number exists, else None."""
            text = text.strip().replace(",", "")
            match = re.search(r"-?\d+(?:\.\d+)?", text)  # ✅ captures negative numbers too
            return float(match.group()) if match else None

        # --- DESCENDING CHECK ---
        cif_values_desc = [extract_numeric_value(cif_locator.nth(i).inner_text()) for i in range(cif_count)]
        cif_values_desc_filtered = [v for v in cif_values_desc if v is not None]

        is_descending = all(
            cif_values_desc_filtered[i] >= cif_values_desc_filtered[i + 1]
            for i in range(len(cif_values_desc_filtered) - 1)
        )

        if is_descending:
            print("✅ CIF values are in descending order.")
        else:
            for i in range(len(cif_values_desc_filtered) - 1):
                if cif_values_desc_filtered[i] < cif_values_desc_filtered[i + 1]:
                    print(
                        f"❌ Descending sort issue at row {i + 1}: "
                        f"{cif_values_desc_filtered[i]} < {cif_values_desc_filtered[i + 1]}"
                    )

        # Click again to sort ascending
        self.page.locator("//div[contains(text(),'CIF Value (USD)')]").click()
        self.page.wait_for_timeout(2000)

        # --- ASCENDING CHECK ---
        cif_values = [extract_numeric_value(cif_locator.nth(i).inner_text()) for i in range(cif_count)]
        cif_values_filtered = [v for v in cif_values if v is not None]

        is_ascending = all(
            cif_values_filtered[i] <= cif_values_filtered[i + 1]
            for i in range(len(cif_values_filtered) - 1)
        )

        if is_ascending:
            print("✅ CIF values are in ascending order.")
        else:
            for i in range(len(cif_values_filtered) - 1):
                if cif_values_filtered[i] > cif_values_filtered[i + 1]:
                    print(
                        f"❌ Ascending sort issue at row {i + 1}: "
                        f"{cif_values_filtered[i]} > {cif_values_filtered[i + 1]}"
                    )

    def Validate_Download_Screen(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        #Click on the Download button in the shipment grid
        self.page.locator("//span[normalize-space()='Download']").click()
        expect(self.page.locator("//span[normalize-space()='All Columns']")).to_be_visible()
        expect(self.page.locator("//span[normalize-space()='Columns Shown in Grid']")).to_be_visible()
        #Choose either option

    def Validate_Cancel_Close_button(self):
        self.page.locator("//span[normalize-space()='All Columns']").click()
        print("All columns is visible")
        #Verify the modal includes:
        expect(self.page.locator("//div[@class='modal-title h4']")).to_contain_text("Download Shipments")
        #Click on Cancel button
        self.page.locator("//button[@class='btn btn-outline-secondary trademo-outline-btn btn-sm mr-2']").click()
        expect(self.page.locator("//div[@class='modal-title h4']")).not_to_be_visible()
        # Click on the Download button in the shipment grid
        self.page.locator("//span[normalize-space()='Download']").click()
        self.page.locator("//span[normalize-space()='All Columns']").click()
        expect(self.page.locator("//div[@class='modal-title h4']")).to_contain_text("Download Shipments")
        self.page.locator("//span[@aria-hidden='true']").click()

    def Validate_Download_GridColumn_Shipment(self,range:str,range1:str):
        self.page.locator("//span[normalize-space()='Columns Shown in Grid']").click()
        expect(self.page.locator("//div[@class='modal-title h4']")).to_contain_text("Download Shipments")
        #Verify the modal includes:
        self.page.locator("//input[@placeholder='To']").clear()
        self.page.locator("//input[@placeholder='To']").fill(range)
        expect(self.page.locator("//button[normalize-space()='Email']")).to_be_visible()
        self.page.locator("//input[@placeholder='To']").clear()
        self.page.locator("//input[@placeholder='To']").fill(range1)
        expect(self.page.locator("//button[normalize-space()='Download']")).to_be_visible()

    def Validate_Download_All_Column_Shipment(self, range: str, range1: str):
        self.page.locator("//span[normalize-space()='All Columns']").click()
        expect(self.page.locator("//div[@class='modal-title h4']")).to_contain_text("Download Shipments")
        # Verify the modal includes:
        self.page.locator("//input[@placeholder='To']").clear()
        self.page.locator("//input[@placeholder='To']").fill(range)
        expect(self.page.locator("//button[normalize-space()='Email']")).to_be_visible()
        self.page.locator("//input[@placeholder='To']").clear()
        self.page.locator("//input[@placeholder='To']").fill(range1)
        expect(self.page.locator("//button[normalize-space()='Download']")).to_be_visible()

    def Validate_Colmn_Shown_In_the_Grid_Download(self):
        import os
        import tempfile
        import openpyxl
        from datetime import datetime

        # --- Trigger download ---
        with self.page.expect_download() as download_info:
            self.page.locator("//button[normalize-space()='Download']").click()
        download = download_info.value

        # --- Save to temp file ---
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, download.suggested_filename)
        download.save_as(file_path)
        print(f"📂 Downloaded to: {file_path}")

        # --- Read workbook ---
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # --- Validate "Report Summary" sheet ---
        if "Report Summary" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Report Summary' sheet not found in downloaded Excel.")
        summary_sheet = wb["Report Summary"]

        EXPECTED_REPORT_TYPE = "Shipment Records - Customised Columns"
        EXPECTED_DOWNLOAD_DATE = datetime.today().strftime("%b %d, %Y")

        report_type_value = str(summary_sheet["B7"].value).strip() if summary_sheet["B7"].value else ""
        download_date_value = str(summary_sheet["B9"].value).strip() if summary_sheet["B9"].value else ""

        if report_type_value == EXPECTED_REPORT_TYPE:
            print("✅ Report Type matches expected value.")
        else:
            print(f"❌ Report Type mismatch: Found '{report_type_value}', Expected '{EXPECTED_REPORT_TYPE}'")

        if download_date_value == EXPECTED_DOWNLOAD_DATE:
            print("✅ Download Date matches today's date.")
        else:
            print(f"❌ Download Date mismatch: Found '{download_date_value}', Expected '{EXPECTED_DOWNLOAD_DATE}'")

        # --- Read "Shipment Data" sheet ---
        if "Shipment Data" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Shipment Data' sheet not found in downloaded Excel.")
        sheet = wb["Shipment Data"]

        # Read Excel headers
        downloaded_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        wb.close()

        # --- Get UI headers ---
        ui_headers = self.page.locator("//table/thead/tr/th").all_inner_texts()
        ui_headers = [h.strip() for h in ui_headers if h not in ("", "View")]

        # --- Normalize headers ---
        def normalize_header(h):
            if not h:
                return ""
            return str(h).strip().replace(".", "").replace("  ", " ").replace("'", "").replace('"', '').lower()

        excel_norm = [normalize_header(h) for h in downloaded_headers]
        ui_norm = [normalize_header(h) for h in ui_headers]

        # --- Columns to ignore in Excel ---
        ignore_cols = [
            "sno",
            "per unit price (usd)",
            "shipper standardized name",
            "consigne standardized name",
            "quantity unit",
            "s no",
            "per unit value (usd)"
        ]

        # Filter Excel and UI columns
        excel_filtered = [col for col in excel_norm if col not in ignore_cols]
        ui_filtered = [col for col in ui_norm if col not in ignore_cols]

        # --- Validation ---
        missing_ui_cols = [col for col in ui_filtered if col not in excel_filtered]
        extra_excel_cols = [col for col in excel_filtered if col not in ui_filtered]

        if not missing_ui_cols:
            print("✅ All relevant UI columns are present in Excel (ignored extra fields).")
        else:
            print("❌ Missing UI columns in Excel:", missing_ui_cols)

        if extra_excel_cols:
            print("⚠️ Extra columns in Excel (not in UI):", extra_excel_cols)

        # --- Cleanup ---
        os.remove(file_path)

    def Validate_All_Columns_Download(self):
        import os
        import tempfile
        import openpyxl
        from datetime import datetime

        # --- Trigger download ---
        with self.page.expect_download() as download_info:
            self.page.locator("//button[normalize-space()='Download']").click()
        download = download_info.value

        # --- Save to temp file ---
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, download.suggested_filename)
        download.save_as(file_path)
        print(f"📂 Downloaded to: {file_path}")

        # --- Read workbook ---
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # --- Validate "Report Summary" sheet ---
        if "Report Summary" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Report Summary' sheet not found in downloaded Excel.")
        summary_sheet = wb["Report Summary"]

        EXPECTED_REPORT_TYPE = "Shipment Records - All Columns"
        EXPECTED_DOWNLOAD_DATE = datetime.today().strftime("%b %d, %Y")

        report_type_value = str(summary_sheet["B7"].value).strip() if summary_sheet["B7"].value else ""
        download_date_value = str(summary_sheet["B9"].value).strip() if summary_sheet["B9"].value else ""

        if report_type_value == EXPECTED_REPORT_TYPE:
            print("✅ Report Type matches expected value.")
        else:
            print(f"❌ Report Type mismatch: Found '{report_type_value}', Expected '{EXPECTED_REPORT_TYPE}'")

        if download_date_value == EXPECTED_DOWNLOAD_DATE:
            print("✅ Download Date matches today's date.")
        else:
            print(f"❌ Download Date mismatch: Found '{download_date_value}', Expected '{EXPECTED_DOWNLOAD_DATE}'")

        # --- Read "Shipment Data" sheet ---
        if "Shipment Data" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Shipment Data' sheet not found in downloaded Excel.")
        sheet = wb["Shipment Data"]

        # Read Excel headers
        downloaded_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        wb.close()

        # Normalize headers (remove extra quotes, spaces, etc.)
        def normalize_header(h):
            if not h:
                return ""
            return str(h).strip().replace("'", "").replace('"', '').replace("  ", " ")

        excel_filtered = [normalize_header(h) for h in downloaded_headers]

        # UI headers (filtered and normalized)
        ui_headers = self.page.locator("//table/thead/tr/th").all_inner_texts()
        ui_headers = [
    "S.No", "Bill of Entry Date", "HS Code", "6 Digit HS Code", "Item Number", "Invoice Number",
    "Mode of Transportation", "Country of Origin", "Port of Lading", "Shipper Name", "Shipper Standardized Name",
    "Shipper Type", "Country of Export", "Shipper Address", "Port of Unlading", "Country of Import",
    "Consignee Name", "Consigne Standardized Name", "Consignee Type", "Consignee State", "Consignee City",
    "Consignee Pincode", "Consignee Address", "Consignee Phone Number", "Consignee Email ID",
    "Custom House Agent Code", "Custom House Agent Name", "Product Description", "Quantity", "Quantity Unit",
    "Per Unit Price (USD)", "CIF Value (USD)", "Duty Paid (USD)", "Shipper Jurisdiction Country",
    "Consignee Jurisdiction Country", "IEC Number"
]

        # --- Validation ---
        if excel_filtered == ui_headers:
            print("✅ All Columns - Downloaded Excel 'Shipment Data' columns match All extra columns exactly.")
        else:
            print("❌ Mismatch between Excel and All columns.")
            print("UI       :", ui_headers)
            print("Excel    :", excel_filtered)

        # --- Cleanup ---
        os.remove(file_path)

    def Validate_Checkbox(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        # Verify the functionality after selecting checkboxes in the shipment grid
        view_checkbox = self.page.locator("input[type='checkbox']").first
        # Locator for all individual checkboxes
        all_checkboxes = self.page.locator("td:nth-child(1) input[type='checkbox']")

        # Listen for the "View" checkbox state and check all if it is checked
        if not view_checkbox.is_checked():
            view_checkbox.check()
            print("✅ 'View' checkbox checked.")

        # Check all individual checkboxes
        count = all_checkboxes.count()
        for i in range(count):
            checkbox = all_checkboxes.nth(i)
            if checkbox.is_checked():
                print(f"✅ Checkbox at row {i + 1} checked.")
            else:
                print(f"ℹ️ Checkbox at row {i + 1} is not checked.")

        #Validate row selected banner
        expect(self.page.locator("span").filter(has_text=re.compile(r"^10$"))).to_be_visible(timeout=6000)

        #Uncheck View checkbox
        view_checkbox.uncheck()
        count = all_checkboxes.count()
        for i in range(count):
            checkbox = all_checkboxes.nth(i)
            if not checkbox.is_checked():
                print(f"✅ Checkbox at row {i + 1} unchecked.")
            else:
                print(f"ℹ️ Checkbox at row {i + 1} is checked.")

        #Check any two checkbox
        all_checkboxes.nth(0).check()
        all_checkboxes.nth(1).check()
        #Validate row selected banner
        expect(self.page.locator("span").filter(has_text=re.compile(r"^2$"))).to_be_visible(timeout=6000)

        #Click on close icon
        self.page.get_by_text("close").click()

        #All selected checkboxes should be cleared and banner should disappear
        expect(self.page.locator("span").filter(has_text=re.compile(r"^2$"))).not_to_be_visible(timeout=6000)
        checked_boxes = []
        for i in range(count):
            checkbox = all_checkboxes.nth(i)
            if checkbox.is_checked():
                checked_boxes.append(i + 1)

        # Validation
        assert not checked_boxes, f"❌ The following checkboxes are checked: {checked_boxes}"
        print("✅ All checkboxes are unchecked.")

    def Validate_View_Shipment(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        # Click on View button
        self.page.locator('tr:nth-child(4) > .sc-dCFHLb.dinpZO').click(timeout=10000)
        self.page.wait_for_timeout(3000)
        #Verify the view details page
        expect(self.page.get_by_role("dialog")).to_contain_text("Detailed Shipment")
        expect(self.page.get_by_role("dialog")).to_contain_text("Shipment Origin Details")
        expect(self.page.get_by_role("dialog")).to_contain_text("Shipment Destination Details")
        expect(self.page.get_by_role("dialog")).to_contain_text("Customs Details")
        expect(self.page.get_by_role("dialog")).to_contain_text("Cargo Details")
        self.page.get_by_role("button", name="Close").click()

    def Validate_Bookmark(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        self.page.pause()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        # Click on View button
        self.page.locator('td:nth-child(2)').first.click()
        self.page.wait_for_timeout(1000)
        #Verify the view details page
        expect(self.page.get_by_role("dialog")).to_contain_text("Detailed Shipment")
        #Verify the presence of the following in detail view:
        expect(self.page.locator('[class="_btn_e8xj1_1 _btn-primaryWithIcon_e8xj1_23 _size-large_e8xj1_96 mx-2 d-flex align-items-center justify-content-center"]')).to_contain_text("Download")
        expect(self.page.get_by_role("button", name="bookmark")).to_be_visible()
        expect( self.page.get_by_role("button", name="icon Take notes")).to_be_visible()
        expect(self.page.locator("//span[@aria-hidden='true']")).to_be_visible()
        #Click on Bookmark icon
        self.page.get_by_role("button", name="bookmark").click()
        expect(self.page.locator("//div[normalize-space()='Add to Bookmarks']")).to_be_visible()
        self.page.locator("//button[normalize-space()='Add Tags']").click()
        expect(self.page.locator("//div[contains(text(),'Shipment Bookmarked Successfully')]")).to_be_visible()
        self.page.locator("//span[@aria-hidden='true']").click()

        #Click on Bookmark
        self.page.locator("//p[normalize-space()='Shipment']").click()
        #validate added date
        from datetime import datetime

        today = datetime.today().strftime("%b %d, %Y")
        print(today)
        expect(self.page.locator('[class="text-wrap"]').nth(3)).to_contain_text(today)
        self.page.locator("//p[normalize-space()='Shipments']").click()

    def Validate_Download(self):
        self.page.locator("//p[normalize-space()='Shipments']").click()
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        # Click on View button
        self.page.locator('td:nth-child(2)').nth(1).click()
        self.page.wait_for_timeout(1000)
        #Verify the view details page
        expect(self.page.get_by_role("dialog")).to_contain_text("Detailed Shipment")
        #Verify the presence of the following in detail view:
        expect(self.page.locator('[class="_btn_e8xj1_1 _btn-primaryWithIcon_e8xj1_23 _size-large_e8xj1_96 mx-2 d-flex align-items-center justify-content-center"]')).to_contain_text("Download")
        expect(self.page.get_by_role("button", name="bookmark")).to_be_visible()
        expect(self.page.get_by_role("button", name="icon Take notes")).to_be_visible()
        expect(self.page.locator("//span[@aria-hidden='true']")).to_be_visible()

        #Click on download
        with self.page.expect_download() as download_info:
            self.page.locator(
                '[class="_btn_e8xj1_1 _btn-primaryWithIcon_e8xj1_23 _size-large_e8xj1_96 mx-2 d-flex align-items-center justify-content-center"]').click()
        download = download_info.value

        # --- Save PDF to a temp location ---
        temp_dir = tempfile.mkdtemp()
        pdf_path = os.path.join(temp_dir, download.suggested_filename)
        download.save_as(pdf_path)

        # --- Read PDF content ---
        pdf_reader = pypdf.PdfReader(pdf_path)
        pdf_text = ""
        for page_num in range(len(pdf_reader.pages)):
            pdf_text += pdf_reader.pages[page_num].extract_text()

        # --- Validate Headers ---
        assert "Shipment Origin Details" in pdf_text
        normalized_text = re.sub(r"\s+", " ", pdf_text)  # replace all whitespace/newlines with single space
        assert "Shipment Destination Details" in normalized_text
        assert "Custom Details" in pdf_text
        assert "Cargo Details" in pdf_text

        print("✅ PDF header content validated successfully.")
        self.page.locator("//span[@aria-hidden='true']").click()

    def Validate_take_note(self, note:str):
        self.page.locator("//p[normalize-space()='Shipments']").click()
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        # Click on View button
        self.page.locator('td:nth-child(2)').first.click()
        self.page.wait_for_timeout(1000)
        #Verify the view details page
        expect(self.page.get_by_role("dialog")).to_contain_text("Detailed Shipment")
        #Verify the presence of the following in detail view:
        expect(self.page.locator('[class="_btn_e8xj1_1 _btn-primaryWithIcon_e8xj1_23 _size-large_e8xj1_96 mx-2 d-flex align-items-center justify-content-center"]')).to_contain_text("Download")
        expect(self.page.get_by_role("button", name="bookmark")).to_be_visible()
        expect(self.page.get_by_role("button", name="icon Take notes")).to_be_visible()
        expect(self.page.locator("//span[@aria-hidden='true']")).to_be_visible()

        #Click on take note
        self.page.get_by_role("button", name="icon Take notes").click()
        self.page.pause()
        self.page.get_by_role("textbox").click()
        self.page.get_by_role("textbox").fill(note)
        self.page.get_by_role("button", name="Add Note").click()
        self.page.get_by_role("button", name="Close").click()
        self.page.get_by_text("format_list_bulletedShipment").click()
        self.page.get_by_role("img", name="comment").click()
        expect(self.page.get_by_role("tooltip")).to_contain_text(note)

    def Validate_Customise_Shipment_Grid(self):
        self.page.locator("//p[normalize-space()='Shipments']").click()
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        #In the shipment grid, click on Customise Shipment Grid button
        self.page.get_by_role("button", name="dashboard_customize Customise").click()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text(
            "Select columns for your grid.*You can select and display maximum of 15 columns.")
        #Verify the available section categories: Primary Details, Shipper Details, Consignee Details, custom details and Cargo Details
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("primary details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("shipper details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("consignee details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("custom details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("cargo details")

    def Customise_grid(self):
        Customise_grid = self.page.get_by_role("button", name="dashboard_customize Customise")
        if Customise_grid.is_visible():
            print("🔍 'Customise grid' is visible, clicking...")
            Customise_grid.click()
            expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text(
                "Select columns for your grid.*You can select and display maximum of 15 columns.")
        else:
            print("ℹ️ 'Customise grid' not visible.")
            self.page.locator("//a[@id='nav-home-tab']").click()
            expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
            # In the shipment grid, click on Customise Shipment Grid button
            self.page.get_by_role("button", name="dashboard_customize Customise").click()
            expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text(
                "Select columns for your grid.*You can select and display maximum of 15 columns.")


    def Validate_Customise_Shipment_Grid_checkbox(self):
        #Click each section and verify related column checkboxes (e.g., Shipment Date, Shipment Value under Primary Details)
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Bill of Entry Date")
        expect(self.page.get_by_role("checkbox", name="Bill of Entry Date")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("HS Code")
        expect(self.page.get_by_role("checkbox", name="HS Code")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Mode of Transportation")
        expect(self.page.get_by_role("checkbox", name="Mode of Transportation")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Product Description")
        expect(self.page.get_by_role("checkbox", name="Product Description")).to_be_visible()
        self.page.get_by_text("shipper details").click()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Country of Origin")
        expect(self.page.get_by_role("checkbox", name="Country of Origin")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Port of Lading")
        expect(self.page.get_by_role("checkbox", name="Port of Lading")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Shipper Name")
        expect(self.page.get_by_role("checkbox", name="Shipper Name")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Shipper Type")
        expect(self.page.get_by_role("checkbox", name="Shipper Type")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Shipper Jurisdiction Coun...")
        expect(self.page.get_by_role("checkbox", name="Shipper Jurisdiction Coun...")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Country of Export")
        expect(self.page.get_by_role("checkbox", name="Country of Export")).to_be_visible()
        self.page.get_by_text("consignee details").click()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Port of Unlading")
        expect(self.page.get_by_role("checkbox", name="Port of Unlading")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("IEC Number")
        expect(self.page.get_by_role("checkbox", name="IEC Number")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Consignee Name")
        expect(self.page.get_by_role("checkbox", name="Consignee Name")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Consignee Jurisdiction Co...")
        expect(self.page.get_by_role("checkbox", name="Consignee Jurisdiction Co...")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Consignee State")
        expect(self.page.get_by_role("checkbox", name="Consignee State")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Consignee City")
        expect(self.page.get_by_role("checkbox", name="Consignee City")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Consignee Pincode")
        expect(self.page.get_by_role("checkbox", name="Consignee Pincode")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Country of Import")
        expect(self.page.get_by_role("checkbox", name="Country of Import")).to_be_visible()
        self.page.get_by_text("custom details").click()
        expect(self.page.locator("label")).to_contain_text("Custom House Agent Name")
        self.page.get_by_text("cargo details").click()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Quantity")
        expect(self.page.get_by_role("checkbox", name="Quantity")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("CIF Value (USD)")
        expect(self.page.get_by_role("checkbox", name="CIF Value (USD)")).to_be_visible()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("Per Unit Value (USD)")
        expect(self.page.get_by_role("checkbox", name="Per Unit Value (USD)")).to_be_visible()
        #Verify presence of Close (X) icon next to each selected column
        self.page.locator("#custom-shipment-grid-modal div").filter(has_text=re.compile(r"^HS Code$")).get_by_role(
            "img").nth(1).click()
        expect(self.page.locator("#custom-shipment-grid-modal div").filter(has_text=re.compile(r"^HS Code$")).get_by_role(
            "img").nth(1)).not_to_be_visible()

    def Validate_Uncheck_check_checkbox(self):
        #Check/uncheck a few columns to view their visibility in the shipment grid
        self.page.get_by_text("primary details").click()
        selected_labels = []
        checkboxes = self.page.locator("label")

        count = checkboxes.count()
        for i in range(count):
            label = checkboxes.nth(i)
            label_text = label.inner_text().strip()
            input_box = label.locator("input[type='checkbox']")

            # ✅ Check the checkbox if it's not checked and enabled
            if input_box.is_visible() and input_box.is_enabled():
                if not input_box.is_checked():
                    input_box.check()
                    print(f"☑️ Checked checkbox: {label_text}")
                else:
                    print(f"✅ Already checked: {label_text}")

            # Collect all checked labels
            if input_box.is_checked():
                selected_labels.append(label_text)

        print("✅ Selected columns in modal (after ensuring all checked):", selected_labels)

        # Save configuration
        self.page.get_by_role("button", name="Save Configuration").click()
        self.page.wait_for_timeout(3000)

        # --- Step 2: Get all table header columns ---
        headers = self.page.locator("table thead tr th")
        header_texts = [headers.nth(i).inner_text().strip() for i in range(headers.count())]
        print("✅ Columns displayed in table:", header_texts)

        # --- Step 3: Validate each selected column is in table headers (case-insensitive) ---
        header_texts_lower = [h.lower() for h in header_texts]

        for col in selected_labels:
            assert col.lower() in header_texts_lower, (
                f"❌ Column '{col}' is selected but not found in the table!"
            )
            print(f"✅ Column '{col}' is correctly displayed in the table (case-insensitive check)")

        #Uncheck checkbox
        self.page.get_by_role("button", name="dashboard_customize Customise").click()
        unselected_labels = []
        checkboxes = self.page.locator("label")

        count = checkboxes.count()
        for i in range(count):
            label = checkboxes.nth(i)
            label_text = label.inner_text().strip()
            input_box = label.locator("input[type='checkbox']")

            # ✅ Check the checkbox if it's not checked and enabled
            if input_box.is_visible() and input_box.is_enabled():
                if input_box.is_checked():
                    input_box.check()
                    print(f"☑️ UnChecked checkbox: {label_text}")
                else:
                    print(f"✅ Already unchecked: {label_text}")

            # Collect all checked labels
            if input_box.is_checked():
                unselected_labels.append(label_text)

        print("✅ unselected columns in modal (after ensuring all unchecked):", selected_labels)

        # Save configuration
        self.page.get_by_role("button", name="Save Configuration").click(timeout=60000)

        # --- Step 2: Get all table header columns ---
        headers = self.page.locator("table thead tr th")
        header_texts = [headers.nth(i).inner_text().strip() for i in range(headers.count())]
        print("✅ Columns displayed in table:", header_texts)

        # --- Step 3: Validate each selected column is in table headers ---
        for col in unselected_labels:
            assert col not in header_texts, f"❌ Column '{col}' is unselected but still found in the table!"
            print(f"✅ Column '{col}' is correctly NOT displayed in the table")

        #Drag and drop columns in Selected Columns section to rearrange order
    def Validate_Drag_and_Drop(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)

        # In the shipment grid, click on Customise Shipment Grid button
        self.page.get_by_role("button", name="dashboard_customize Customise").click()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text(
            "Select columns for your grid.*You can select and display maximum of 15 columns."
        )

        # --- Step 1: Locate source and target elements for drag & drop ---
        source = self.page.locator('[class="selected-items"]').nth(1)  # Example: 2nd item
        target = self.page.locator('[class="selected-items"]').nth(2)  # Example: 3rd item

        # Get the text for reference
        source_text = source.inner_text().strip()
        target_text = target.inner_text().strip()

        # --- Step 2: Perform drag and drop ---
        source.drag_to(target)
        print(f"✅ Dragged '{source_text}' and dropped after '{target_text}'")

        # --- Step 3: Capture new order from modal ---
        items = self.page.locator('[class="selected-items"]')
        new_order = [items.nth(i).inner_text().strip() for i in range(items.count())]
        print("📌 New order in modal:", new_order)

        # --- Step 4: Save configuration and close modal ---
        self.page.get_by_role("button", name="Save Configuration").click(timeout=60000)
        self.page.wait_for_timeout(3000)

        # --- Step 5: Capture table headers after modal is closed ---
        headers = self.page.locator("table thead tr th")
        table_order = [headers.nth(i).inner_text().strip() for i in range(headers.count())]
        print("📌 Columns in table:", table_order)

        # --- Step 6: Validate table order matches modal order (ignore case + spaces, offset +3) ---
        normalize = lambda text: "".join(text.split()).lower()

        for i, col in enumerate(new_order):
            actual = table_order[i + 3]  # ✅ offset by 3
            expected = col

            assert normalize(actual) == normalize(expected), (
                f"❌ Column order mismatch at position {i + 1}: "
                f"Expected '{expected}', Found '{actual}'"
            )
            print(f"✅ Column '{expected}' is correctly placed at position {i + 1}")

    def Validate_Reset_default_view(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        # In the shipment grid, click on Customise Shipment Grid button
        self.page.get_by_role("button", name="dashboard_customize Customise").click()
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text(
            "Select columns for your grid.*You can select and display maximum of 15 columns.")
        # Verify the available section categories: Primary Details, Shipper Details, Consignee Details, custom details and Cargo Details
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("primary details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("shipper details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("consignee details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("custom details")
        expect(self.page.locator("#custom-shipment-grid-modal")).to_contain_text("cargo details")
        #Click Reset Default View
        self.page.get_by_role("button", name="Reset Default View").click()
        expect(self.page.get_by_text("user configuration set")).to_be_visible()
        self.page.wait_for_timeout(3000)
        #Validate all columns
        expected_columns = [
            "S No",
            "Bill of Entry Date",
            "HS Code",
            "Product Description",
            "Mode of Transportation",
            "Country of Export",
            "Shipper Name",
            "Consignee Name",
            "Port of Lading",
            "Port of Unlading",
            "Mode of Transportation",
            "Quantity",
            "CIF Value (USD)",
            "Per Unit Value (USD)"
        ]

        # Locate all column headers in the table
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        # Collect all header texts
        header_texts = [headers.nth(i).inner_text().strip() for i in range(header_count)]
        print("Found columns:", header_texts)

        # Validate each expected column is present
        for col in expected_columns:
            assert col in header_texts, f"❌ Column '{col}' not found in table headers"
            print(f"✅ Column '{col}' is present in the table")

    def Validate_HSCode_Hyperlink(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        self.page.wait_for_timeout(2000)

        hs_code_link = self.page.locator("table tbody tr td:nth-child(5) a")
        count = hs_code_link.count()
        print(f"🔗 Found {count} HS Code links")

        for i in range(count):
            # --- Get HS Code text from the table before clicking ---
            code_text = hs_code_link.nth(i).inner_text().strip()
            print(f"➡ Clicking HS Code: {code_text}")

            # --- Step 2: Click the link and capture the new tab ---
            with self.page.context.expect_page() as new_page_info:
                hs_code_link.nth(i).click()
            new_page = new_page_info.value

            # --- Step 3: Wait for new page to load ---
            new_page.wait_for_load_state("domcontentloaded")

            # --- Step 4: Validate HS Code on the new page ---
            page_content = new_page.locator(".ml-3 > h5").inner_text()
            assert code_text in page_content, f"❌ HS Code {code_text} not found on the new page"
            print(f"✅ HS Code {code_text} is present on the new page")

            # --- Close new tab and return to main page ---
            new_page.close()

    def Validate_Shippername_Hyperlink(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        self.page.wait_for_timeout(2000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        for i in range(row_count):

            # Get company name from data-tip or fallback to visible text
            shipper_name_link = self.page.locator("table tbody tr td:nth-child(8) a")
            Shipper_name = rows.nth(i).locator("td").nth(7).locator("[data-tip]").first
            Shipper_tip = Shipper_name.get_attribute("data-tip") if Shipper_name.count() > 0 else None
            if Shipper_tip and Shipper_tip.strip():
                shipper_text = Shipper_tip.strip()
            else:
                shipper_text = rows.nth(i).locator("td").nth(7).inner_text().strip()

            print(f"➡ Clicking shipper name: {shipper_text}")

            # --- Step 2: Click the link and capture the new tab ---
            with self.page.context.expect_page() as new_page_info:
                shipper_name_link.nth(i).click()
            new_page = new_page_info.value

            # --- Step 3: Wait for new page to load ---
            new_page.wait_for_load_state("domcontentloaded")

            # --- Step 4: Validate Shipper name on the new page (case-insensitive) ---
            page_content = new_page.locator("[class='_portName_j7u5w_19']").inner_text().strip()
            print(f"Shipper name on new page: {page_content}")

            assert shipper_text.lower() in page_content.lower(), (
                f"❌ Shipper name '{shipper_text}' not found on the new page (Got: {page_content})"
            )
            print(f"✅ Shipper name '{shipper_text}' is present on the new page")

            # --- Close new tab and return to main page ---
            new_page.close()

    def Validate_Consigneename_Hyperlink(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        self.page.wait_for_timeout(2000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        for i in range(row_count):

            # Get company name from data-tip or fallback to visible text
            consignee_name_link = self.page.locator("table tbody tr td:nth-child(9) a")
            consignee_name = rows.nth(i).locator("td").nth(8).locator("[data-tip]").first
            consignee_tip = consignee_name.get_attribute("data-tip") if consignee_name.count() > 0 else None
            if consignee_tip and consignee_tip.strip():
                consignee_text = consignee_tip.strip()
            else:
                consignee_text = rows.nth(i).locator("td").nth(8).inner_text().strip()

            print(f"➡ Clicking consignee name: {consignee_text}")

            # --- Step 2: Click the link and capture the new tab ---
            with self.page.context.expect_page() as new_page_info:
                consignee_name_link.nth(i).click()
            new_page = new_page_info.value

            # --- Step 3: Wait for new page to load ---
            new_page.wait_for_load_state("domcontentloaded")

            # --- Step 4: Validate Shipper name on the new page (case-insensitive) ---
            page_content = new_page.locator("[class='_portName_j7u5w_19']").inner_text().strip()
            print(f"consignee name on new page: {page_content}")

            assert consignee_text.lower() in page_content.lower(), (
                f"❌ consignee name '{consignee_text}' not found on the new page (Got: {page_content})"
            )
            print(f"✅ consignee name '{consignee_text}' is present on the new page")
            # Verify the presence of the following tabs
            expect(new_page.locator("//div[normalize-space()='Trade as Buyer']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='Trade as Supplier']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='Logistics Partners']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='Corporate Hierarchy']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='Due Diligence Reports']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='Contact Details']")).to_be_visible()

            # Click View Shipments for any country
            new_page.locator("//button[normalize-space()='View Shipments']").click()
            expect(new_page.locator("//span[normalize-space()='Universal Search']")).to_be_visible()
            # --- Close new tab and return to main page ---
            new_page.close()

    def Validate_Port_Laiding_Hyperlink(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        self.page.wait_for_timeout(2000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        for i in range(row_count):

            # Get port laiding name from data-tip or fallback to visible text
            port_laiding_name_link = self.page.locator("table tbody tr td:nth-child(10) a")
            port_laiding_name = rows.nth(i).locator("td").nth(9).locator("[data-tip]").first
            port_laiding_tip =  port_laiding_name.get_attribute("data-tip") if  port_laiding_name.count() > 0 else None
            if port_laiding_tip and port_laiding_tip.strip():
                port_laiding = port_laiding_tip.strip()
            else:
                port_laiding = rows.nth(i).locator("td").nth(9).inner_text().strip()
            port_laiding_text = port_laiding.split(",")[0].strip()

            print(f"➡ Clicking port laiding name: {port_laiding_text}")

            # --- Step 2: Click the link and capture the new tab ---
            with self.page.context.expect_page() as new_page_info:
                port_laiding_name_link.nth(i).click()
            new_page = new_page_info.value

            # --- Step 3: Wait for new page to load ---
            new_page.wait_for_load_state("domcontentloaded")

            # --- Step 4: Validate Shipper name on the new page (case-insensitive) ---
            page_content = new_page.locator("[class='_portName_j7u5w_19']").inner_text().strip()
            print(f"consignee name on new page: {page_content}")

            assert port_laiding_text.lower() in page_content.lower(), (
                f"❌ port laiding name '{port_laiding_text}' not found on the new page (Got: {page_content})"
            )
            print(f"✅ port laiding name '{port_laiding_text}' is present on the new page")
            # Verify the presence of the following tabs
            expect(new_page.locator("//div[normalize-space()='As Port of Lading']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='As Port of Unlading']")).to_be_visible()
            # --- Close new tab and return to main page ---
            new_page.close()

    def Validate_Port_Unlaiding_Hyperlink(self):
        self.page.locator("//a[@id='nav-home-tab']").click()
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
        self.page.wait_for_timeout(2000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        for i in range(row_count):

            # Get port laiding name from data-tip or fallback to visible text
            port_unlaiding_name_link = self.page.locator("table tbody tr td:nth-child(11) a")
            port_unlaiding_name = rows.nth(i).locator("td").nth(10).locator("[data-tip]").first
            port_unlaiding_tip =  port_unlaiding_name.get_attribute("data-tip") if  port_unlaiding_name.count() > 0 else None
            if port_unlaiding_tip and port_unlaiding_tip.strip():
                port_unlaiding = port_unlaiding_tip.strip()
            else:
                port_unlaiding = rows.nth(i).locator("td").nth(10).inner_text().strip()
            port_unlaiding_text = port_unlaiding.split(",")[0].strip()

            print(f"➡ Clicking port unlaiding name: {port_unlaiding_text}")

            # --- Step 2: Click the link and capture the new tab ---
            with self.page.context.expect_page() as new_page_info:
                port_unlaiding_name_link.nth(i).click()
            new_page = new_page_info.value

            # --- Step 3: Wait for new page to load ---
            new_page.wait_for_load_state("domcontentloaded")

            # --- Step 4: Validate Shipper name on the new page (case-insensitive) ---
            page_content = new_page.locator("[class='_portName_j7u5w_19']").inner_text().strip()
            print(f"port unlaiding on new page: {page_content}")

            assert port_unlaiding_text.lower() in page_content.lower(), (
                f"❌ port unlaiding name '{port_unlaiding_text}' not found on the new page (Got: {page_content})"
            )
            print(f"✅ port unlaiding name '{port_unlaiding_text}' is present on the new page")
            # Verify the presence of the following tabs
            expect(new_page.locator("//div[normalize-space()='As Port of Lading']")).to_be_visible()
            expect(new_page.locator("//div[normalize-space()='As Port of Unlading']")).to_be_visible()
            # --- Close new tab and return to main page ---
            new_page.close()

    def _validate_date_format(self, date_text: str) -> bool:
        """Check if date is in format MMM DD, YYYY (e.g., Jul 31, 2025)."""
        pattern = r"^[A-Z][a-z]{2} \d{1,2}, \d{4}$"
        if re.match(pattern, date_text):
            try:
                datetime.strptime(date_text, "%b %d, %Y")
                return True
            except ValueError:
                return False
        return False

    def _save_results_to_excel(self):
        """Save invalid dates and mismatches into Excel file."""
        if not self.invalid_date_entries and not self.date_mismatches:
            print("📘 No invalid Bill Entry Dates found.")
            return

        os.makedirs("results", exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invalid Bill Entry Date"

        # Headers
        sheet.append(["Sl. No", "Bill Entry (Grid)", "Bill Entry (View)"])

        # Red fill style
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Write invalid date entries
        for item in self.invalid_date_entries:
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1, value=item["slNo"])
            cell = sheet.cell(row=row, column=2, value=item["Bill Entry Date"])
            cell.fill = red_fill

        # Write mismatches
        for item in self.date_mismatches:
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1, value=item["slNo"])
            grid_cell = sheet.cell(row=row, column=2, value=item["Bill Entry Grid"])
            view_cell = sheet.cell(row=row, column=3, value=item["Bill Entry View"])
            grid_cell.fill = red_fill
            view_cell.fill = red_fill

        # Save file
        workbook.save("results/invalid_bill_entry_dates.xlsx")
        print("📁 Results saved to 'results/invalid_bill_entry_dates.xlsx'")

    def Validate_Bill_Entry_Data_Column(self, first_call=True):
        """Validate Bill Entry Date column against format and View details."""

        # Step 1: Navigate & get headers (only on first call)
        if first_call:
            self.page.locator("//a[@id='nav-home-tab']").click()
            expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=6000)
            expect(self.page.get_by_text("Bill of Entry Date")).to_be_visible(timeout=6000)

            # Find 'Bill of Entry Date' column index
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()
            self.bill_entry_index = None

            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "bill of entry date":
                    self.bill_entry_index = i
                    break

            if self.bill_entry_index is None:
                raise Exception("❌ Bill of Entry Date column not found in the table headers")

            print(f"✅ Found 'Bill of Entry Date' column at index: {self.bill_entry_index}")

        # Step 2: Validate current page rows
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        for i in range(row_count):
            view = rows.nth(i).locator("td").nth(1)  # View button column
            bill_entry_date_grid = rows.nth(i).locator("td").nth(self.bill_entry_index).inner_text().strip()
            sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()

            # Verify that the date format is consistent across all rows
            if self._validate_date_format(bill_entry_date_grid):
                print(f"✅ Valid Date: {bill_entry_date_grid}")
            else:
                print(f"❌ Invalid Date: {bill_entry_date_grid}")
                self.invalid_date_entries.append(
                    {"slNo": sl_no, "Bill Entry Date": bill_entry_date_grid}
                )

            # Click on View Details for any shipment
            view.click()
            self.page.locator('[class="tw-capitalize "]').first.wait_for(state="visible", timeout=60000)
            # Verify the view details page
            expect(self.page.get_by_role("dialog")).to_contain_text("Detailed Shipment")
            expect(self.page.get_by_role("dialog")).to_contain_text("Shipment Origin Details")
            expect(self.page.get_by_role("dialog")).to_contain_text("Shipment Destination Details")
            expect(self.page.get_by_role("dialog")).to_contain_text("Customs Details")
            expect(self.page.get_by_role("dialog")).to_contain_text("Cargo Details")

            bill_entry_date_view = self.page.locator('[class="tw-capitalize "]').first.inner_text().strip()
            #Compare the Bill of Entry Date in the detail view with the grid
            if bill_entry_date_view == bill_entry_date_grid:
                print(f"✅ Matched Date in View: {bill_entry_date_view} (Sl. No: {sl_no})")
            else:
                print(
                    f"❌ Mismatch - Grid: {bill_entry_date_grid}, View: {bill_entry_date_view} (Sl. No: {sl_no})"
                )
                self.date_mismatches.append(
                    {
                        "slNo": sl_no,
                        "Bill Entry Grid": bill_entry_date_grid,
                        "Bill Entry View": bill_entry_date_view,
                    }
                )

            self.page.get_by_role("button", name="Close").click(timeout=60000)

        #Step 4: Handle pagination (recursive call)
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
        #         state="visible", timeout=100000
        #     )
        #     # recursive call (not first_call anymore)
        #     self.Validate_Bill_Entry_Data_Column(first_call=False)
        # else:
        #     print("✅ All pages validated")
        #     self._save_results_to_excel()

    def validate_shipment_fields_Shipment_Origin_Details(self, labels_to_check, first_call=True):
        """
        Validate shipment fields across Grid and View Details screen.
        Invalid data is stored in Excel after traversing all pages.
        """
        expected_fields = [
            "Country of Origin",
            "Port of Lading",
            "Shipper Name",
            "Shipper Type",
            "Country of Export",
            "Shipper Jurisdiction Country",
        ]

        # --- Open shipment grid / column selection panel ---
        self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
        self.page.wait_for_timeout(1000)

        # --- Example: Click on group ---
        self.page.locator("//span[normalize-space()='shipper details']").click()

        # --- Normalize for comparison ---
        labels_to_check_lower = [lbl.lower() for lbl in expected_fields]

        checkboxes = self.page.locator("label")
        for i in range(checkboxes.count()):
            label = checkboxes.nth(i)
            label_text = label.inner_text().strip()
            input_box = label.locator("input[type='checkbox']")

            if label_text.lower() in labels_to_check_lower:
                if input_box.is_visible() and input_box.is_enabled():
                    if not input_box.is_checked():
                        input_box.check()
                        print(f"☑️ Checked checkbox: {label_text}")
                    else:
                        print(f"✅ Already checked: {label_text}")
            else:
                print(f"⏩ Skipped: {label_text}")

        # --- Save configuration ---
        self.page.get_by_role("button", name="Save Configuration").click(timeout=60000)
        self.page.wait_for_timeout(2000)

        # --- Re-fetch headers ---
        headers = self.page.locator("table thead tr th")
        self.header_map = {}
        for field in expected_fields:
            for i in range(headers.count()):
                header_text = headers.nth(i).inner_text().strip()
                if header_text.lower() == field.lower():
                    self.header_map[field] = i
                    break

        print("✅ Header Map:", self.header_map)

        # --- Validate rows ---
        rows = self.page.locator("table tbody tr")
        for i in range(rows.count()):
            sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()
            grid_data = {
                field: rows.nth(i).locator("td").nth(idx).inner_text().strip()
                for field, idx in self.header_map.items()
            }

            # --- Open View Details ---
            rows.nth(i).locator("td").nth(1).click()
            self.page.locator('[class="tw-capitalize "]').first.wait_for(
                state="visible", timeout=60000
            )

            # --- Field validation loop ---
            detail_locators = self.page.locator('[class="col-7 bold-data text-capitalize d-flex"]')
            link_locators = self.page.locator('[class="col-7  text-capitalize d-flex"]')

            checks = {
                "Country of Origin": detail_locators.nth(0),
                "Port of Lading": link_locators.nth(0),
                "Shipper Name": detail_locators.nth(1),
                "Shipper Type": detail_locators.nth(2),
                "Country of Export": detail_locators.nth(3),
                "Shipper Jurisdiction Country": detail_locators.nth(5),
            }

            for field, expected_value in grid_data.items():
                if not expected_value:
                    print(f"⚠️ {field} not found in grid_data")
                    continue

                locator = checks.get(field)
                if locator is None:
                    print(f"⚠️ No locator mapped for {field}")
                    continue

                actual_text = locator.inner_text().strip()

                # --- Soft + Fuzzy Validation ---
                similarity = difflib.SequenceMatcher(None, expected_value.lower(), actual_text.lower()).ratio()

                if (
                        expected_value.lower() == actual_text.lower()
                        or expected_value.lower() in actual_text.lower()
                        or actual_text.lower() in expected_value.lower()
                        or similarity > 0.9  # 90% similarity threshold
                ):
                    print(f"✅ {field} validated: '{expected_value}' ~ '{actual_text}' at sl.No: {sl_no}")
                else:
                    raise AssertionError(
                        f"❌ {field} mismatch: Expected '{expected_value}' but found '{actual_text}' (similarity={similarity:.2f})"
                    )

            # --- Close dialog ---
            self.page.locator("//span[@aria-hidden='true']").click()

    def validate_shipment_fields_Shipment_Destination_Details(self, labels_to_check, first_call=True):
        """
        Validate shipment fields across Grid and View Details screen.
        Invalid data is stored in Excel after traversing all pages.
        """
        expected_fields = [
            "Port of Unlading",
            "Consignee Name",
            "Country of Import",
            "Consignee State",
            "Consignee City",
            "Consignee Pincode",
            "Consignee Jurisdiction Country"
        ]

        # --- Open shipment grid / column selection panel ---
        self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click(timeout=10000)
        self.page.wait_for_timeout(1000)

        # --- Example: Click on group ---
        self.page.get_by_text("consignee details").click()
        self.page.locator("label", has_text="Consignee Name").wait_for(timeout=10000)
        self.page.pause()
        self.page.locator("#custom-shipment-grid-modal div").filter(has_text=re.compile(r"^HS Code$")).get_by_role(
            "img").nth(1).click()
        self.page.locator("#custom-shipment-grid-modal div").filter(
            has_text=re.compile(r"^Product Description$")).get_by_role("img").nth(1).click()

        # --- Normalize for comparison ---
        labels_to_check_lower = [lbl.lower() for lbl in expected_fields]

        checkboxes = self.page.locator("label")
        count = checkboxes.count()
        for i in range(count):
            label = checkboxes.nth(i)
            label_text = label.inner_text().strip()
            input_box = label.locator("input[type='checkbox']")

            # Allow partial matching (important for truncated text like "Consignee Jurisdiction Co...")
            if any(label_text.lower().startswith(exp[:20].lower()) or exp.lower().startswith(label_text.lower())
                   for exp in labels_to_check_lower):
                if input_box.is_visible() and input_box.is_enabled():
                    if not input_box.is_checked():
                        input_box.check()
                        print(f"☑️ Checked checkbox: {label_text}")
                    else:
                        print(f"✅ Already checked: {label_text}")
            else:
                print(f"⏩ Skipped: {label_text}")

        # --- Save configuration ---
        self.page.get_by_role("button", name="Save Configuration").click(timeout=60000)
        self.page.wait_for_timeout(2000)

        # --- fetch headers ---
        headers = self.page.locator("table thead tr th")
        self.header_map = {}
        for field in expected_fields:
            for i in range(headers.count()):
                header_text = headers.nth(i).inner_text().strip()
                if header_text.lower() == field.lower():
                    self.header_map[field] = i
                    break

        print("✅ Header Map:", self.header_map)

        # --- Validate rows ---
        rows = self.page.locator("table tbody tr")
        for i in range(rows.count()):
            sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()
            grid_data = {
                field: rows.nth(i).locator("td").nth(idx).inner_text().strip()
                for field, idx in self.header_map.items()
            }

            # --- Open View Details ---
            rows.nth(i).locator("td").nth(1).click()
            self.page.locator('[class="tw-capitalize "]').first.wait_for(
                state="visible", timeout=60000
            )

            # --- Field validation loop ---
            detail_locators = self.page.locator('[class="col-7 bold-data text-capitalize d-flex"]')
            link_locators = self.page.locator('[class="col-7  text-capitalize d-flex"]')

            checks = {
                "Port of Unlading": link_locators.nth(2),
                "Consignee Name": detail_locators.nth(6),
                "Country of Import": detail_locators.nth(8),
                "Consignee State": detail_locators.nth(10),
                "Consignee City": detail_locators.nth(11),
                "Consignee Pincode": detail_locators.nth(12),
                "Consignee Jurisdiction Country": detail_locators.last,
            }

            for field, expected_value in grid_data.items():
                if not expected_value:
                    print(f"⚠️ {field} not found in grid_data")
                    continue

                locator = checks.get(field)
                if locator is None:
                    print(f"⚠️ No locator mapped for {field}")
                    continue

                actual_text = locator.inner_text().strip()

                # --- Soft + Fuzzy Validation ---
                similarity = difflib.SequenceMatcher(None, expected_value.lower(), actual_text.lower()).ratio()

                if (
                        expected_value.lower() == actual_text.lower()
                        or expected_value.lower() in actual_text.lower()
                        or actual_text.lower() in expected_value.lower()
                        or similarity > 0.9  # 90% similarity threshold
                ):
                    print(f"✅ {field} validated: '{expected_value}' ~ '{actual_text}' at sl.No: {sl_no}")
                else:
                    raise AssertionError(
                        f"❌ {field} mismatch: Expected '{expected_value}' but found '{actual_text}' (similarity={similarity:.2f})"
                    )

            # --- Close dialog ---
            self.page.locator("//span[@aria-hidden='true']").click()

    def validate_shipment_fields_Cargo_Details(self, labels_to_check, first_call=True):
        """
        Validate shipment fields across Grid and View Details screen.
        Invalid data is stored in Excel after traversing all pages.
        """
        expected_fields = [
            "Quantity",
            "CIF Value (USD)",
            "Per Unit Value (USD)"
        ]

        # --- Open shipment grid / column selection panel ---
        self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click(timeout=10000)
        self.page.wait_for_timeout(1000)

        # --- Example: Click on group ---
        self.page.get_by_text("cargo details").click()
        self.page.locator("label", has_text="Quantity").wait_for(timeout=10000)
        # --- Normalize for comparison ---
        labels_to_check_lower = [lbl.lower() for lbl in expected_fields]

        checkboxes = self.page.locator("label")
        count = checkboxes.count()
        for i in range(count):
            label = checkboxes.nth(i)
            label_text = label.inner_text().strip()
            input_box = label.locator("input[type='checkbox']")

            # Allow partial matching (important for truncated text like "Consignee Jurisdiction Co...")
            if any(label_text.lower().startswith(exp[:20].lower()) or exp.lower().startswith(label_text.lower())
                   for exp in labels_to_check_lower):
                if input_box.is_visible() and input_box.is_enabled():
                    if not input_box.is_checked():
                        input_box.check()
                        print(f"☑️ Checked checkbox: {label_text}")
                    else:
                        print(f"✅ Already checked: {label_text}")
            else:
                print(f"⏩ Skipped: {label_text}")

        # --- Save configuration ---
        self.page.get_by_role("button", name="Save Configuration").click(timeout=60000)
        self.page.wait_for_timeout(2000)

        # --- fetch headers ---
        headers = self.page.locator("table thead tr th")
        self.header_map = {}
        for field in expected_fields:
            for i in range(headers.count()):
                header_text = headers.nth(i).inner_text().strip()
                if header_text.lower() == field.lower():
                    self.header_map[field] = i
                    break

        print("✅ Header Map:", self.header_map)

        # --- Validate rows ---
        rows = self.page.locator("table tbody tr")
        for i in range(rows.count()):
            sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()
            grid_data = {
                field: rows.nth(i).locator("td").nth(idx).inner_text().strip()
                for field, idx in self.header_map.items()
            }

            # --- Open View Details ---
            rows.nth(i).locator("td").nth(1).click()
            self.page.locator('[class="tw-capitalize "]').first.wait_for(
                state="visible", timeout=60000
            )

            # --- Field validation loop ---
            detail_locators = self.page.locator('[class="col-7 bold-data text-capitalize "]')

            checks = {
                "Quantity": detail_locators.nth(3),
                "CIF Value (USD)": detail_locators.nth(4),
                "Per Unit Value (USD)": detail_locators.nth(5)

            }

            for field, expected_value in grid_data.items():
                if not expected_value:
                    print(f"⚠️ {field} not found in grid_data")
                    continue

                locator = checks.get(field)
                if locator is None:
                    print(f"⚠️ No locator mapped for {field}")
                    continue

                actual_text = locator.inner_text().strip()

                # --- Soft + Fuzzy Validation ---
                similarity = difflib.SequenceMatcher(None, expected_value.lower(), actual_text.lower()).ratio()

                if (
                        expected_value.lower() == actual_text.lower()
                        or expected_value.lower() in actual_text.lower()
                        or actual_text.lower() in expected_value.lower()
                        or similarity > 0.9  # 90% similarity threshold
                ):
                    print(f"✅ {field} validated : '{expected_value}' ~ '{actual_text}' at sl.No: {sl_no}")
                else:
                    raise AssertionError(
                        f"❌ {field} mismatch: Expected '{expected_value}' but found '{actual_text}' (similarity={similarity:.2f})"
                    )

            # --- Close dialog ---
            self.page.locator("//span[@aria-hidden='true']").click()

    def validate_shipment_fields_HS_Product_Details(self, labels_to_check, first_call=True):
        """
        Validate shipment fields across Grid and View Details screen.
        Invalid data is stored in Excel after traversing all pages.
        """
        expected_fields = [
            "HS Code",
            "Product Description",
        ]

        # --- Open shipment grid / column selection panel ---
        self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click(timeout=10000)
        self.page.wait_for_timeout(1000)

        # --- Example: Click on group ---
        self.page.locator("label", has_text="HS Code").wait_for(timeout=10000)
        # --- Normalize for comparison ---
        labels_to_check_lower = [lbl.lower() for lbl in expected_fields]

        checkboxes = self.page.locator("label")
        count = checkboxes.count()
        for i in range(count):
            label = checkboxes.nth(i)
            label_text = label.inner_text().strip()
            input_box = label.locator("input[type='checkbox']")

            # Allow partial matching (important for truncated text like "Consignee Jurisdiction Co...")
            if any(label_text.lower().startswith(exp[:20].lower()) or exp.lower().startswith(label_text.lower())
                   for exp in labels_to_check_lower):
                if input_box.is_visible() and input_box.is_enabled():
                    if not input_box.is_checked():
                        input_box.check()
                        print(f"☑️ Checked checkbox: {label_text}")
                    else:
                        print(f"✅ Already checked: {label_text}")
            else:
                print(f"⏩ Skipped: {label_text}")

        # --- Save configuration ---
        self.page.get_by_role("button", name="Save Configuration").click(timeout=60000)
        self.page.wait_for_timeout(2000)

        # --- fetch headers ---
        headers = self.page.locator("table thead tr th")
        self.header_map = {}
        for field in expected_fields:
            for i in range(headers.count()):
                header_text = headers.nth(i).inner_text().strip()
                if header_text.lower() == field.lower():
                    self.header_map[field] = i
                    break

        print("✅ Header Map:", self.header_map)

        # --- Validate rows ---
        rows = self.page.locator("table tbody tr")
        for i in range(rows.count()):
            sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()
            grid_data = {}

            # --- Collect grid values for all mapped headers ---
            for field, idx in self.header_map.items():
                td_cell = rows.nth(i).locator("td").nth(idx)

                # If `data-tip` exists, prefer that over plain text
                data_tip_el = td_cell.locator("[data-tip]").first
                if data_tip_el.count() > 0:
                    field_tip = data_tip_el.get_attribute("data-tip")
                    if field_tip and field_tip.strip():
                        grid_data[field] = field_tip.strip()
                    else:
                        grid_data[field] = td_cell.inner_text().strip()
                else:
                    grid_data[field] = td_cell.inner_text().strip()

            # --- Open View Details ---
            rows.nth(i).locator("td").nth(1).click()
            self.page.locator('[class="tw-capitalize "]').first.wait_for(
                state="visible", timeout=60000
            )

            # --- Field locators in detail modal ---
            detail_locators = self.page.locator('[class="col-7 bold-data text-capitalize "]')
            checks = {
                "Product Description": detail_locators.nth(2),
                "HS Code": detail_locators.nth(1),
                # add more mappings if needed
            }

            # --- Validation loop ---
            for field, expected_value in grid_data.items():
                if not expected_value:
                    print(f"⚠️ {field} not found in grid_data")
                    continue

                locator = checks.get(field)
                if locator is None:
                    print(f"⚠️ No locator mapped for {field}")
                    continue

                actual_text = locator.inner_text().strip()

                # --- Soft + Fuzzy Validation ---
                similarity = difflib.SequenceMatcher(
                    None, expected_value.lower(), actual_text.lower()
                ).ratio()

                if (
                        expected_value.lower() == actual_text.lower()
                        or expected_value.lower() in actual_text.lower()
                        or actual_text.lower() in expected_value.lower()
                        or similarity > 0.9  # 90% similarity threshold
                ):
                    print(
                        f"✅ {field} validated : '{expected_value}' ~ '{actual_text}' at sl.No: {sl_no}"
                    )
                else:
                    print(
                        f"❌ {field} mismatch: Expected '{expected_value}' but found '{actual_text}' (similarity={similarity:.2f})"
                    )

            # --- Close dialog ---
            self.page.locator("//span[@aria-hidden='true']").click()







