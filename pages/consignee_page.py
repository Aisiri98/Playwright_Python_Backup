import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page
from playwright.sync_api import expect

invalid_entries =[]
import difflib
from decimal import Decimal, ROUND_HALF_UP
def format_number(num_str: str) -> str:
    """Convert number string like '149,760' → '149.76K', '268,000' → '268K'.
       Uses ROUND_HALF_UP and trims unnecessary trailing zeros."""
    num_str = num_str.replace(",", "").strip()

    try:
        num = int(num_str)
    except ValueError:
        raise ValueError(f"Invalid number string: {num_str}")

    def format_decimal(val: Decimal, suffix: str) -> str:
        # Round to 2 decimals
        val = val.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # Convert to string and strip trailing zeros/decimal if not needed
        val_str = f"{val:.2f}".rstrip("0").rstrip(".")
        return f"{val_str}{suffix}"

    if num >= 1_000_000:
        return format_decimal(Decimal(num) / Decimal(1_000_000), "M")
    elif num >= 1_000:
        return format_decimal(Decimal(num) / Decimal(1_000), "K")
    else:
        return str(num)
class ConsigneeFilter:
    def __init__(self, page: Page):
        self.C_O_field = None
        self.C_E_field = None
        self.S_JC_field = None
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def normalize_company_name(self, name: str) -> str:
        """
        Normalize company names by shortening common terms.
        Example: "Limited" → "Ltd", "Company" → "Co".
        """
        # Lowercase and strip spaces
        name = name.lower().strip()

        # Replace common full forms with abbreviations
        replacements = {
            r"\blimited\b": "ltd",
        }

        for pattern, replacement in replacements.items():
            name = re.sub(pattern, replacement, name)

        # Collapse multiple spaces
        name = re.sub(r"\s+", " ", name)

        return name

    def check_consignee_filter(self):
        # Wait for the Consignee Standardized Name label
        consignee_standardized_locator = self.page.locator(
            "//div[contains(text(),'Consignee Standardized Name')]"
        )
        expect(consignee_standardized_locator).to_be_visible(timeout=100000)

        # Locate Consignee filter options
        elements = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count = elements.count()
        print(f"Found {count} Consignee filter options (before toggle)")

        # Loop through each element
        for i in range(count):
            element = elements.nth(i)
            if element.is_visible():
                text = element.inner_text()
                print(f"✅ Consignee option {i + 1} is visible: {text}")
            else:
                print(f"❌ Consignee option {i + 1} is not visible")

        # Click on the Consignee filter again (collapse/close)
        self.page.locator("//span[normalize-space()='Consignee']").click()
        self.page.wait_for_timeout(2000)  # give time for collapse

        # Re-locate elements again after DOM change
        elements_after = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count_after = elements_after.count()
        print(f"Checking {count_after} Consignee filter options")

        # Expect no options should be visible
        for i in range(count_after):
            element = elements_after.nth(i)
            expect(element).not_to_be_visible(timeout=5000)

    def open_consignee_dropdown(self):
        consignee_standardized_locator = self.page.locator("//div[contains(text(),'Consignee Standardized Name')]")

        # Check if "consignee Standardized Name" is visible
        if not consignee_standardized_locator.is_visible():
            print("📌 consignee dropdown is closed. Clicking to open...")
            self.page.locator("//span[normalize-space()='Consignee']").click()
        else:
            print("✅ consignee dropdown is already open.")

    def Click_C_SName(self):
        self.page.locator("//div[contains(text(),'Consignee Standardized Name')]").click()
        self.page.wait_for_timeout(2000)
        expect(self.page.locator("//h5[normalize-space()='Filter by Consignee Standardized Name']")).to_be_visible()

    def Consignee_Filter_Search_Company(self, C_S_Name: str):
        # Use the search bar in the modal to search for a consignee (e.g., "Walmart")
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").last.fill(C_S_Name)
        self.page.wait_for_timeout(2000)
        # Check for the presence of the company name in the results
        locators = self.page.locator('[class="text-no-transform"]')
        count = locators.count()
        print(f"🔍 Found {count} elements with class 'text-no-transform'")

        if count == 0:
            print("❌ FAIL: No elements found for search")
            self.page.locator('input[type="text"]').last.clear()
        else:
            found_match = False

            for i in range(count):
                locator = locators.nth(i)

                if locator.is_visible():
                    actual_text = locator.inner_text().strip()

                    if C_S_Name.lower() in actual_text.lower():  # case-insensitive compare
                        print(f"✅ PASS: Search is working → '{C_S_Name}' found in element {i + 1}: '{actual_text}'")
                        expect(locator).to_contain_text(C_S_Name)
                        found_match = True
                    else:
                        print(f"❌ FAIL: Search text '{C_S_Name}' not found in element {i + 1}. Actual: '{actual_text}'")
                else:
                    print(f"⚠️ Element {i + 1} is not visible")

            if not found_match:
                print(f"❌ FAIL: Search text '{C_S_Name}' not found in any visible element")
                self.page.locator('input[type="text"]').clear()

    def Apply_Filter_Functionality(self):
        last_input = self.page.locator('input[type="text"]').last
        value = last_input.input_value().strip()

        if value == "":
            print("ℹ️  input box is already empty. Skipping clear()...")
        else:
            print(f"🧹 Clearing input box (current value: '{value}')")
            last_input.clear()
        self.page.wait_for_timeout(2000)
        # Click on the checkbox of the first result
        expect(self.page.locator('[class="text-no-transform"]').nth(1)).to_be_visible()
        self.Company_name = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Selected value is : {self.Company_name}")

        import re
        # Calculate total from all .grey-text elements
        total = 0

        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        shipper_record_clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        shipper_record_num = float(shipper_record_clean)
        if 0 <= shipper_record_num < 100_000:
            self.total_in_millions = f"{shipper_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{shipper_record_num / 1_000_000:.2f}M"

        print(self.total_in_millions)
        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        return False

    def Validate_ShipmentRecord_Count(self):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=10000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(3000)
        # Validate total shipment count in 'Shipment' tab
        # expect(self.page.locator(
        #     "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']")).to_contain_text(
        #     f"{self.total_in_millions}")
        record_count = self.page.locator(
            "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']"
        )

        # Get actual text
        actual_text = record_count.inner_text()

        # If–else validation
        if self.total_in_millions == actual_text:
            print(f"✅ Record count matches: {actual_text}")
        else:
            print(f"❌ Record count mismatch. Expected {self.total_in_millions}, Found {actual_text}")

    def Validate_Total_RecordCount_In_the_Table(self):
        total_record_text = self.page.locator('[class="trademo-table-count-text"]').inner_text()
        total_from_summary = total_record_text.split("of")[1].strip()
        print(f"🔢 Value after 'of': {total_from_summary}")
        print(f"total record count:{self.Total_Record_Count}")
        if total_from_summary == self.Total_Record_Count:
            print(f"✅ Record count matched: {total_from_summary}")
        else:
            print(f"❌ Record count mismatch! Summary: {total_from_summary}, Table: {self.Total_Record_Count}")

    def Validate_C_S_Name_label(self):
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')

        # Click on the consignee filter dropdown
        self.page.locator("//span[normalize-space()='Consignee']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        self.page.locator("//span[normalize-space()='Consignee']").click()
        self.page.wait_for_timeout(1000)

    def Validate_Importer_Tab_Count(self):
        #Navigate to the Importers tab after applying the filter
        self.page.locator("//a[@id='nav-profile-tab']").click()

        # Get record text like "Showing 1-10 of 149,760"
        total_record_text = self.page.locator(".trademo-table-count-text").inner_text().strip()

        if "of" in total_record_text:
            total_count_raw = total_record_text.split("of")[-1].strip()
        else:
            total_count_raw = total_record_text

        # Format total count (e.g. 149.76k)
        total_count = format_number(total_count_raw)
        self.page.wait_for_timeout(2000)
        print(f"✅ Total Exporter count: {total_count}")
        # Verify total record count
        badge_locator = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').first

        if badge_locator.is_visible():
            actual_text = badge_locator.inner_text().strip()
            if actual_text == str(total_count):
                print(f"✅ PASS: Badge count matches → {actual_text}")
            else:
                print(f"❌ FAIL: Importer count mismatch. Expected: {total_count}, Found: {actual_text}")
        else:
            print("❌ FAIL: Badge locator not visible on the page")
        return total_count

    def validate_duplicate_company_names(self):
        # Locate all company names in the "Company Name" column
        company_elements = self.page.locator('[class="trademo-link text-left px-0 text-capitalize "]')
        count = company_elements.count()

        print(f"🔍 Total companies found: {count}")

        seen = set()
        duplicates = []

        for i in range(count):
            name = company_elements.nth(i).inner_text().strip()

            # Normalize text (convert Limited -> Ltd, lowercase etc.)
            normalized_name = name.lower().replace("limited", "ltd").strip()

            if normalized_name in seen:
                print(f"❌ Duplicate found: '{name}' at row {i + 1}")
                duplicates.append(name)
            else:
                seen.add(normalized_name)

        if not duplicates:
            print("✅ No duplicate company names found!")
        else:
            print(f"⚠️ Total Duplicates Found: {len(duplicates)} -> {duplicates}")

    def validate_company_names(self, validate: bool = True):
        company_elements = self.page.locator('[class*="trademo-link"][class*="text-left"]')
        count = company_elements.count()
        if not hasattr(self, "invalid_shipper_entries"):
            self.invalid_shipper_entries = []

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(1).inner_text().strip()

                # Get company name from data-tip or fallback to visible text
                company = company_elements.nth(i).inner_text().strip()
                if company.lower() == self.Company_name.lower() :
                    print(f"✅ Valid company name found: {company} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid company name found: {company} at Sl. No: {sl_no}")


    def Manual_Search(self, C_S_Name:str,search_value):
        self.page.wait_for_timeout(3000)
        # In the search bar, manually enter a Shipper Name and click Search
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible()
        # Manual suggest search for Shipper
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(C_S_Name)
        self.page.locator(".tw-bg-primary-purple-500").click()
        expect(self.page.locator("table tbody tr")).to_have_count(10, timeout=100000)
        self.page.wait_for_timeout(5000)
        #Note the counts in all three tabs-Shipments,Importers and Exporters
        Shipment_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').nth(
            0).inner_text().strip()
        print(f"shipment count:{Shipment_count}")
        Importer_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            0).inner_text().strip()
        print(f"Importer count:{Importer_count}")
        Exporter_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            1).inner_text().strip()
        print(f"Exporter count:{Exporter_count}")
        # Clear the search bar → Select “Consignee ” from category → Type partial name
        self.page.wait_for_timeout(2000)
        self.page.locator('input[type="text"]').clear()
        self.page.wait_for_timeout(2000)
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.locator("(//div[contains(text(),'Consignee')])[1]").click()
        self.page.get_by_role("textbox", name="Search upto 20 Consignee's name or address").fill(search_value)
        self.page.wait_for_timeout(1000)
        elements = self.page.locator('[class="tw-font-medium tw-bg-transparent tw-p-0"]')
        count = elements.count()

        print(f"🔍 Total elements found: {count}")

        for i in range(count):
            text = elements.nth(i).inner_text().strip()
            print(f"Checking element {i + 1}: {text}")

            if text.lower() == search_value.lower():  # case-insensitive match
                print(f"✅ Found '{search_value}' at index {i + 1}")
                elements.nth(i).click()
                return True  # ✅ exit function immediately after finding

        print(f"❌ '{search_value}' not found in {count} elements")
        self.page.wait_for_timeout(3000)
        self.page.locator("//a[@id='nav-home-tab']").click()
        self.page.wait_for_timeout(3000)
        #Observe updated counts in Shipment, Importers, Exporters tabs
        Check_Shipment_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').nth(0).inner_text().strip()
        if Shipment_count == Check_Shipment_count:
            print("✅ Shipment count is matching")
        else:
            print("❌ Shows wrong shipment count")

        Check_Importer_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            0).inner_text().strip()
        if Importer_count == Check_Importer_count:
            print("✅ Importer count is matching")
        else:
            print("❌ Shows wrong Importer count")
        Check_Exporter_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            1).inner_text().strip()
        if Exporter_count == Check_Exporter_count:
            print("✅ Exporter count is matching")
        else:
            print("❌ Shows wrong Exporter count")
        return False

    def Click_ConsigneeName_Option(self):
        self.page.get_by_text('Consignee Name').first.click()
        expect(self.page.locator("//h5[normalize-space()='Filter by Consignee Name']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(2000)

    def check_Consignee_Name_intheGrid_View(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_Name_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif "consignee" in header_text.lower():
                self.C_Name_field = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.C_Name_field is None:
            raise Exception("❌ 'Consignee name' column not found in table headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Consignee Name: {self.C_Name_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- define function to extract one page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()
                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                validation_note = ""  # extra column for Excel

                if validate:
                    consignee_name = row_data[self.C_Name_field]
                    expected = self.normalize_company_name(self.Company_name)
                    actual = self.normalize_company_name(consignee_name)
                    sl_no = row_data[self.S_No]

                    if expected == actual:
                        print(f"✅ PASS: Correct consignee name '{actual}' at Sl. No: {sl_no}")
                    else:
                        print(f"❌ FAIL: Invalid consignee name '{consignee_name}' at Sl. No: {sl_no} "
                              f"(Expected: {self.Company_name})")
                        self.invalid_entries.append({"slNo": sl_no, "Company": consignee_name})
                        validation_note = "❌ Invalid Consignee Name"

                    # --- Popup validation ---
                    rows.nth(i).locator("td").nth(1).click()
                    popup_value = self.page.locator(
                        "//span[normalize-space(text())='Consignee Name']"
                        "/ancestor::div[contains(@class,'col-5')]"
                        "/following-sibling::div[contains(@class,'col-7')]//div[1]"
                    ).inner_text().strip()

                    popup_normalized = self.normalize_company_name(popup_value)

                    if expected == popup_normalized:
                        print(f"✅ Shows valid: {popup_value} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_value} for Sl. No: {sl_no} in view details")
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_value, "Source": "View"})
                        validation_note = "❌ Invalid Consignee Name (View)"

                    # Close popup safely
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_note)
                all_rows_data.append(row_data)

        # --- if pagination enabled, loop pages ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save to Excel only if there were failures ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee Data"

            # Write headers (+ an extra "Validation" column)
            sheet.append(table_headers + ["Validation"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "Invalid" in row_data[-1]:
                    consignee_col_index = self.C_Name_field + 1
                    sheet.cell(row=sheet.max_row, column=consignee_col_index).fill = red_fill

            file_path = f"results/consignee_data_{self.Company_name}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid consignee names highlighted in red)")

        else:
            print("📘 No invalid consignee name found. Excel not generated.")

        return self.invalid_entries

    def check_Consignee_Standardized_Name_Filter(self, validate: bool = True, use_pagination: bool = False):
        """Validate Consignee Standardized Name across all pages and save invalid entries to Excel"""

        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_consignee_entries"):
            self.invalid_consignee_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_S_Name_field = None

        # --- Identify column indexes ---
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif "consignee" in header_text.lower():
                self.C_S_Name_field = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.C_S_Name_field is None:
            raise Exception("❌ 'Consignee Name' column not found in table headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Consignee Name: {self.C_S_Name_field}")

        all_rows_data = []

        while True:
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    row_data.append(cells.nth(j).inner_text().strip())

                # --- Validation ---
                if validate:
                    sl_no = row_data[self.S_No]
                    consignee_name = row_data[self.C_S_Name_field]

                    normalized_expected = re.sub(r"\s+", " ", self.Company_name).strip().lower()
                    normalized_actual = re.sub(r"\s+", " ", consignee_name).strip().lower()

                    # grid validation
                    if normalized_actual == normalized_expected:
                        print(f"✅ Grid: Correct consignee '{normalized_actual}' at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid: Invalid consignee '{consignee_name}' at Sl. No: {sl_no}")
                        self.invalid_consignee_entries.append({"slNo": sl_no, "Company": consignee_name})
                        row_data.append("❌ Invalid Grid Consignee")

                    # popup validation
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()

                    Consignee_Standardized_Name = self.page.locator("//span[normalize-space(text())='Consignee Standardized Name']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//a").inner_text().strip()

                    if normalized_expected == Consignee_Standardized_Name.strip().lower():
                        print(f"✅ View Details: Correct standardized name '{Consignee_Standardized_Name}' at Sl. No: {sl_no}")
                    else:
                        print(f"❌ View Details: Wrong standardized name '{Consignee_Standardized_Name}' at Sl. No: {sl_no}")
                        self.invalid_consignee_entries.append(
                            {"slNo": sl_no, "Company": Consignee_Standardized_Name}
                        )
                        row_data.append("❌ Invalid Standardized Name")

                    consignee_link = self.page.locator("//span[normalize-space(text())='Consignee Standardized Name']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//a")
                    consignee_page1 = consignee_link.inner_text().strip()
                    with self.page.expect_popup() as page1_info:
                        self.page.get_by_role("dialog").get_by_text(consignee_page1).click()
                    page1 = page1_info.value
                    self.page.wait_for_timeout(2000)
                    # expect(page1.get_by_role("main")).to_contain_text(self.Company_name)
                    expect(page1.locator('[class="_portName_9369m_19"]')).to_contain_text(self.Company_name)
                    page1.close()
                    # close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                all_rows_data.append(row_data)

            # --- Pagination handling ---
            if not use_pagination:
                break

            next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
            if next_button.is_enabled():
                next_button.click()
                self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
            else:
                break

        # --- Save results to Excel ---
        if self.invalid_consignee_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee Data"

            sheet.append(table_headers + ["Validation"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and any("❌" in col for col in row_data):
                    for cell in sheet[sheet.max_row]:
                        cell.fill = red_fill

            file_path = f"results/consignee_standardized_data_{self.Company_name}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid rows highlighted)")
        else:
            print("📘 No invalid consignee standardized name found. Excel not generated.")

        return self.invalid_consignee_entries

    def Consignee_Juridiction_option(self):
        self.page.wait_for_timeout(1000)
        self.page.locator("//div[contains(text(),'Consignee Jurisdiction Country')]").first.click()
        expect(self.page.get_by_text("Filter by Consignee Jurisdiction Country")).to_be_visible(timeout=10000)

    def Consignee_Country_Filter_Search(self, country: str):

        self.page.wait_for_timeout(1000)
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").fill(country, timeout=100000)
        self.page.wait_for_timeout(2000)

        # Get search results
        elements = self.page.locator("//span[@data-for='localize-filter-tooltip']")
        count = elements.count()

        found = False
        for i in range(count):
            text = elements.nth(i).text_content()  # safer than inner_text()
            if text:
                text = text.strip().lower()
                print(f"🔎 Found element text: {text}")  # debug log

                if country.lower() in text:
                    print(f"✅ Searched state '{country}' is shown in the filter → {text}")
                    found = True
                    break

        if not found:
            print(f"❌ State '{country}' not found in filter list")

        self.page.locator('input[type="text"]').last.clear()
        self.page.wait_for_timeout(3000)
        # Click on the checkbox of the first result
        all_checkboxes = self.page.locator("//span[@data-for='localize-filter-tooltip']")
        all_count = all_checkboxes.count()
        print("Checkbox count after clearing:", all_count)

        # Find at least one checkbox that is NOT "Maharashtra"
        found_other = False
        for i in range(all_count):
            name = all_checkboxes.nth(i).inner_text().strip()
            if country not in name:
                found_other = True
                break

        if found_other:
            print("✅ Clearing Consignee State shows all the checkboxes (not just searched values).")
        else:
            print(f"❌ Only {country} is still visible after clearing — bug in filter.")

    def Consignee_Country_Apply_Filter(self):
        Juridiction_countrywithcount = self.page.locator('[data-for = "localize-filter-tooltip"]').nth(
            0).inner_text().strip()
        match = re.match(r"([a-zA-Z\s]+)\(([\d,]+)\)", Juridiction_countrywithcount)
        if match:
            self.country_selelcted = match.group(1).strip()
            self.Total_Record_Count = match.group(2).strip()
            print("Selected state/country:", self.country_selelcted)
            print("Total record count:", self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        Country_record_clean = re.sub(r"[()\sA-Za-z,]", "", self.Total_Record_Count)

        # Convert to float
        country_record_num = float(Country_record_clean)
        if 0 <= country_record_num < 100_000:
            self.total_in_millions = f"{country_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{country_record_num / 1_000_000:.2f}M"

        # Convert to millions
        total_in_millions = f"{country_record_num / 1_000_000:.2f}"
        print(total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(1).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()

    def validate_duplicate_country_names(self):
        # Locate all company names in the "Company Name" column
        country_elements = self.page.locator("//div[img[@class='trademo-search-flag mr-2']]")
        count = country_elements.count()

        print(f"🔍 Total countries found: {count}")

        seen = set()
        duplicates = []

        for i in range(count):
            name = country_elements.nth(i).inner_text().strip()

            if name in seen:
                print(f"❌ Duplicate found: '{name}' at row {i + 1}")
                duplicates.append(name)
            else:
                seen.add(name)

        if not duplicates:
            print("✅ No duplicate company names found!")
        else:
            print(f"⚠️ Total Duplicates Found: {len(duplicates)} -> {duplicates}")

    def validate_country_names(self, validate: bool = True):
        country_elements = self.page.locator("//div[img[@class='trademo-search-flag mr-2']]")
        count = country_elements.count()
        if not hasattr(self, "invalid_shipper_entries"):
            self.invalid_country_entries = []

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(1).inner_text().strip()
                country = country_elements.nth(i+1).inner_text().strip()
                if country.lower() == self.country_selelcted.lower():
                    print(f"✅ Valid country name found: {country} at Sl. No: {sl_no} in Importer tab")
                else:
                    print(f"❌ Invalid country name found: {country} at Sl. No: {sl_no} in Importer tab")

    def check_Consignee_Juridiction_Filter(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_JC_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "consignee jurisdiction country":
                self.C_JC_field = i

        # --- Handle missing "Consignee JC" column ---
        if self.C_JC_field is None:
            self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
            self.page.locator("//span[normalize-space()='consignee details']").click()
            self.page.locator("//span[@data-for='Consignee Jurisdiction Countryelipsis']").click()
            self.page.locator("//button[normalize-space()='Save Configuration']").click()
            self.page.wait_for_timeout(2000)
            expect(self.page.get_by_text("View").last).to_be_visible(timeout=10000)

            # 🔄 Re-fetch headers after update
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()

            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "consignee jurisdiction country":
                    self.C_JC_field = i
                    print("✅ 'Consignee Jurisdiction Country' column added successfully!")

        if self.C_JC_field is None:
            raise Exception("❌ 'Consignee Jurisdiction Country' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Consignee JC: {self.C_JC_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                consignee_country = row_data[self.C_JC_field]
                consignee_country_norm = re.sub(r"\s+", " ", consignee_country).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.country_selelcted.lower() == consignee_country_norm:
                        print(f"✅ Grid valid: {consignee_country} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {consignee_country} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": consignee_country, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    popup_country = self.page.locator("//span[normalize-space(text())='Consignee Jurisdiction Country']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip().lower()

                    if self.country_selelcted.lower() == popup_country:
                        print(f"✅ View details valid: {popup_country} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ View details invalid: {popup_country} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in view details"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_country, "Source": "Popup"})

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee Jurisdiction Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.C_JC_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.C_JC_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/consignee_juridiction_{self.country_selelcted}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid consignee jurisdiction found. Excel not generated.")

        return self.invalid_entries

    def Click_ConsigneeType_Option(self):
        self.page.locator("//div[contains(text(),'Consignee Type')]").first.click()
        expect(self.page.locator("//h5[normalize-space()='Filter by Consignee Type']")).to_be_visible(timeout=10000)
        expect(self.page.locator("//span[normalize-space()='Importer']")).to_be_visible()
        expect(self.page.locator("//span[normalize-space()='Logistics Provider']")).to_be_visible()
        expect(self.page.locator("//span[normalize-space()='Potential Individual']")).to_be_visible()

    def check_Consignee_Type_inthegrid_view(self, validate: bool = True, use_pagination: bool = False):
        # Navigate to grid view
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)

            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                sl_no = row_data[self.S_No]

                if validate:
                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    popup_type = self.page.locator("//span[normalize-space(text())='Consignee Type']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip().lower()

                    if self.Company_name.strip().lower() == popup_type:
                        print(f"✅ Shows valid: {popup_type} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_type} for Sl. No: {sl_no} in view details")
                        row_data.append("❌ Invalid data in view")
                        row_data.append("Popup")
                        self.invalid_entries.append(row_data)

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save only invalid results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Consignee Type Data"

            # Headers + Validation columns
            sheet.append(table_headers + ["Validation", "ErrorType"])

            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in self.invalid_entries:
                sheet.append(row_data)
                if row_data[-1] == "Popup":
                    col_idx = self.S_No + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/invalid_consignee_type_{self.Company_name}.xlsx"
            workbook.save(file_path)

            print(f"📁 Only invalid rows saved to {file_path} (highlighted in orange).")
        else:
            print("✅ No invalid entries found. Excel not created.")

        return self.invalid_entries

    def Country_of_Import_Option(self):
        self.page.get_by_text("Country of Import").first.click()
        expect(self.page.get_by_text("Filter by Country of Import")).to_be_visible(timeout=10000)

    def check_country_of_import_inthegrid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_I_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() in ["country of import", "consignee of import"]:
                self.C_I_field = i

        # --- Handle missing "Country of Import" column ---
        if self.C_I_field is None:
            self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
            self.page.locator("//span[normalize-space()='consignee details']").click()
            self.page.locator("//span[@data-for='Country of Importelipsis']").click()
            self.page.locator("//button[normalize-space()='Save Configuration']").click()
            expect(self.page.get_by_text("View").last).to_be_visible(timeout=10000)

            # 🔄 Re-fetch headers after update
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()
            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "country of import":
                    self.C_I_field = i
                    print("✅ 'Country of import' column added successfully!")

        if self.C_I_field is None:
            raise Exception("❌ 'Country of import' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Country of import: {self.C_I_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                consignee_country = row_data[self.C_I_field]
                consignee_country_norm = re.sub(r"\s+", " ", consignee_country).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.country_selelcted.lower() == consignee_country_norm:
                        print(f"✅ Grid valid: {consignee_country} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {consignee_country} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": consignee_country, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    popup_country = self.page.locator("//span[normalize-space(text())='Country of Import']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip().lower()

                    if self.country_selelcted.lower() == popup_country:
                        print(f"✅ Shows valid: {popup_country} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_country} for Sl. No: {sl_no} in view details")
                        validation_status = "❌ Invalid data in View details modal"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_country, "Source": "Popup"})

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Country of Import Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.C_I_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.C_I_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/country_of_import_{self.country_selelcted}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid country of import found. Excel not generated.")

        return self.invalid_entries

    def Consignee_Country_Import_Filter_Search(self, country: str):
        self.page.wait_for_timeout(1000)

        # Fill search box with the country
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").fill(country, timeout=100000)
        self.page.wait_for_timeout(2000)

        # Get search results
        elements = self.page.locator("//span[@data-for='localize-filter-tooltip']")
        count = elements.count()

        found = False
        for i in range(count):
            text = elements.nth(i).text_content()  # safer than inner_text()
            if text:
                text = text.strip().lower()
                print(f"🔎 Found element text: {text}")  # debug log

                if country.lower() in text:
                    print(f"✅ Searched country '{country}' is shown in the filter → {text}")
                    found = True
                    break

        if not found:
            print(f"❌ Country '{country}' not found in filter list")

        # Clear the input afterwards
        self.page.locator('input[type="text"]').last.clear()
        # Ensure last element is visible before clicking
        expect(elements.last).to_be_visible(timeout=5000)
        self.page.wait_for_timeout(3000)

    def Consignee_State_Option(self):
        self.page.get_by_text("Consignee State").first.click()
        expect(self.page.get_by_text("Filter by Consignee State")).to_be_visible(timeout=10000)

    def check_consignee_state_inthegrid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_S_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() in ["consignee state", "Consignee State"]:
                self.C_S_field = i

        # --- Handle missing "Country of Import" column ---
        if self.C_S_field is None:
            self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
            self.page.locator("//span[normalize-space()='consignee details']").click()
            self.page.locator("//span[@data-for='Consignee Stateelipsis']").click()
            self.page.locator("//button[normalize-space()='Save Configuration']").click()
            self.page.wait_for_timeout(2000)

            # 🔄 Re-fetch headers after update
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()
            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "consignee state":
                    self.C_S_field = i
                    print("✅ 'Consignee State' column added successfully!")

        if self.C_S_field is None:
            raise Exception("❌ 'Consignee State' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Consignee State: {self.C_S_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                consignee_state = row_data[self.C_S_field]
                consignee_state_norm = re.sub(r"\s+", " ", consignee_state).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.country_selelcted.lower() == consignee_state_norm:
                        print(f"✅ Grid valid: {consignee_state} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {consignee_state} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": consignee_state, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    popup_country = self.page.locator("//span[normalize-space(text())='Consignee State']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip().lower()

                    if self.country_selelcted.lower() == popup_country:
                        print(f"✅ Shows valid: {popup_country} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_country} for Sl. No: {sl_no} in view details")
                        validation_status = "❌ Invalid data in View details modal"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_country, "Source": "Popup"})

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee State Invalid Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.C_S_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.C_S_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/consignee_state_{self.country_selelcted}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid consignee state found. Excel not generated.")

        return self.invalid_entries

    def Consignee_City_Option(self):
        self.page.get_by_text("Consignee City").first.click()
        expect(self.page.get_by_text("Filter by Consignee City")).to_be_visible(timeout=10000)

    def check_consignee_city_inthegrid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_C_field = None  # Consignee City field index

        # --- Extract headers ---
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "consignee city":
                self.C_C_field = i

        # --- Handle missing "Consignee City" column ---
        if self.C_C_field is None:
            self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
            self.page.locator("//span[normalize-space()='consignee details']").click()
            self.page.locator("//span[@data-for='Consignee Cityelipsis']").click()
            self.page.locator("//button[normalize-space()='Save Configuration']").click()
            expect(self.page.get_by_text("View").last).to_be_visible(timeout=100000)
            # 🔄 Re-fetch headers after update
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()
            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "consignee city":
                    self.C_C_field = i
                    print("✅ 'Consignee City' column added successfully!")

        if self.C_C_field is None:
            raise Exception("❌ 'Consignee City' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Consignee City: {self.C_C_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                consignee_city = row_data[self.C_C_field]
                consignee_city_norm = re.sub(r"\s+", " ", consignee_city).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.country_selelcted.lower() == consignee_city_norm:
                        print(f"✅ Grid valid: {consignee_city} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {consignee_city} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": consignee_city, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    popup_consignee_city = self.page.locator(
                        "//span[normalize-space()='Consignee City']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//div[1]"
                    ).inner_text().strip().lower()

                    if self.country_selelcted.lower() == popup_consignee_city:
                        print(f"✅ Shows valid: {popup_consignee_city} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_consignee_city} for Sl. No: {sl_no} in view details")
                        validation_status = "❌ Invalid data in View details modal"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_consignee_city, "Source": "Popup"})

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee City Invalid Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.C_C_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.C_C_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/consignee_city_{self.country_selelcted}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid consignee city found. Excel not generated.")

        return self.invalid_entries

    def Consignee_Pincode_Option(self):
        self.page.get_by_text("Consignee Pincode").first.click()
        expect(self.page.get_by_text("Filter by Consignee Pincode")).to_be_visible(timeout=10000)

    def check_consignee_pincode_inthegrid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_P_field = None  # Consignee Pincode field index

        # --- Extract headers ---
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "consignee pincode":
                self.C_P_field = i

        # --- Handle missing "Consignee Pincode" column ---
        if self.C_P_field is None:
            self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
            self.page.locator("//span[normalize-space()='consignee details']").click()
            self.page.locator("//span[@data-for='Consignee Pincodeelipsis']").click()
            self.page.locator("//button[normalize-space()='Save Configuration']").click()
            expect(self.page.get_by_text("View").last).to_be_visible(timeout=100000)
            # 🔄 Re-fetch headers after update
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()
            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "consignee pincode":
                    self.C_P_field = i
                    print("✅ 'Consignee Pincode' column added successfully!")

        if self.C_P_field is None:
            raise Exception("❌ 'Consignee Pincode' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Consignee Pincode: {self.C_P_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                consignee_pincode = row_data[self.C_P_field]
                consignee_pincode_norm = re.sub(r"\s+", " ", consignee_pincode).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.Company_name.lower() == consignee_pincode_norm:
                        print(f"✅ Grid valid: {consignee_pincode} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {consignee_pincode} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": consignee_pincode, "Source": "Grid"})

                    # --- Popup check ---
                    rows.nth(i).locator("td").nth(1).click()
                    popup_consignee_pincode = self.page.locator(
                        "//span[normalize-space()='Consignee Pincode']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//div[1]"
                    ).inner_text().strip().lower()

                    if self.Company_name.lower() == popup_consignee_pincode:
                        print(f"✅ Shows valid: {popup_consignee_pincode} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_consignee_pincode} for Sl. No: {sl_no} in view details")
                        validation_status = "❌ Invalid data in View details modal"
                        error_type = "Popup"
                        self.invalid_entries.append(
                            {"slNo": sl_no, "Value": popup_consignee_pincode, "Source": "Popup"})

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee Pincode Invalid Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.C_P_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.C_P_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/consignee_pincode_{self.Company_name}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid consignee pincode found. Excel not generated.")

        return self.invalid_entries




