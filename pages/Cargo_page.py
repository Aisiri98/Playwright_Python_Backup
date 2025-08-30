import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page, expect
from playwright.sync_api import expect, TimeoutError as PlaywrightTimeoutError
invalid_entries =[]

class CargoFilter:
    def __init__(self, page: Page):
        self.page = page

    def check_India_import(self):
        text_content = self.page.locator(".tw-text-nowrap > span").text_content()

        # Select any country from the Data Source filter and click Apply
        if "India Imports" not in text_content:
            self.page.locator("[class='tw-flex tw-items-center tw-gap-3 tw-border-none tw-text-sm']").click()
            clear_all_btn = self.page.get_by_role("button", name="Clear All")
            # Wait for the element to be attached to DOM
            if clear_all_btn.is_visible() and clear_all_btn.is_enabled():
                clear_all_btn.click()
                print("✅ 'Clear All' button clicked.")
            else:
                print("ℹ️ 'Clear All' button is either not visible or disabled.")
            # Loop through each row
            rows = self.page.locator('[class="tw-flex tw-items-center tw-w-full"]')
            count = rows.count()

            for i in range(count):
                row_text = rows.nth(i).inner_text().strip()
                if "India" in row_text:
                    rows.nth(i).locator('input[type="checkbox"]').first.click()
                    break

            # Click on Apply button
            self.page.locator("//span[normalize-space()='Apply']").click()
            print("India import is selected in filter")

        else:
            print("India import is already selected in filter. No action needed.")

    def Verify_Reset_button(self):
        reset_locator = self.page.get_by_text(" Reset").first
        if reset_locator.is_visible():
            print("🔍 ' Reset' is visible, clicking...")
            reset_locator.click()
        else:
            print("ℹ️ ' Reset' not visible.")

    def verify_Cargo_Filter_Option(self):
        # Click on the cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # cargo filter opens with sub-filter options
        expect(self.page.locator("//div[contains(text(),'Product Category')]")).to_be_visible()
        expect(self.page.locator('[class="_filteritemtext_a6k5f_88"]').nth(1)).to_contain_text("HS Code")
        expect(self.page.locator("//div[contains(text(),'Shipment Value (USD)')]")).to_be_visible()
        expect(self.page.locator("//div[contains(text(),'6 Digit HS Code')]")).to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Quantity')]")).to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Per Unit Value (USD)')]")).to_be_visible()

        # Close Cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # Verify the Cargo filter is closed
        expect(self.page.locator("//div[contains(text(),'Product Category')]")).not_to_be_visible()
        expect(self.page.locator('[class="_filteritemtext_a6k5f_88"]').nth(1)).not_to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Shipment Value (USD)')]")).not_to_be_visible()
        expect(self.page.locator("//div[contains(text(),'6 Digit HS Code')]")).not_to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Quantity')]")).not_to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Per Unit Value (USD)')]")).not_to_be_visible()

    def ProductCategory_Filter(self ):
        # Click on the cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # Verify the Cargo filter is open
        expect(self.page.locator("//div[contains(text(),'Product Category')]")).to_be_visible(timeout=10000)
        self.page.locator("//div[contains(text(),'Product Category')]").click(timeout=10000)
        #Verify the Product Category filter is open
        expect(self.page.locator("//h5[normalize-space()='Filter by Product Category']")).to_be_visible(timeout=10000)

    def Cargo_Filter_Search(self, search_product: str):
        #Use the search bar in the modal to search for a product category (e.g., "Electronics")
        self.page.locator("//input[@placeholder='Search in Product Category Filter']").fill(search_product)
        # Wait for the search results to load
        expect(self.page.locator("//span[@class='text-no-transform']")).to_contain_text(search_product, timeout=10000)
        # Click on the first checkbox in the search results
        self.page.locator("//input[@placeholder='Search in Product Category Filter']").clear()
        expect(self.page.locator("//span[normalize-space()='Fabric']")).to_be_visible(timeout=10000)

    def Cargo_Apply_Filter(self):
        self.Product_Name = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Selected Product Name: {self.Product_Name}")

        import re
        # Calculate total from all .grey-text elements
        total = 0

        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(f"Total Product Record Count: {self.Total_Record_Count}")
        # Remove parentheses, spaces, letters like 'M', and commas
        Product_record_clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        Product_record_num = float(Product_record_clean)

        # Convert to millions
        self.total_in_millions = f"{Product_record_num / 1_000_000:.2f}"
        print(self.total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()

    def Verify_Shipment_Record_count(self):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=100000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=100000)
        self.page.wait_for_timeout(3000)
        # Validate total shipment count in 'Shipment' tab
        expect(self.page.locator(
            "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']")).to_contain_text(
            f"{self.total_in_millions}"
        )
        # Validate the total shipment record count from 1-10 of X
        total_record_text = self.page.locator('[class="trademo-table-count-text"]').inner_text()
        total_from_summary = total_record_text.split("of")[1].strip()
        print(f"🔢 Total shipment record count in shipment tab': {total_from_summary}")

        assert total_from_summary == self.Total_Record_Count, "Shipment Record count mismatch!"

    def Verify_Product_Label(self):
        # In the Shipper filter section, the "Product" label shows the count as (1)
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')
        # Click on the shipper filter dropdown
        self.page.locator("//span[normalize-space()='Cargo']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        self.page.locator("//span[normalize-space()='Cargo']").click()
        self.page.wait_for_timeout(1000)

    def Verify_HSCode_Filter(self):
        self.page.pause()
        # Click on the cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # Verify the Cargo filter is open
        self.page.locator('[class="_filteritemtext_a6k5f_88"]').nth(1).click(timeout=10000)
        # Verify the Product Category filter is open
        expect(self.page.locator("//span[@class='modal_header']")).to_contain_text("HS Code Filters", timeout=10000)

    def HSCode_Filter_Search(self, search_hscode: str):
        # Use the search bar in the modal to search for a HS Code (e.g., "123456")
        self.page.locator("//input[@id='hscode-search-input']").fill(search_hscode)
        # Wait for the search results to load
        hs_code_locator = self.page.locator("//label[@class='form-check']//span[@class='form-check-label']")
        # ✅ Wait until the element is visible before extracting text
        hs_code_locator.wait_for(state="visible", timeout=100000)  # waits max 10s
        # Now safely get the text
        hs_code_text = hs_code_locator.inner_text().strip()
        print(f"HS Code text: {hs_code_text}")
        expected_text = f"HS Code {search_hscode}"

        # Partial match check
        if expected_text in hs_code_text:
            print("✅ Contains HS code:", hs_code_text)
        else:
            print("❌ No match found:", hs_code_text)

        # Click on the first checkbox in the search results
        self.page.locator("//input[@id='hscode-search-input']").clear()
        self.page.wait_for_timeout(2000)

    def HSCode_Apply_Filter(self):
        # Get the element text
        text = self.page.locator('[class="form-check-label"]').nth(0).inner_text().strip()
        print(f"Raw text from page: {text}")

        match1 = re.match(r"(HS Code\s+\d+)", text)
        if match1:
            hs_code_text = match1.group(1)  # "HS Code 85"
            print(hs_code_text)
        else:
            raise ValueError(f"Could not extract HS Code from: {text}")
        # Regex to match: HS Code 85 (58,191,911)
        match = re.match(r"HS Code (\d+)\s*\(([\d,]+)\)", text)

        if match:
            self.hs_code = match.group(1)  # "85"
            self.Total_Record_Count = match.group(2)  # "58,191,911"
            print(f"Select HS Code is: {self.hs_code}")
            print(f"Total Record in filter page: {self.Total_Record_Count}")
        else:
            raise ValueError(f"Text format did not match: {text}")

        # Remove parentheses, spaces, letters like 'M', and commas
        HS_Code = re.sub(r"[()\sA-Za-z,]", "", self.Total_Record_Count)

        # Convert to float
        HS_Code_count = float(HS_Code)
        if 0 <= HS_Code_count < 100_000:
            self.total_in_millions = f"{HS_Code_count / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{HS_Code_count / 1_000_000:.2f}M"

        print(self.total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        #validate the selected HS code
        hs_code_elements = self.page.locator('.badge.filter-no.tw-pl-2')
        self.count = hs_code_elements.count()

        found = False
        for i in range(self.count):
            text = hs_code_elements.nth(i).inner_text().strip()
            print(f"Found HScode begins with{hs_code_text}: {text}")
            if hs_code_text.lower() in text.lower():
                found = True

        if found:
            print(f"✅ HS code starting with {self.hs_code} is present in the selected hscode filters.")
        else:
            print(f"❌ No HS code starting with {self.hs_code} found in the selected hscode filters.")

        # Click on 'Apply Filter' button
        self.page.locator("//button[normalize-space()='Apply Filter']").click()

    def Verify_HSCode_Label(self):
        # In the Shipper filter section, the "Product" label shows the count as (1)
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text(self.count)
        # Click on the shipper filter dropdown
        self.page.locator("//span[normalize-space()='Cargo']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text(self.count)
        self.page.locator("//span[normalize-space()='Cargo']").click()
        self.page.wait_for_timeout(1000)

    def Check_CargoFilter_By_HSCode(self, validate: bool = True):
        # --- Initialize only once ---
        if not hasattr(self, "invalid_HSCode_entries"):
            self.invalid_HSCode_entries = []

        # --- Step 1: Find HS Code column index dynamically ---
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()
        hs_code_index = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            if header_text == "hs code":
                hs_code_index = i
                break

        if hs_code_index is None:
            raise Exception("❌ HS Code column not found in the table headers")

        print(f"✅ Found 'HS Code' column at index: {hs_code_index}")

        # --- Step 2: Iterate rows and validate HS Code ---
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:  # ✅ Only validate if flag is True
            for i in range(row_count):
                view = rows.nth(i).locator("td").nth(1)

                hs_code_column = rows.nth(i).locator("td").nth(hs_code_index).inner_text()
                sl_no = rows.nth(i).locator("td").nth(2).inner_text()

                hs_code_column = hs_code_column.strip()
                sl_no = sl_no.strip()

                # --- Validate HS Code ---
                if hs_code_column.startswith(self.hs_code):
                    print(f"✅ Valid HS Code found: {hs_code_column} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid HS Code found: {hs_code_column} at Sl. No: {sl_no}")
                    self.invalid_HSCode_entries.append({"slNo": sl_no, "hsCode": hs_code_column})

                # --- Step 3: Click on View details ---
                view.click()
                self.page.wait_for_timeout(2000)

                HS_Code_View = (
                    self.page.locator("//span[normalize-space(text())='HS Code']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]//a")
                    .inner_text()
                    .strip()
                )
                #if HS_Code_View.startswith(self.hs_code):
                if HS_Code_View == hs_code_column:
                    print(f"✅ Correct HS Code in View: {HS_Code_View} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Incorrect HS Code in View: {HS_Code_View} at Sl. No: {sl_no}")
                    self.invalid_HSCode_entries.append({"slNo": sl_no, "HS_Code_View": HS_Code_View})

                # Scroll into view
                self.page.get_by_role("dialog").get_by_text(HS_Code_View).nth(1).click()

                with self.page.expect_popup() as page1_info:
                    self.page.locator("span").filter(
                        has_text=f"{HS_Code_View}{HS_Code_View}"
                    ).locator("a").click()

                page1 = page1_info.value
                page1.wait_for_load_state("domcontentloaded")

                expect(page1.get_by_role("heading")).to_have_text(
                    f"HS Code - {HS_Code_View}", timeout=10000
                )

                # ✅ handle unload dialog if present
                page1.once("dialog", lambda dialog: dialog.accept())

                # ✅ ensure page is ready before closing
                page1.wait_for_timeout(500)
                page1.close()

                # Close modal
                self.page.locator("//span[@aria-hidden='true']").click(timeout=100000)

        # --- Step 3: Handle pagination ---
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
        #         state="visible", timeout=100000
        #     )
        #     self.Check_CargoFilter_By_HSCode(validate=validate)  # recursive check for next page
        # else:
        #     print("✅ All pages validated")

            if not self.invalid_HSCode_entries:
                print("📘 No invalid HS code found.")
                return

            # --- Save invalid entries to Excel ---
            import os
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid HS Codes"

            # Write headers
            sheet.append(["Sl. No", "HS Code", "Matching Field"])

            # Red fill style
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in self.invalid_HSCode_entries:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=item["slNo"])
                hs_code_cell = sheet.cell(row=row, column=2, value=item.get("hsCode"))
                matching_field_cell = sheet.cell(row=row, column=3, value=item.get("matching_field", ""))
                hs_code_cell.fill = red_fill
                matching_field_cell.fill = red_fill

            file_path = f"results/invalid_hs_codes_{self.hs_code}.xlsx"
            workbook.save(file_path)
            print(f"📁 Invalid HS Code data saved to '{file_path}'")

    def Verify_HSCode_6_Filter(self):
        # Click on the cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # Verify the Cargo filter is open
        self.page.locator("//div[contains(text(),'6 Digit HS Code')]").click()
        # Verify the HSCode filter is open
        expect(self.page.locator("//h5[normalize-space()='Filter by 6 digit HS Code']")).to_be_visible(timeout=10000)

    def HSCode_6_Filter_Search(self, search_6digit_hscode: str):
        # Use the search bar in the modal to search for a HS Code (e.g., "123456")
        self.page.locator("//input[@placeholder='Type to Search 6 digit HS Code']").fill(search_6digit_hscode)
        # Wait for the search results to load
        hs_code_locator = self.page.locator("//span[@class='text-no-transform']")
        hs_code_text = hs_code_locator.inner_text().strip()
        expected_text = f"{search_6digit_hscode}"

        # Partial match check
        if expected_text in hs_code_text:
            print("✅ Contains HS code:", hs_code_text)
        else:
            print("❌ No match found:", hs_code_text)

        self.page.locator("//input[@placeholder='Type to Search 6 digit HS Code']").clear()
        self.page.wait_for_timeout(2000)

    def HSCode_6_Apply_Filter(self):
        # Get the element text
        text = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Raw text from page: {text}")
        self.hs_code = text.split()[0][:6]
        print(self.hs_code)

        #Get the total record count
        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        HS_Code_Clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        HS_Code_record_num = float(HS_Code_Clean)

        # Convert to millions
        self.total_in_millions = f"{HS_Code_record_num / 1_000_000:.2f}"
        print(self.total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()


    def Verify_Shipment_Value_Filter(self):
        # Click on the cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # Verify the Cargo filter is open
        self.page.locator('[class="_filteritemtext_a6k5f_88"]').nth(2).click()
        # Verify the HSCode filter is open
        expect(self.page.locator("//h1[normalize-space()='Filter by Shipment Value']")).to_be_visible(timeout=10000)

    def Shipment_Value_Apply(self, min: str, max: str):
        # Use the search bar in the modal to search for a HS Code (e.g., "123456")
        self.page.locator('input[type="number"]').nth(0).clear()
        self.page.locator('input[type="number"]').nth(0).fill(min)
        self.Min_Range = self.page.locator('input[type="number"]').nth(0).input_value()
        print(f"{self.Min_Range}")
        self.page.locator('input[type="number"]').nth(1).clear()
        self.page.locator('input[type="number"]').nth(1).fill(max)
        self.Max_Range = self.page.locator('input[type="number"]').nth(1).input_value()
        print(f"{self.Max_Range}")
        # Click on 'Apply Filter' button
        self.page.locator("//button[normalize-space()='Apply Filter']").click()

    def Validate_Shipment_Value_in_the_grid(self, validate: bool = True):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=100000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=100000)
        self.page.wait_for_timeout(3000)

        # Initialize storage for invalid entries only once
        if validate and not hasattr(self, "invalid_shipmentvalue_entries"):
            self.invalid_shipmentvalue_entries = []

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                shipment_value_text = rows.nth(i).locator("td").nth(13).inner_text().strip()
                sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()

                try:
                    # --- Convert shipment value to float (remove commas, $ etc.) ---
                    shipment_value_num = float(shipment_value_text.replace(",", "").replace("$", "").strip())

                    # --- Convert min & max to float ---
                    min_range_num = float(self.Min_Range.replace(",", "").strip())
                    max_range_num = float(self.Max_Range.replace(",", "").strip())

                    # --- Format for printing only ---
                    formatted_MinRange = f"{min_range_num:,.2f}"
                    formatted_MaxRange = f"{max_range_num:,.2f}"
                    formatted_ShipmentValue = f"{shipment_value_num:,.2f}"

                    # --- Validate shipment value ---
                    if min_range_num <= shipment_value_num <= max_range_num:
                        print(f"✅ Valid shipment value: {formatted_ShipmentValue} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Invalid shipment value: {formatted_ShipmentValue} at Sl. No: {sl_no}")
                        self.invalid_shipmentvalue_entries.append(
                            {"slNo": sl_no, "shipment_value": formatted_ShipmentValue}
                        )

                except ValueError:
                    # Handle blank or invalid values gracefully
                    print(f"⚠️ Skipped invalid/blank shipment value '{shipment_value_text}' at Sl. No: {sl_no}")
                    self.invalid_shipmentvalue_entries.append(
                        {"slNo": sl_no, "shipment_value": shipment_value_text}
                    )

        # --- Check next page ---
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)
        #     self.Validate_Shipment_Value_in_the_grid(validate=True)  # recursive check
        # else:
        #     print("✅ All pages validated")

            if not self.invalid_shipmentvalue_entries:
                print("📘 No invalid shipment values found.")
                return

            # Save invalid entries to Excel
            import os
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Shipment Values"

            # Write headers
            sheet.append(["Sl. No", "Shipment_Value"])

            # Red fill style for invalid cells
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in self.invalid_shipmentvalue_entries:
                sheet.append([item["slNo"], item["shipment_value"]])
                # Fill last cell with red color
                sheet.cell(row=sheet.max_row, column=2).fill = red_fill

            workbook.save("results/invalid_shipment_Cargo.xlsx")
            print("📁 Invalid shipment value data saved to 'results/invalid_shipment_Cargo.xlsx'")

    def Open_Quanity_Filter(self):
        # Click on the cargo filter
        self.page.locator("//span[normalize-space()='Cargo']").click(timeout=10000)
        # Verify the Cargo filter is open
        self.page.get_by_text("Quantity").nth(0).click()
        # Verify the Quantity filter is open
        expect(self.page.locator("//button[normalize-space()='Quantity Units']")).to_be_visible(timeout=10000)
        expect(self.page.locator("//button[normalize-space()='Quantity Value']")).to_be_visible(timeout=10000)

    def Verify_Quanity_Units_Search_Filter(self , search_quantity_unit: str):
        # Navigate to the search bar in the Quantity Unit tab to search a specific unit
        self.page.locator("//input[@placeholder='Search Quantity Unit']").fill(search_quantity_unit)
        # Wait for the search results to load
        Unit_locator = self.page.locator("//span[@class='text-no-transform']")
        Unit_Code_text = Unit_locator.inner_text().strip()
        self.unit = f"{Unit_Code_text}"
        #print(f"Expected text: {expected_text}")

        # # Partial match check
        if search_quantity_unit in self.unit :
            print("✅ Shows valid search:", self.unit)
        else:
            print("❌ Shows invalid search:", self.unit)

        self.page.locator("//input[@placeholder='Search Quantity Unit']").clear()
        self.page.wait_for_timeout(2000)

    def Quantity_Unit_Apply_Filter(self):
        # Get the element text
        self.Unit = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Selected Checkbox is : {self.Unit}")
        #Get the total record count
        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        Unit_Clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        Unit_record_num = float(Unit_Clean)

        # Convert to millions
        self.total_in_millions = f"{Unit_record_num / 1_000_000:.2f}"
        print(self.total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        self.page.wait_for_timeout(2000)

    def Verify_Quantity_Unit_Filter_After_Reopen(self):
        # Previously selected unit should still appear checked
        expect(self.page.locator('input[type="checkbox"]').nth(0)).to_be_checked()
        #Click on close modal
        self.page.locator('[aria-label="Close"]').click()

    def Validate_Quantity_in_the_grid(self, validate: bool = True, invalid_entries=None):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=100000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=100000)
        self.page.wait_for_timeout(3000)

        # keep track of invalid entries across pages
        if invalid_entries is None:
            invalid_entries = []

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                Quantity_unit = rows.nth(i).locator("td").nth(12).inner_text().strip()
                sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()

                print(f"Unit to check: {self.Unit.lower()}")

                if self.Unit.lower() in Quantity_unit.lower():
                    print(f"✅ Valid quantity found : {Quantity_unit} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid quantity found : {Quantity_unit} at Sl. No: {sl_no}")
                    invalid_entries.append({"slNo": sl_no, "quantity": Quantity_unit})

        # --- check next page ---
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)
        #     self.Validate_Quantity_in_the_grid(validate, invalid_entries)  # recursive call with same list
        # else:
        #     print("✅ All pages validated")

            if not invalid_entries:
                print("📘 No invalid quantity values found.")
                return

            # --- save invalid entries to Excel ---
            import os
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Quantity Values"

            # headers
            sheet.append(["Sl. No", "Quantity_Value"])

            # red fill for invalid cells
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in invalid_entries:
                sheet.append([item["slNo"], item["quantity"]])
                sheet.cell(row=sheet.max_row, column=2).fill = red_fill

            file_name = f"results/invalid_quantity_Cargo_{Quantity_value}.xlsx"
            workbook.save(file_name)
            print(f"📁 Invalid quantity data saved to {file_name}")

    def Quantity_Value_Apply_Filter(self , quantity:str, min:str , max:str):
        #Click on Quantity value
        self.page.locator("//button[normalize-space()='Quantity Value']").click()
        #From the Select Quantity Unit dropdown, choose a unit (e.g., Kilogram)
        self.page.get_by_text("Select Quantity", exact=True).click()
        self.page.locator('input[type="text"]').last.fill(quantity)
        expect(self.page.locator('[class=" css-dp3zo9-option"]')).to_be_visible()
        #Choose any unit
        self.page.pause()
        self.Units = self.page.locator('[class=" css-dp3zo9-option"]').inner_text()
        self.page.locator('[class=" css-dp3zo9-option"]').click()
        print(f"Selected value : {self.Units}")

        self.page.locator('input[type="number"]').nth(0).clear()
        self.page.locator('input[type="number"]').nth(0).fill(min)
        self.Min_Range = self.page.locator('input[type="number"]').nth(0).input_value().strip()
        self.page.locator('input[type="number"]').nth(1).clear()
        self.page.locator('input[type="number"]').nth(1).fill(max)
        self.Max_Range = self.page.locator('input[type="number"]').nth(1).input_value().strip()
        print(f"Max: {self.Max_Range}")
        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        self.page.wait_for_timeout(2000)

    def Validate_Quantity_Value_in_the_grid(self, validate: bool = True, invalid_entries=None):
        # Initialize invalid_entries only once
        if invalid_entries is None:
            invalid_entries = []

        # Go to grid tab
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=100000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=100000)
        self.page.wait_for_timeout(3000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                quantity_text = rows.nth(i).locator("td").nth(12).inner_text().strip()
                sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()

                # --- Extract number + unit ---
                match = re.match(r"([\d,\.]+)\s*(.*)", quantity_text)
                if not match:
                    continue  # skip invalid rows
                number_str, unit = match.groups()

                # Convert to number
                try:
                    quantity_value = float(number_str.replace(",", "").strip())
                except ValueError:
                    continue  # skip non-numeric values

                unit = unit.strip() if unit else None

                # Convert min/max ranges
                min_range_num = float(self.Min_Range.replace(",", "").strip())
                max_range_num = float(self.Max_Range.replace(",", "").strip())

                # --- Validate numeric range ---
                if min_range_num <= quantity_value <= max_range_num:
                    print(f"✅ Valid quantity: {quantity_value:,.2f} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid quantity: {quantity_value:,.2f} at Sl. No: {sl_no}")
                    invalid_entries.append(
                        {"slNo": sl_no, "quantity": f"{quantity_value:,.2f}", "unit": unit or "N/A"}
                    )

                # --- Validate unit ---
                print(f"DEBUG Units → Expected: {self.Units.lower()}, Found: {unit.lower() if unit else 'N/A'}")
                if unit and unit.lower() in self.Units.lower():
                    print(f"✅ Valid quantity unit: {unit} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid quantity unit: {unit or 'N/A'} at Sl. No: {sl_no}")
                    invalid_entries.append(
                        {"slNo": sl_no, "quantity": f"{quantity_value:,.2f}", "unit": unit or "N/A"}
                    )

        # --- Check next page ---
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
        #         state="visible", timeout=100000
        #     )
        #     # Recursive call with same invalid_entries
        #     return self.Validate_Quantity_Value_in_the_grid(validate, invalid_entries)
        #
        # else:
        #     print("✅ All pages validated")

            if not invalid_entries:
                print("📘 No invalid quantity values found.")
                return

            # --- Save invalid entries to Excel ---
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid quantity Values"

            # Headers
            sheet.append(["Sl. No", "Quantity_Value", "Unit"])

            # Red fill style
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in invalid_entries:
                sheet.append([item["slNo"], item["quantity"], item.get("unit", "")])
                sheet.cell(row=sheet.max_row, column=2).fill = red_fill
                sheet.cell(row=sheet.max_row, column=3).fill = red_fill

            # Unique file name (timestamped)
            unit_name = str(self.Units).replace(" ", "_")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"results/invalid_quantityunit_Cargo_{unit_name}_{timestamp}.xlsx"
            workbook.save(file_name)

            print(f"📁 Invalid quantity data saved to '{file_name}'")

    def Open_PerUnitUSD_Filter(self):
        # Click on the cargo
        self.page.wait_for_timeout(2000)
        self.page.locator("//span[normalize-space()='Cargo']").click()
        # Verify the Cargo filter is open
        self.page.locator("//div[contains(text(),'Per Unit Value (USD)')]").first.click()
        # Verify the Quantity filter is open
        expect(self.page.locator("//h5[normalize-space()='Filter by Per Unit Value (USD)']")).to_be_visible(timeout=100000)

    def Verify_Per_Unit_Search_Filter(self , search_quantity_unit: str):
        # Navigate to the search bar in the Quantity Unit tab to search a specific unit
        self.page.locator("//input[@placeholder='Search Quantity Unit']").fill(search_quantity_unit)
        # Wait for the search results to load
        Unit_locator = self.page.locator("//span[@class='text-no-transform']")
        Unit_Code_text = Unit_locator.inner_text().strip()
        self.unit = f"{Unit_Code_text}"
        #print(f"Expected text: {expected_text}")

        # # Partial match check
        if search_quantity_unit in self.unit :
            print("✅ Shows valid search:", self.unit)
        else:
            print("❌ Shows invalid search:", self.unit)

        self.page.locator("//input[@placeholder='Search Quantity Unit']").clear()
        self.page.wait_for_timeout(2000)

    def PerUnitUSD_Apply_Filter(self, min: str, max: str,quantity:str):
        # Click the “+” icon near the filter field to add another filter row.
        self.page.wait_for_timeout(2000)
        self.page.locator("#searchIndividualFilterPanel").get_by_role("img").click(force=True)
        self.page.wait_for_timeout(2000)
        expect(self.page.get_by_text("Select quantity unit").nth(1)).to_be_visible(timeout=10000)
        expect(self.page.get_by_text("Minimum range").nth(1)).to_be_visible()
        expect(self.page.get_by_text("Maximum range").nth(1)).to_be_visible()

        # Click the “–” icon next to the added row to remove it.
        self.page.locator("#searchIndividualFilterPanel").get_by_role("img").first.click()
        expect(self.page.get_by_text("Select quantity unit").nth(1)).not_to_be_visible(timeout=10000)
        expect(self.page.get_by_text("Minimum range").nth(1)).not_to_be_visible()
        expect(self.page.get_by_text("Maximum range").nth(1)).not_to_be_visible()

        # From the Select Quantity Unit dropdown, choose a unit (e.g., Kilogram)
        self.page.get_by_text("Select Quantity", exact=True).click()
        self.page.locator('input[type="text"]').last.fill(quantity)
        expect(self.page.locator('[class=" css-dp3zo9-option"]')).to_be_visible()
        # Choose any unit
        self.page.pause()
        self.Units = self.page.locator('[class=" css-dp3zo9-option"]').inner_text()
        self.page.locator('[class=" css-dp3zo9-option"]').click()
        print(f"Selected value : {self.Units}")

        #Enter a specific Per Unit Value range and select a Quantity Unit from the dropdown.
        self.page.locator('input[type="number"]').nth(0).clear()
        self.page.locator('input[type="number"]').nth(0).fill(min)
        self.Min_Range = self.page.locator('input[type="number"]').nth(0).input_value().strip()
        self.page.locator('input[type="number"]').nth(1).clear()
        self.page.locator('input[type="number"]').nth(1).fill(max)
        self.Max_Range = self.page.locator('input[type="number"]').nth(1).input_value().strip()
        print(f"Max: {self.Max_Range}")

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        self.page.wait_for_timeout(2000)

    def Validate_Quantity_Per_Unit_in_the_grid(self, validate: bool = True):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=100000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=100000)
        self.page.wait_for_timeout(3000)

        # Keep invalid_entries consistent across calls
        if not hasattr(self, "invalid_shipper_entries"):
            self.invalid_shipper_entries = []
        invalid_entries = self.invalid_shipper_entries

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                Quantity_unit = rows.nth(i).locator("td").nth(12).inner_text().strip()
                Per_unit = rows.nth(i).locator("td").nth(14).inner_text().strip()
                sl_no = rows.nth(i).locator("td").nth(2).inner_text().strip()
                view = rows.nth(i).locator("td").nth(1)

                # --- Split number and unit ---
                match = re.match(r"([\d,\.]+)\s*(.*)", Per_unit)
                number, unit = None, None
                if match:
                    number = match.group(1)  # "5,000.00"
                    unit = match.group(2)  # "per carat"
                    print("Number:", number)
                    print("Unit:", unit)

                if number:
                    quantity_value = float(number.replace(",", "").strip())
                    min_range_num = float(self.Min_Range.replace(",", "").strip())
                    max_range_num = float(self.Max_Range.replace(",", "").strip())

                    # --- Validate per unit value ---
                    if min_range_num <= quantity_value <= max_range_num:
                        print(f"✅ Valid quantity found : {Quantity_unit} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Invalid quantity found : {Quantity_unit} at Sl. No: {sl_no}")
                        invalid_entries.append({
                            "slNo": sl_no,
                            "Quantity_Value": Quantity_unit,
                            "Per_unit": Per_unit
                        })

                # --- Validate per unit unit string ---
                expected_unit = f"per {self.Units}".lower()
                if unit.lower() in expected_unit:
                    print(f"✅ Valid quantity unit found : {Quantity_unit} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid quantity unit found : {Quantity_unit} at Sl. No: {sl_no}")
                    invalid_entries.append({
                        "slNo": sl_no,
                        "Quantity_Value": Quantity_unit,
                        "Per_unit": Per_unit
                    })

                # --- Validate in view screen ---
                view.click()
                self.page.wait_for_timeout(2000)

                Quantity_view = self.page.locator("//span[@class='text-no-transform']").inner_text().strip()
                if Quantity_view.lower() in self.Units.lower():
                    print(f"✅ Valid quantity found  : {Quantity_view} at Sl. No: {sl_no} in view screen")
                else:
                    print(f"❌ Shows incorrect quantity {Quantity_view} at Sl. No: {sl_no} in view screen")

                PerUnit_view1 = self.page.locator('[class="col-7 bold-data text-capitalize "]')
                count1 = PerUnit_view1.count()

                if count1 > 0:
                    # Get the second last element's text
                    PerUnit_view = PerUnit_view1.nth(count1 - 2).inner_text().strip()
                    print(f"Last element text: {PerUnit_view}")
                else:
                    print("⚠ No matching elements found.")
                    PerUnit_view = ""
                if  Per_unit.lower() in PerUnit_view.lower():
                    print(f"✅ Valid unit found : {PerUnit_view} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid unit found : {PerUnit_view} at Sl. No: {sl_no}")
                    invalid_entries.append({
                        "slNo": sl_no,
                        "Quantity_Value": Quantity_unit,
                        "Per_unit": Per_unit,
                        "Quantity_view": Quantity_view,
                        "PerUnit_view": PerUnit_view
                    })

                self.page.locator('[class="close"]').click(timeout=10000)

        # --- Check next page ---
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
        #         state="visible", timeout=100000
        #     )
        #     self.Validate_Quantity_Per_Unit_in_the_grid()  # recursive check
        # else:
        #     print("✅ All pages validated")

            if not invalid_entries:
                print("📘 No invalid quantity values found.")
                return

            # --- Save invalid entries to Excel ---
            import os
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid quantity Values"

            # Write headers
            sheet.append(["Sl. No", "Quantity_unit", "Per_unit", "Quantity_view", "PerUnit_view"])

            # Red fill style for invalid cells
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in invalid_entries:
                sheet.append([
                    item.get("slNo", ""),
                    item.get("Quantity_Value", ""),
                    item.get("Per_unit", ""),
                    item.get("Quantity_view", ""),
                    item.get("PerUnit_view", "")
                ])
                # Fill last cell with red color
                sheet.cell(row=sheet.max_row, column=2).fill = red_fill

            workbook.save("results/invalid_quantity_Cargo.xlsx")
            print("📁 Invalid quantity data saved to 'results/invalid_quantity_Cargo.xlsx'")









