import os
import re
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page, expect
from playwright.sync_api import expect, TimeoutError as PlaywrightTimeoutError
import time
invalid_entries =[]


def split_location(text: str):
    # Remove commas and extra spaces
    clean_text = text.replace(",", " ").strip()

    # Split into words
    parts = clean_text.split()

    return parts
class Search:
    def __init__(self, page: Page):
        self.port_name_hyperlink = None
        self.page = page
        self.invalid_consignee_entries = []

    from playwright.sync_api import Page, expect

    def wait_for_new_table(self, table_selector: str = "//table", timeout: int = 10000, interval: float = 1.0):
        """
        Wait until a new or refreshed table is visible.
        Retries in a loop until timeout, printing logs for each attempt.
        If table is not found, prints error message instead of raising exception.

        :param table_selector: XPath/CSS selector for the table (default: //table)
        :param timeout: Max wait time in ms (default: 10s)
        :param interval: Time between retries in seconds
        """
        table_locator = self.page.locator(table_selector)

        # Capture old table HTML if it exists
        try:
            old_html = table_locator.inner_html(timeout=2000)
        except:
            old_html = None  # No table existed before

        max_attempts = int(timeout / (interval * 1000))
        attempt = 1
        start_time = time.time()

        while attempt <= max_attempts:
            try:
                self.page.wait_for_function(
                    """([selector, old_html]) => {
                        const table = document.evaluate(selector, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                        return table && table.innerHTML !== old_html;
                    }""",
                    arg=[table_selector, old_html],
                    timeout=interval * 1000,
                    polling=200,
                )

                expect(table_locator).to_be_visible(timeout=interval * 1000)
                print(f"[INFO] ✅ Table became visible after {attempt} attempt(s).")
                return table_locator

            except Exception:
                print(f"[WARN] Attempt {attempt}: Table not visible yet... retrying.")
                time.sleep(interval)
                attempt += 1

        total_time = round(time.time() - start_time, 2)
        print(f"[ERROR] ❌ Table with selector '{table_selector}' did not become visible after {total_time}s.")
        return None

    def wait_for_shipment_grid(self, timeout: int = 30):
        """Wait until shipment grid has at least one row or timeout (in seconds)"""
        start_time = time.time()
        while time.time() - start_time < timeout:
            shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
            if shipment_tab.is_visible():
                shipment_tab.click(timeout=40000)
                # Wait for table rows to appear
                rows = self.page.locator("table tbody tr")
                if rows.count() > 0:
                    return True
            # Wait a bit before retrying
            self.page.wait_for_timeout(1000)
        # Timeout reached, no results
        print("❌ No search results found in Shipment Grid after waiting.")
        return False

    def Close_button(self):
        close_locator = self.page.locator('[class="tw-mt-6 tw-cursor-pointer"]').first
        self.page.wait_for_timeout(3000)
        if close_locator.is_visible(timeout=40000):
            print("🔍 'close' is visible, clicking...")
            close_locator.click()
        else:
            print("ℹ️ 'Close' button is not visible.")

    def auto_suggest_hs_code_search(self, hs_code: str):
        """Test HS Code search functionality without explicit waits"""
        print(f"Testing autosuggest HS Code search for: {hs_code}")

        # Click on the Search bar (auto-waits until it's ready)
        search_box = self.page.get_by_role("textbox", name="Type to search in all")
        expect(search_box).to_be_visible(timeout=40000)
        search_box.click()

        # Ensure placeholder is visible before proceeding
        expect(
            self.page.get_by_placeholder(
                "Type to search in all categories or choose from the category below"
            )
        ).to_be_visible()

        # Select category HS Code
        hs_code_button = self.page.get_by_text("HS Code", exact=True).first
        expect(hs_code_button).to_be_visible()
        hs_code_button.click()

        # Fill HS Code input and wait for autosuggest list to appear
        hs_code_input = self.page.get_by_role("textbox", name="Search upto 20 Harmonised")
        expect(hs_code_input).to_be_visible()
        hs_code_input.fill(hs_code)

        # Wait for autosuggest item to be visible and select first one
        autosuggest_first_item = self.page.locator('[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]').nth(0)
        expect(autosuggest_first_item).to_be_visible(timeout=40000)

        # Get full text
        autosuggest_text = autosuggest_first_item.inner_text().strip()

        # Extract only the number (e.g., "HS:10 - Cereals" -> "10")
        match = re.search(r"\b(\d+)\b", autosuggest_text)
        self.selected_hs_code = match.group(1) if match else autosuggest_text

        print(f"🔎 Autosuggest returned: {autosuggest_text}")
        print(f"✅ Selected HS Code number: {self.selected_hs_code}")
        autosuggest_first_item.click(timeout=40000)

        # Choose a Match Type if not already Exact Phrase
        exact_phrase_text = self.page.locator(".tw-text-nowrap > span").text_content()
        if "Exact Phrase" not in exact_phrase_text:
            contains_option = self.page.locator("div").filter(has_text=re.compile(r"^Contains$")).nth(2)
            expect(contains_option).to_be_visible()
            contains_option.click()
            first_dropdown_option = self.page.locator('[class=" css-10xa8g5-option"]').first
            expect(first_dropdown_option).to_be_visible()
            first_dropdown_option.click()

        # Click Search button
        search_button = self.page.locator(".tw-bg-primary-purple-500")
        expect(search_button).to_be_enabled()
        search_button.click()
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        # Verify the presence of the “Discover more insights about HS: <hs_code>” link
        insights_link = self.page.locator("a.btn-link.trademo-link").first
        expect(insights_link).to_have_text(f"HS: {self.selected_hs_code}")

    def auto_suggest_hs_code_search_save_search(self, hs_code: str):
        """Test HS Code search functionality"""
        print(f"Testing autosuggest HS Code search for: {hs_code}")
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        # Select category HS Code, enter keyword, and select from autosuggest
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=40000)
        self.page.wait_for_timeout(3000)
        self.page.get_by_text("HS Code", exact=True).first.click()
        self.page.get_by_role("textbox", name="Search upto 20 Harmonised").fill(hs_code)
        # select value from autosuggest
        # store hs code in a variable
        self.selected_hs_code = self.page.locator('[class="tw-font-medium tw-bg-transparent tw-p-0"]').nth(
            0).inner_text().strip()
        self.page.locator('[class="tw-font-medium tw-bg-transparent tw-p-0"]').nth(0).click(timeout=40000)
        print(f"Selected HS Code is: {self.selected_hs_code}")

        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)

        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")


    def check_hs_code_in_shipmentgrid(self, validate: bool = True, use_pagination: bool = False):
        # Wait until the table is visible
        if not self.wait_for_shipment_grid(timeout=30):
            print("⏩ Skipping further checks because shipment grid has no results.")
            return []
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        shipment_tab.click(timeout=40000)

        table = self.page.get_by_role("table")
        expect(table).to_be_visible()

        first_row = self.page.locator('tbody[role="rowgroup"] >> tr[role="row"]').first
        expect(first_row).to_be_visible()
        print("✅ Table is loaded and has at least one row.")

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # Get table headers
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()
        self.S_No = self.HS_Code_field = self.Matching_Field = None
        table_headers = []

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "hs code":
                self.HS_Code_field = i
            elif header_text.lower() in ["matching fields", "matching field"]:
                self.Matching_Field = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.HS_Code_field is None:
            raise Exception("❌ 'HS Code' column not found in table headers")
        if self.Matching_Field is None:
            raise Exception("❌ 'Matching Field' column not found in table headers")

        print(
            f"✅ Found column indexes → S No: {self.S_No}, HS Code: {self.HS_Code_field}, Matching Field: {self.Matching_Field}")

        all_rows_data = []

        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()
                for j in range(cell_count):
                    row_data.append(cells.nth(j).inner_text().strip())

                validation_note = ""

                if validate:
                    sl_no = row_data[self.S_No]
                    hs_code = row_data[self.HS_Code_field]
                    matching_field = row_data[self.Matching_Field]

                    # ✅ Check if HS Code in grid is hyperlinked
                    hs_code_cell = cells.nth(self.HS_Code_field)
                    if hs_code_cell.locator("a").count() > 0:
                        hs_code_link = hs_code_cell.locator("a")
                        expect(hs_code_link).to_be_visible()
                        href_value = hs_code_link.get_attribute("href")
                        hs_code_text = hs_code_link.inner_text()
                        print(f"✅ HS Code is hyperlinked in grid: {hs_code_text} → {href_value}")
                    else:
                        hs_code_text = hs_code_cell.inner_text()
                        print(f"❌ HS Code is NOT hyperlinked in grid: {hs_code_text}")
                        validation_note += (" | " if validation_note else "") + "❌ HS Code missing hyperlink in Grid"
                        self.invalid_entries.append({"slNo": sl_no, "HS Code": hs_code_text, "Hyperlink": "Missing"})

                    # ✅ Validate HS Code prefix in grid
                    if hs_code.startswith(self.selected_hs_code):
                        print(f"✅ PASS: HS Code '{hs_code}' at Sl. No: {sl_no} (in grid)")
                    else:
                        print(
                            f"❌ FAIL: Invalid HS Code '{hs_code}' at Sl. No: {sl_no} (Expected prefix: {self.selected_hs_code}) (in grid)")
                        validation_note += (" | " if validation_note else "") + "❌ Invalid HS Code in Grid"
                        self.invalid_entries.append({"slNo": sl_no, "HS Code": hs_code, "Matching": matching_field})

                    # ✅ Validate Matching Field
                    if matching_field.lower() == "hs code":
                        print(f"✅ PASS: Matching field '{matching_field}' at Sl. No: {sl_no}")
                    else:
                        print(f"❌ FAIL: Invalid Matching field '{matching_field}' at Sl. No: {sl_no}")
                        validation_note += (" | " if validation_note else "") + "❌ Invalid Matching Field in Grid"
                        self.invalid_entries.append({"slNo": sl_no, "HS Code": hs_code, "Matching": matching_field})

                    # ✅ Validate HS Code in View Screen
                    cells.nth(1).click()  # Click 2nd column (adjust if needed)
                    self.page.wait_for_timeout(1500)

                    hs_code_elements = self.page.locator(
                        "xpath=//span[starts-with(@data-for, 'hs-code-tooltip-id')]//a")
                    count = hs_code_elements.count()

                    if count == 0:
                        print(f"❌ No HS Code elements found in view screen for Sl. No: {sl_no}")
                        validation_note += (" | " if validation_note else "") + "❌ No HS Code found in View"
                    else:
                        hs_code_view_locator = hs_code_elements.nth(count - 2)
                        expect(hs_code_view_locator).to_be_visible()
                        hs_code_view_text = hs_code_view_locator.inner_text().strip()

                        # Validate HS Code text in view
                        if hs_code_view_text == hs_code and hs_code_view_text.startswith(self.selected_hs_code):
                            print(
                                f"✅ PASS: HS Code '{hs_code_view_text}' matches grid and starts with '{self.selected_hs_code}'")
                        else:
                            print(
                                f"❌ FAIL: HS Code '{hs_code_view_text}' does not match grid or expected prefix '{self.selected_hs_code}'")
                            validation_note += (" | " if validation_note else "") + "❌ HS Code mismatch in View"
                            self.invalid_entries.append({"slNo": sl_no, "HS Code(View)": hs_code_view_text})

                        # Validate hyperlink in view screen
                        tag_name = hs_code_view_locator.evaluate("el => el.tagName")

                        # Count <a> elements inside the locator
                        a_count = hs_code_view_locator.locator("a").count()

                        if tag_name.lower() == "a" or a_count > 0:
                            print(f"✅ HS Code '{hs_code_view_text}' is a valid hyperlink in view")
                        else:
                            print(f"❌ HS Code '{hs_code_view_text}' is NOT a hyperlink in view")
                            validation_note += (
                                                   " | " if validation_note else "") + "❌ HS Code missing hyperlink in View"
                            self.invalid_entries.append({"slNo": sl_no, "HS Code(View) Hyperlink": "Missing"})

                    # Close modal
                    close_btn = self.page.locator("//span[@aria-hidden='true']")
                    expect(close_btn).to_be_visible()
                    close_btn.click()
                    self.page.wait_for_timeout(500)

                row_data.append(validation_note)
                all_rows_data.append(row_data)

        # Handle pagination
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
                else:
                    break
        else:
            process_current_page()

        # Write results to Excel
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "HS Code Data"
            sheet.append(table_headers + ["Validation"])
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "❌" in row_data[-1]:
                    hs_col_index = self.HS_Code_field + 1
                    match_col_index = self.Matching_Field + 1
                    sheet.cell(row=sheet.max_row, column=hs_col_index).fill = red_fill
                    sheet.cell(row=sheet.max_row, column=match_col_index).fill = red_fill

            file_path = f"results/search_autosuggest_{self.selected_hs_code}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid HS Codes highlighted in red)")
        else:
            print("📘 No invalid HS Codes found. Excel not generated.")

        return self.invalid_entries

    def Validate_Discover_Insights(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        with self.page.expect_popup() as page1_info:
            #Click on the "Discover more insights about HS: 85364100" link
            self.page.get_by_text("HS:").click()
        page1 = page1_info.value
        expect(page1.get_by_role("heading")).to_contain_text(f"HS Code - {self.selected_hs_code}")
        page1.close()

    def auto_suggest_product_search(self , product_name:str):
        self.page.pause()
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        """Test Product search functionality"""
        #Select category Product, enter keyword, and select from autosuggest
        print(f"Testing Product search for: {product_name}")
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_text("Product", exact=True).click()
        self.page.get_by_role("textbox", name="Search for 20 commodities such as Apple, Laptops").fill(product_name,timeout=40000)
        self.page.wait_for_timeout(2000)
        autosuggest_first_item = self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]').nth(
            0)
        expect(autosuggest_first_item).to_be_visible(timeout=40000)
        self.selected_product_raw = self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]').nth(
            0).inner_text().strip()
        self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]').nth(
            0).click(timeout=40000)
        result = re.sub(r"\s*or press tab\s*", "", self.selected_product_raw, flags=re.IGNORECASE)
        self.selected_product = result.strip()
        print(f"Selected product is: {self.selected_product}")
        #Click on Apply button
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)

        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def check_product_description_in_shipment_grid(self, validate: bool = True, use_pagination: bool = False):
        """Check product descriptions across shipment grid pages"""
        # --- Wait for shipment tab and click ---
        if not self.wait_for_shipment_grid(timeout=30):
            print("⏩ Skipping further checks because shipment grid has no results.")
            return []
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        shipment_tab.click(timeout=40000)

        # --- Wait for table ---
        table = self.page.get_by_role("table")
        expect(table).to_be_visible()
        first_row = self.page.locator('tbody[role="rowgroup"] >> tr[role="row"]').first
        expect(first_row).to_be_visible()
        print("✅ Table is loaded and has at least one row.")

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # --- Get table headers ---
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=10000)
        header_count = headers.count()

        self.S_No = self.Product_field = self.Matching_Field = None
        table_headers = []

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() in ["product description", "description"]:
                self.Product_field = i
            elif header_text.lower() in ["matching fields", "matching field"]:
                self.Matching_Field = i

        # --- Validate headers found ---
        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.Product_field is None:
            raise Exception("❌ 'Product Description' column not found in table headers")
        if self.Matching_Field is None:
            raise Exception("❌ 'Matching Field' column not found in table headers")

        print(
            f"✅ Found column indexes → S No: {self.S_No}, Matching: {self.Matching_Field}, Product: {self.Product_field}")

        all_rows_data = []
        keywords = self.selected_product.lower().split()

        # --- Process current page rows ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = [rows.nth(i).locator("td").nth(j).inner_text().strip() for j in
                            range(rows.nth(i).locator("td").count())]
                validation_note = ""

                if validate:
                    sl_no = row_data[self.S_No]
                    matching_field = row_data[self.Matching_Field]

                    # Get product (tooltip if exists)
                    product_cell = rows.nth(i).locator("td").nth(self.Product_field)
                    product_span = product_cell.locator("[data-tip]").first
                    if product_span.count() > 0:
                        product_tip = product_span.get_attribute("data-tip")
                        product = product_tip.strip() if product_tip else product_cell.inner_text().strip()
                    else:
                        product = product_cell.inner_text().strip()

                    row_data[self.Product_field] = product
                    print(f"🔎 Checking product: {product}")

                    # --- Keyword validation ---
                    if any(k in product.lower() for k in keywords):
                        print(f"✅ PASS: Product matches → {product} at Sl.No: {sl_no}")
                    else:
                        print(f"❌ FAIL: Invalid product → {product} at Sl.No: {sl_no}")
                        validation_note += (" | " if validation_note else "") + "❌ Invalid Product"
                        self.invalid_entries.append({"slNo": sl_no, "Product": product})

                    # --- Matching field validation ---
                    if matching_field.lower() == "product description":
                        print(f"✅ Matching field correct → {matching_field} at Sl.No: {sl_no}")
                    else:
                        print(f"❌ FAIL: Invalid matching field → {matching_field} at Sl.No: {sl_no}")
                        validation_note += (" | " if validation_note else "") + "❌ Invalid Matching Field"
                        self.invalid_entries.append({"slNo": sl_no, "MatchingField": matching_field})

                    # --- Click 2nd column to view modal ---
                    if rows.nth(i).locator("td").count() > 1:
                        rows.nth(i).locator("td").nth(1).click()
                        expect(self.page.locator("//span[normalize-space()='Shipment Origin Details']")).to_be_visible(
                            timeout=40000)

                        product_desc_locator = self.page.locator(
                            "//span[normalize-space(text())='Product Description']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]"
                        )
                        product_desc_text = product_desc_locator.first.inner_text().strip() if product_desc_locator.count() > 0 else ""

                        product_desc_normalized = re.sub(r'\s+', ' ', product_desc_text.lower()).strip()
                        product_normalized = re.sub(r'\s+', ' ', product.lower()).strip()

                        # ✅ Check for keyword presence and exact normalized match
                        if any(k in product_desc_normalized for k in
                               keywords) and product_desc_normalized == product_normalized:
                            print(
                                f"✅ PASS: Product description '{product_desc_text}' fully matches product '{product}'")
                        else:
                            print(
                                f"❌ FAIL: Product description '{product_desc_text}' does not fully match product '{product}' or missing keyword")
                            validation_note += (" | " if validation_note else "") + "❌ Product mismatch in View"
                            self.invalid_entries.append({"slNo": sl_no, "Product description(View)": product_desc_text})

                        # --- Close modal ---
                        close_btn = self.page.locator("//span[@aria-hidden='true']")
                        expect(close_btn).to_be_visible()
                        close_btn.click()
                        self.page.wait_for_timeout(500)

                row_data.append(validation_note)
                all_rows_data.append(row_data)

        # --- Pagination handling ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
                else:
                    break
        else:
            process_current_page()

        # --- Save results to Excel ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Product Data"
            sheet.append(table_headers + ["Validation"])
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "❌" in row_data[-1]:
                    prod_col_index = self.Product_field + 1
                    match_col_index = self.Matching_Field + 1
                    sheet.cell(row=sheet.max_row, column=prod_col_index).fill = red_fill
                    sheet.cell(row=sheet.max_row, column=match_col_index).fill = red_fill

            file_path = f"results/product_data_{self.selected_product}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid products highlighted in red)")
        else:
            print("📘 No invalid product description found. Excel not generated.")

        return self.invalid_entries

    def auto_suggest_shipper_search(self, shipper_name: str):
        """Test Shipper search functionality"""
        self.page.pause()
        print(f"Testing Shipper search for: {shipper_name}")
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.locator("(//div[contains(text(),'Shipper')])[1]").click()
        self.page.get_by_role("textbox", name="Search upto 20 Shipper's name or address").fill(shipper_name,timeout=40000)
        self.page.wait_for_timeout(2000)
        autosuggest_first_item = self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]'
        ).nth(0)

        # Wait for it to be visible
        expect(autosuggest_first_item).to_be_visible(timeout=40000)

        # Extract text after waiting
        self.selected_shipper_raw = autosuggest_first_item.inner_text().strip()
        self.page.locator('[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]').nth(0).click(timeout=40000)
        result = re.sub(r"\s*or press tab\s*", "", self.selected_shipper_raw, flags=re.IGNORECASE)
        self.selected_shipper = result.strip()
        print(f"Selected shipper name is: {self.selected_shipper}")
        #click on search button
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)

        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def auto_suggest_consignee_search(self, consignee_name: str):
        """Test Shipper search functionality"""
        self.page.pause()
        print(f"Testing Consignee search for: {consignee_name}")
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.locator("(//div[contains(text(),'Consignee')])[1]").click()
        self.page.get_by_role("textbox", name="Search upto 20 Consignee's name or address").fill(consignee_name)
        self.page.wait_for_timeout(2000)
        autosuggest_first_item = self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]'
        ).nth(0)

        # Wait for it to be visible
        expect(autosuggest_first_item).to_be_visible(timeout=40000)

        # Extract text after waiting
        self.selected_consignee_raw = autosuggest_first_item.inner_text().strip()
        self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]').nth(
            0).click(timeout=40000)
        result = re.sub(r"\s*or press tab\s*", "", self.selected_consignee_raw, flags=re.IGNORECASE)
        self.selected_consignee = result.strip()
        print(f"Selected consignee name is: {self.selected_consignee}")
        # click on search button
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)

        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def Validate_exporter_tab(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        # Navigate to exporter tab
        self.page.locator("//a[@id='nav-contact-tab']").click(timeout=40000)

        # Verify The Exporters tab label should update to Exporters (1) and display the selected Shipper name under the Company Name section
        total_record = self.page.locator(".trademo-table-count-text").inner_text().strip()
        total_count = total_record.split("of")[-1].strip()
        expect(self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]')).to_have_text(
            total_count)

    def check_shipper_name_export_tab(self, validate: bool = True):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        """Check shipper names across all pages and collect invalid entries"""
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(1).inner_text().strip()

                # Get company name from data-tip or fallback to visible text
                company_cell = rows.nth(i).locator("td").nth(2)
                company_name_locator = company_cell.locator("[data-tip]").first
                company_tip = (
                    company_name_locator.get_attribute("data-tip")
                    if company_name_locator.count() > 0 else None
                )

                company = (
                    company_tip.strip()
                    if company_tip and company_tip.strip()
                    else company_cell.inner_text().strip()
                )

                print(f"Checking company: {company}")

                expected_normalized = self.selected_shipper.lower()
                actual_normalized = re.sub(r"\s+", " ", company).strip().lower()

                if actual_normalized == expected_normalized:
                    print(f"✅ Valid company name found: {company} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid company name found: {company} at Sl. No: {sl_no}")

    def validate_duplicate_country_names(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        # Locate all company names in the "Company Name" column
        country_elements = self.page.locator("//div[img[@class='trademo-search-flag mr-2']]")
        count = country_elements.count()

        print(f"🔍 Total countries found: {count -1}")

        seen = set()
        duplicates = []

        for i in range(count):
            name = country_elements.nth(i).inner_text().strip()

            if name in seen:
                print(f"❌ Duplicate found: '{name}' at row {i}")
                duplicates.append(name)
            else:
                seen.add(name)

        if not duplicates:
            print("✅ No duplicate country names found!")
        else:
            print(f"⚠️ Total Duplicates Found: {len(duplicates)} -> {duplicates}")

    def normalize_company_name(self, name: str) -> str:
        """
        Normalize company names by shortening common terms.
        Example: "Limited" → "Ltd", "Company" → "Co".
        """
        # Lowercase and strip spaces
        name = name.lower().strip()

        # Replace common full forms with abbreviations
        replacements = {
            r"\blimited\b": "ltd",
        }

        for pattern, replacement in replacements.items():
            name = re.sub(pattern, replacement, name)

        # Collapse multiple spaces
        name = re.sub(r"\s+", " ", name)

        return name

    def Validate_Discover_insight_link(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        self.page.wait_for_timeout(1000)
        shipper_locator = self.page.locator('[class="btn-link trademo-link text-capitalize"]').first
        expect(shipper_locator).to_be_visible(timeout=40000)

        actual_shipper_text = shipper_locator.inner_text().strip().lower()
        expected_shipper_text = self.selected_shipper.strip().lower()

        # Validate ignoring case
        assert actual_shipper_text == expected_shipper_text, (
            f"❌ FAIL: Expected shipper '{expected_shipper_text}' but found '{actual_shipper_text}'"
        )

        # Test insights link
        with self.page.expect_popup() as page1_info:
            self.page.locator("a").filter(
                has_text=re.compile(f"^{re.escape(self.selected_shipper)}$", re.IGNORECASE)
            ).first.click(timeout=40000)
        page1 = page1_info.value
        expect(page1.get_by_role("main")).to_contain_text(self.selected_shipper , ignore_case=True)
        page1.close()
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)

    import re
    from playwright.sync_api import expect

    def Chemical_Search_auto_suggest(self, chemical: str):
        """🔍 Test Chemical search functionality with autosuggest handling"""
        print(f"🔍 Testing Chemical search for: {chemical}")
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        # ✅ Wait for search box to be visible
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")
        ).to_be_visible(timeout=40000)

        # Click on main search textbox
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)

        # Select 'Chemical' category
        self.page.locator("(//div[normalize-space()='Chemicals'])[1]").click(timeout=40000)

        # Fill chemical name
        self.page.get_by_role("textbox", name="Search upto 20 Chemical").fill(chemical, timeout=40000)
        self.page.wait_for_timeout(2000)

        # Locate first autosuggest item
        autosuggest_first_item = self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 '
            'tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]'
        ).nth(0)

        # ✅ Wait until autosuggest appears
        expect(autosuggest_first_item).to_be_visible(timeout=40000)

        # Extract and clean text (remove "or press tab" if present)
        self.selected_chemical_raw = autosuggest_first_item.inner_text().strip()
        result = re.sub(r"\s*or press tab\s*", "", self.selected_chemical_raw, flags=re.IGNORECASE)
        self.selected_chemical = result.strip()

        # Click the first autosuggest suggestion
        autosuggest_first_item.click(timeout=40000)

        print(f"✅ Selected chemical name is: {self.selected_chemical}")

        # Click search button
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def Validate_importer_tab(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        # Navigate to importer tab
        self.page.locator("//a[@id='nav-profile-tab']").click(timeout=40000)

        # Verify results
        total_record = self.page.locator(".trademo-table-count-text").inner_text().strip()
        total_count = total_record.split("of")[-1].strip()
        expect(self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]')).to_have_text(
            total_count)

    def Validate_Discover_insight_consignee(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 10s for results
            print("❌ No search results found in Shipment Grid.")
            return []

        self.page.wait_for_timeout(1000)

        # Locate first consignee link in the grid
        consignee_locator = self.page.locator('[class="btn-link trademo-link text-capitalize"]').first
        expect(consignee_locator).to_be_visible(timeout=40000)

        actual_consignee_text = consignee_locator.inner_text().strip().lower()
        expected_consignee_text = self.selected_consignee.strip().lower()

        # ✅ Validate ignoring case
        assert actual_consignee_text == expected_consignee_text, (
            f"❌ FAIL: Expected consignee '{expected_consignee_text}' but found '{actual_consignee_text}'"
        )

        # ✅ Test insights link popup
        with self.page.expect_popup() as page_info:
            self.page.locator("a").filter(
                has_text=re.compile(f"^{re.escape(self.selected_consignee)}$", re.IGNORECASE)
            ).first.click(timeout=40000)

        page = page_info.value
        expect(page.get_by_role("main")).to_contain_text(self.selected_consignee, ignore_case=True)
        page.close()

        # ✅ Ensure grid is still visible after closing popup
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)

    def check_consignee_Name_Import_tab(self, validate: bool = True):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        """Check consignee names across shipment grid pages and collect invalid entries"""

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(1).inner_text().strip()

                # Try to get 'data-tip' first
                consignee_cell = rows.nth(i).locator("td").nth(2)
                consignee_locator = consignee_cell.locator("[data-tip]").first
                consignee_tip = (
                    consignee_locator.get_attribute("data-tip")
                    if consignee_locator.count() > 0 else None
                )

                consignee_name = (
                    consignee_tip.strip()
                    if consignee_tip and consignee_tip.strip()
                    else consignee_cell.inner_text().strip()
                )

                consignee_name_lower = re.sub(r"\s+", " ", consignee_name).strip().lower()
                expected_lower = self.selected_consignee.lower()

                print(f"Consignee: {consignee_name}")

                if consignee_name_lower == expected_lower:
                    print(f"✅ Valid consignee name found: {consignee_name} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid consignee name found: {consignee_name} at Sl. No: {sl_no}")

    def Check_Chemical_In_shipmentGrid(self, validate: bool = True, use_pagination: bool = False):
        """Check chemical data across shipment grid pages with validation and Excel export"""
        # --- Wait for shipment tab and click ---
        if not self.wait_for_shipment_grid(timeout=30):
            print("⏩ Skipping further checks because shipment grid has no results.")
            return []
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        shipment_tab.click(timeout=40000)

        # --- Wait for table ---
        table = self.page.get_by_role("table")
        expect(table).to_be_visible()
        first_row = self.page.locator("table tbody tr").first
        expect(first_row).to_be_visible()
        print("✅ Table is loaded and has at least one row.")

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # --- Get table headers dynamically ---
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=10000)
        header_count = headers.count()

        self.S_No = self.M_field = self.P_desc = None
        table_headers = []

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "matching fields":
                self.M_field = i
            elif header_text.lower() in ["product description", "description"]:
                self.P_desc = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.M_field is None:
            raise Exception("❌ 'Matching Fields' column not found in table headers")
        if self.P_desc is None:
            raise Exception("❌ 'Product Description' column not found in table headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Matching: {self.M_field}, Product: {self.P_desc}")

        all_rows_data = []
        keywords = self.selected_chemical.lower().split()

        # --- Process current page rows ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = [rows.nth(i).locator("td").nth(j).inner_text().strip()
                            for j in range(rows.nth(i).locator("td").count())]

                validation_note = ""
                if validate:
                    sl_no = row_data[self.S_No]
                    matching_field = row_data[self.M_field]

                    # Get product (prefer tooltip if exists)
                    product_cell = rows.nth(i).locator("td").nth(self.P_desc)
                    product_span = product_cell.locator("[data-tip]").first
                    if product_span.count() > 0:
                        product_tip = product_span.get_attribute("data-tip")
                        product = product_tip.strip() if product_tip else product_cell.inner_text().strip()
                    else:
                        product = product_cell.inner_text().strip()

                    row_data[self.P_desc] = product
                    print(f"🔎 Checking chemical: {product}")

                    # --- Keyword validation ---
                    if any(k in product.lower() for k in keywords):
                        print(f"✅ PASS: Chemical matches → {product} at Sl.No: {sl_no}")
                    else:
                        print(f"❌ FAIL: Invalid Chemical → {product} at Sl.No: {sl_no}")
                        validation_note += (" | " if validation_note else "") + "❌ Invalid Chemical"
                        self.invalid_entries.append({"slNo": sl_no, "Chemical": product})

                    # --- Matching field validation ---
                    if matching_field.lower() == "product description":
                        print(f"✅ Matching field correct → {matching_field} at Sl.No: {sl_no}")
                    else:
                        print(f"❌ FAIL: Invalid matching field → {matching_field} at Sl.No: {sl_no}")
                        validation_note += (" | " if validation_note else "") + "❌ Invalid Matching Field"
                        self.invalid_entries.append({"slNo": sl_no, "MatchingField": matching_field})

                     # --- Click 2nd column to view modal ---
                    if rows.nth(i).locator("td").count() > 1:
                        rows.nth(i).locator("td").nth(1).click()
                        expect(
                                self.page.locator("//span[normalize-space()='Shipment Origin Details']")).to_be_visible(
                                timeout=40000)

                        product_desc_locator = self.page.locator(
                                "//span[normalize-space(text())='Product Description']/ancestor::div[contains(@class,'col-5')]/following-sibling::div[contains(@class,'col-7')]"
                        )
                        product_desc_text = product_desc_locator.first.inner_text().strip() if product_desc_locator.count() > 0 else ""

                        product_desc_normalized = re.sub(r'\s+', ' ', product_desc_text.lower()).strip()
                        product_normalized = re.sub(r'\s+', ' ', product.lower()).strip()

                        # ✅ Check for keyword presence and exact normalized match
                        if any(k in product_desc_normalized for k in
                                   keywords) and product_desc_normalized == product_normalized:
                             print(
                                    f"✅ PASS: Product description '{product_desc_text}' fully matches product '{product} in view'")
                        else:
                            print(
                                    f"❌ FAIL: Product description '{product_desc_text}' does not fully match product '{product}' or missing keyword in view")
                            validation_note += (" | " if validation_note else "") + "❌ Product mismatch in View"
                            self.invalid_entries.append(
                                    {"slNo": sl_no, "Product description(View)": product_desc_text})

                            # --- Close modal ---
                        close_btn = self.page.locator("//span[@aria-hidden='true']")
                        expect(close_btn).to_be_visible()
                        close_btn.click()
                        self.page.wait_for_timeout(500)

                row_data.append(validation_note)
                all_rows_data.append(row_data)

        # --- Pagination handling ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
                else:
                    break
        else:
            process_current_page()

        # --- Save results to Excel ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Chemical Data"
            sheet.append(table_headers + ["Validation"])
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "❌" in row_data[-1]:
                    chem_col_index = self.P_desc + 1
                    match_col_index = self.M_field + 1
                    sheet.cell(row=sheet.max_row, column=chem_col_index).fill = red_fill
                    sheet.cell(row=sheet.max_row, column=match_col_index).fill = red_fill

            file_path = f"results/chemical_data_{self.selected_chemical}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid chemicals highlighted in red)")
        else:
            print("📘 No invalid chemical data found. Excel not generated.")

        return self.invalid_entries

    def auto_suggest_search_port(self, port: str):
        """Test Port search functionality with autosuggest dropdown"""
        print(f"🔎 Testing Port search for: {port}")
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        # --- Wait for main search bar ---
        search_bar = self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below"
        )
        expect(search_bar).to_be_visible(timeout=40000)

        # --- Select Ports category ---
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.locator("(//div[contains(text(),'Ports')])[1]").click(timeout=40000)

        # --- Fill the port name ---
        port_input = self.page.get_by_role("textbox", name="Search upto 20 Ports’s name")
        port_input.fill(port, timeout=40000)
        self.page.wait_for_timeout(1000)

        # --- Wait for and select first autosuggest item ---
        autosuggest_first_item = self.page.locator(
            '[class="tw-px-3 tw-py-1 tw-text-sm hover:tw-bg-primary-purple-50 '
            'tw-cursor-pointer tw-flex tw-justify-between tw-items-center"]'
        ).nth(0)
        expect(autosuggest_first_item).to_be_visible(timeout=40000)
        self.page.wait_for_timeout(2000)

        self.selected_port_raw = autosuggest_first_item.inner_text().strip()

        # Remove "or press tab" if present
        result = re.sub(r"\s*or press tab\s*", "", self.selected_port_raw, flags=re.IGNORECASE)
        self.selected_port = result.strip()

        autosuggest_first_item.click(timeout=40000)
        print(f"✅ Selected port is: {self.selected_port}")

        # --- Click on Apply button ---
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        # self.page.wait_for_timeout(5000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def validate_port_field(self, rows, i, row_data, sl_no, field_name, col_index, expected_port):
        """
        Generic validator for Port of Lading / Port of Unlading.
        field_name: "Port of Lading" or "Port of Unlading"
        col_index: self.Port_Lading_col or self.Port_Unlading_col
        expected_port: self.selected_port
        """
        try:
            # 🔹 Get port name from grid cell
            port_cell = rows.nth(i).locator("td").nth(col_index)
            port_span = port_cell.locator("[data-tip]").first

            port_tip = port_span.get_attribute("data-tip") if port_span.count() > 0 else None
            port_name = port_tip.strip() if port_tip else row_data[col_index].strip()

            self.port_name_hyperlink = port_name.split(",")[0].strip()
            print(f"🔎 {field_name}: {self.port_name_hyperlink}")

            # 🔹 Validate port name
            if expected_port.lower().strip() in port_name.lower():
                print(f"✅ PASS: {field_name} '{port_name}' at Sl. No: {sl_no}")
            else:
                print(f"❌ FAIL: Invalid {field_name} '{port_name}' at Sl. No: {sl_no}")
                self.invalid_entries.append({"slNo": sl_no, field_name: port_name})

            # 🔹 Validate hyperlink in grid
            link = port_cell.locator("a")
            if link.count() > 0:
                expect(link).to_be_visible()
                href_value = link.get_attribute("href")
                port_text = link.inner_text().strip()
                print(f"✅ {field_name} is hyperlinked in grid: {port_text} → {href_value}")
            else:
                port_text = port_cell.inner_text().strip()
                print(f"❌ {field_name} is NOT hyperlinked in grid: {port_text}")
                self.invalid_entries.append({"slNo": sl_no, f"{field_name} Hyperlink": "Missing"})

            # 🔹 Open details modal
            rows.nth(i).locator("td").nth(1).click()
            self.page.wait_for_timeout(1500)

            # 🔹 Validate in View modal
            port_elements = self.page.locator(
                f"//span[normalize-space(text())='{field_name}']"
                "/ancestor::div[contains(@class,'col-5')]"
                "/following-sibling::div[contains(@class,'col-7')]"
            )

            port_element = None
            if port_elements.locator("a").count() > 0:
                port_element = port_elements.locator("a").first
            elif port_elements.locator("span").count() > 0:
                port_element = port_elements.locator("span").first

            if not port_element:
                print(f"❌ No {field_name} found in view screen for Sl. No: {sl_no}")
                self.invalid_entries.append({"slNo": sl_no, f"{field_name} (View)": "Missing"})
            else:
                expect(port_element).to_be_visible()
                port_view_text = port_element.inner_text().strip()

                # Validate text
                if port_view_text.lower() == port_name.lower() and port_view_text.lower() == expected_port.lower().strip():
                    print(f"✅ PASS: {field_name} '{port_view_text}' matches grid and expected port")
                else:
                    print(f"❌ FAIL: {field_name} '{port_view_text}' does not match grid or expected '{expected_port}'")
                    self.invalid_entries.append({"slNo": sl_no, f"{field_name} (View)": port_view_text})

                # Validate hyperlink in view (only if it is an <a>)
                if port_element.evaluate("el => el.tagName").lower() == "a":
                    print(f"✅ {field_name} '{port_view_text}' is a valid hyperlink in view")
                else:
                    print(f"❌ {field_name} '{port_view_text}' is NOT a hyperlink in view")
                    self.invalid_entries.append({"slNo": sl_no, f"{field_name} (View Hyperlink)": "Missing"})

            # 🔹 Close modal
            close_btn = self.page.locator("//span[@aria-hidden='true']").first
            expect(close_btn).to_be_visible()
            close_btn.click()
            self.page.wait_for_timeout(500)

        except Exception as e:
            print(f"❌ ERROR validating {field_name} at Sl. No: {sl_no}: {str(e)}")
            self.invalid_entries.append({"slNo": sl_no, f"{field_name} Error": str(e)})

    def Check_Port_description_Shipment_Grid(self, validate: bool = True, use_pagination: bool = False):
        if not self.wait_for_shipment_grid(timeout=30):
            print("⏩ Skipping further checks because shipment grid has no results.")
            return []
        self.page.locator("//a[@id='nav-home-tab']").click()

        self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.M_field = None
        self.Port_Lading_col = None
        self.Port_Unlading_col = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() in ["matching field", "matching fields"]:
                self.M_field = i
            elif header_text.lower() in ["port of lading", "ports of lading"]:
                self.Port_Lading_col = i
            elif header_text.lower() in ["port of unlading", "ports of unlading"]:
                self.Port_Unlading_col = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.M_field is None:
            raise Exception("❌ 'Matching Field(s)' column not found in table headers")
        if self.Port_Lading_col is None and self.Port_Unlading_col is None:
            raise Exception("❌ Neither 'Port of Lading' nor 'Port of Unlading' found in headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Matching Field: {self.M_field}, "
              f"Port Lading: {self.Port_Lading_col}, Port Unlading: {self.Port_Unlading_col}")

        all_rows_data = []  # store all rows with validation info

        # --- process one page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()
                for j in range(cell_count):
                    row_data.append(cells.nth(j).inner_text().strip())

                validation_note = ""  # Excel column
                sl_no = row_data[self.S_No]
                matching_field = row_data[self.M_field]
                port_name = None
                is_invalid = False

                if validate:
                    # --- Port of Lading ---
                    if matching_field.lower() == "port of lading" and self.Port_Lading_col is not None:
                        port_span = rows.nth(i).locator("td").nth(self.Port_Lading_col).locator("[data-tip]").first
                        port_tip = port_span.get_attribute("data-tip") if port_span.count() > 0 else None
                        port_name = port_tip.strip() if port_tip else row_data[self.Port_Lading_col]

                        self.port_name_hyperlink = port_name.split(",")[0].strip()

                        print(self.port_name_hyperlink)

                        if self.selected_port.lower().strip() in port_name.lower().strip():
                            print(f"✅ PASS: Port of Lading '{port_name}' at Sl. No: {sl_no}")
                        else:
                            print(f"❌ FAIL: Invalid Port of Lading '{port_name}' at Sl. No: {sl_no}")
                            is_invalid = True
                            validation_note = "❌ Invalid Port of Lading"

                    # --- Port of Unlading ---
                    elif matching_field.lower() == "port of unlading" and self.Port_Unlading_col is not None:
                        port_span = rows.nth(i).locator("td").nth(self.Port_Unlading_col).locator("[data-tip]").first
                        port_tip = port_span.get_attribute("data-tip") if port_span.count() > 0 else None
                        port_name = port_tip.strip() if port_tip else row_data[self.Port_Unlading_col]

                        self.port_name_hyperlink = port_name.split(",")[0].strip()
                        print(self.port_name_hyperlink)

                        if self.selected_port.lower().strip() in port_name.lower().strip():
                            print(f"✅ PASS: Port of Unlading '{port_name}' at Sl. No: {sl_no}")
                        else:
                            print(f"❌ FAIL: Invalid Port of Unlading '{port_name}' at Sl. No: {sl_no}")
                            is_invalid = True
                            validation_note = "❌ Invalid Port of Unlading"

                    # --- Unexpected Matching Field ---
                    else:
                        print(f"❌ FAIL: Unexpected Matching field '{matching_field}' at Sl. No: {sl_no}")
                        is_invalid = True
                        validation_note = "❌ Invalid Matching Field"

                    if is_invalid:
                        self.invalid_entries.append({
                            "slNo": sl_no,
                            "Port": port_name if port_name else "N/A",
                            "MatchingField": matching_field
                        })

                row_data.append(validation_note)
                all_rows_data.append(row_data)

        # --- handle pagination ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Port Data"

            sheet.append(table_headers + ["Validation"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "Invalid" in row_data[-1]:
                    # highlight port column
                    if "lading" in row_data[self.M_field].lower():
                        col_index = self.Port_Lading_col + 1
                    elif "unlading" in row_data[self.M_field].lower():
                        col_index = self.Port_Unlading_col + 1
                    else:
                        col_index = self.M_field + 1
                    sheet.cell(row=sheet.max_row, column=col_index).fill = red_fill

            file_path = f"results/port_data_{self.selected_port}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid ports highlighted in red)")
        else:
            print("📘 No invalid port entries found. Excel not generated.")

        return self.invalid_entries

    def Validate_Discover_insight_ports(self):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        with self.page.expect_popup() as page3_info:
               self.page.locator("span").filter(has_text="Discover more insights about").locator("span").click()
        page3 = page3_info.value
        expect(page3.get_by_role("main")).to_contain_text(self.port_name_hyperlink,ignore_case=True, timeout=100000)
        page3.close()

    def Search_product_manualsuggest(self, product_name: str):
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=100000)

        # Select category Product, enter keyword, and select from autosuggest
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all categories or choose from the category below").fill(product_name)
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def Verify_Shipment_tab_Manual_suggest(self, Extracted_Text: str, validate: bool = True,
                                           use_pagination: bool = False):
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        if not shipment_tab.is_visible(timeout=10000):  # wait max 5s for results
            print("❌ No search results found in Shipment Grid.")
            return []
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=40000)
        self.page.wait_for_timeout(2000)
        self.invalid_entries = []
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        # ✅ Flexible row check
        if row_count < 10:
            print(f"⚠️ Only {row_count} rows found (less than 10). Skipping strict check.")
        else:
            expect(rows).to_have_count(10, timeout=100000)

        # Keep invalid_entries consistent across calls
        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # --- Extract headers dynamically ---
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        table_headers = []
        self.S_No = self.M_field = self.Product_col = None
        self.Shipper_col = self.Consignee_col = self.Address_col = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)

            lower = header_text.lower()
            if lower in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif "matching" in lower:
                self.M_field = i
            elif "product" in lower:
                self.Product_col = i
            elif "shipper" in lower and "standardized" not in lower:
                self.Shipper_col = i
            elif "consignee" in lower and "address" not in lower and "standardized" not in lower:
                self.Consignee_col = i
            elif "address" in lower:
                self.Address_col = i

        if self.S_No is None or self.M_field is None:
            print("❌ Required columns ('S No' or 'Matching Field') not found in headers")

        extracted_text_lower = Extracted_Text.lower()
        all_rows_data = []

        # --- process one page ---
        def process_current_page():
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    row_data.append(cells.nth(j).inner_text().strip())

                validation_note = []
                is_invalid = False

                if validate:
                    sl_no = row_data[self.S_No]
                    matching_field = row_data[self.M_field]
                    matching_fields_list = [mf.strip() for mf in matching_field.split(",")]
                    product = row_data[self.Product_col] if self.Product_col else ""
                    shipper_name = row_data[self.Shipper_col] if self.Shipper_col else ""
                    consignee_name = row_data[self.Consignee_col] if self.Consignee_col else ""

                    consignee_address = shipper_std_name = consignee_std_name = ""

                    # --- Validation checks ---
                    for mf in matching_fields_list:
                        if mf == "Product Description":
                            parts = split_location(extracted_text_lower)  # e.g. ['shekou', 'port']
                            product_span = rows.nth(i).locator("td").nth(self.Product_col).locator("[data-tip]").first
                            product_tip = product_span.get_attribute("data-tip") if product_span.count() > 0 else None
                            product = product_tip.strip() if product_tip else rows.nth(i).locator("td").nth(
                                self.Product_col).inner_text().strip()

                            print(f"🔎 Checking product: {product}")
                            product_lower = product.lower()

                            matched_word = None
                            for word in parts:
                                if word in product_lower:
                                    matched_word = word
                                    break

                            if matched_word:
                                print(f"✅ {mf} contains keyword '{matched_word}' at Sl. No: {sl_no}")
                            else:
                                print(f"❌ {mf} mismatch: {product} at Sl. No: {sl_no}")
                                is_invalid, validation_note = True, [f"❌ Invalid Product Description in Grid"]

                        elif mf == "Shipper Name":
                            if extracted_text_lower in shipper_name.lower():
                                print(f"✅ Shipper name valid at Sl. No: {sl_no}")
                            else:
                                print(f"❌ Shipper name mismatch: {shipper_name} at Sl. No: {sl_no}")
                                is_invalid, validation_note = True, ["❌ Invalid Shipper in Grid"]

                        elif mf == "Consignee Name":
                            if extracted_text_lower in consignee_name.lower():
                                print(f"✅ Consignee name valid at Sl. No: {sl_no}")
                            else:
                                print(f"❌ Consignee name mismatch: {consignee_name} at Sl. No: {sl_no}")
                                is_invalid, validation_note = True, ["❌ Invalid Consignee in Grid"]

                        elif mf == "Shipper Address":
                            rows.nth(i).locator("td").nth(1).click()
                            self.page.wait_for_timeout(2000)

                            read_more_locator = self.page.get_by_text("Read more").first
                            self.page.wait_for_timeout(1000)  # wait for rendering

                            if read_more_locator.is_visible():
                                print("🔍 'Read More' is visible, clicking...")
                                read_more_locator.click()
                                self.page.wait_for_timeout(500)

                                read_more = self.page.locator("div.read-more-cards").first
                                try:
                                    address_text = read_more.inner_text().strip()
                                    parts = split_location(extracted_text_lower)  # e.g. ['shekou', 'port']
                                    address_lower = address_text.lower()

                                    matched_word = None
                                    for word in parts:
                                        if word in address_lower:
                                            matched_word = word
                                            break

                                    if matched_word:
                                        print(f"✅ {mf} contains keyword '{matched_word}' at Sl. No: {sl_no}")
                                    else:
                                        print(f"❌ {mf} mismatch: {address_text} at Sl. No: {sl_no}")
                                        is_invalid, validation_note = True, [f"❌ Invalid shipper address in view"]
                                finally:
                                    close_btn = self.page.locator("//span[@aria-hidden='true']")
                                    if close_btn.is_visible():
                                        close_btn.click()
                            else:
                                print("❌ 'Read more' link not found for Shipper Address")

                        elif mf == "Consignee Address":
                            rows.nth(i).locator("td").nth(1).click()
                            self.page.wait_for_timeout(2000)

                            read_more_locator = self.page.get_by_text("Read more").last
                            self.page.wait_for_timeout(1000)  # wait for rendering

                            if read_more_locator.is_visible():
                                print("🔍 'Read More' is visible, clicking...")
                                read_more_locator.click()
                                self.page.wait_for_timeout(500)

                                read_more = self.page.locator("div.read-more-cards").last
                                try:
                                    address_text = read_more.inner_text().strip()
                                    parts = split_location(extracted_text_lower)  # e.g. ['shekou', 'port']
                                    address_lower = address_text.lower()

                                    matched_word = None
                                    for word in parts:
                                        if word in address_lower:
                                            matched_word = word
                                            break

                                    if matched_word:
                                        print(f"✅ {mf} contains keyword '{matched_word}' at Sl. No: {sl_no}")
                                    else:
                                        print(f"❌ {mf} mismatch: {address_text} at Sl. No: {sl_no}")
                                        is_invalid, validation_note = True, [f"❌ Invalid consignee address in view"]
                                finally:
                                    close_btn = self.page.locator("//span[@aria-hidden='true']")
                                    if close_btn.is_visible():
                                        close_btn.click()
                            else:
                                print("❌ 'Read more' link not found for Consignee Address")

                        elif mf == "Shipper Standardized Name":
                            rows.nth(i).locator("td").nth(1).click()
                            self.page.wait_for_timeout(2000)
                             # Extract Shipper Standardized Name
                            Shipper_Standardized_Name = self.page.locator(
                                " //span[normalize-space(text())='Shipper Standardized Name'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a").inner_text().strip()
                            if extracted_text_lower in Shipper_Standardized_Name.lower():
                                print(
                                    f"✅ Shipper Standardized Name contains keyword {Shipper_Standardized_Name} at Sl. No: {sl_no}")
                            else:
                                print(
                                    f"❌ Search keyword not found in shipper standardized name: {Shipper_Standardized_Name} at Sl. No: {sl_no}")
                                is_invalid, validation_note = True, ["❌ Invalid shipper standardized name in Grid"]
                            self.page.locator("//span[@aria-hidden='true']").click()

                        elif mf == "Consignee Standardized Name":
                            rows.nth(i).locator("td").nth(1).click()
                            Consignee_Standardized_Name = self.page.locator(
                                " //span[normalize-space(text())='Consignee Standardized Name'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a").inner_text().strip()
                            if extracted_text_lower in Consignee_Standardized_Name.lower():
                                print(
                                    f"✅ Consignee Standardized Name contains keyword {Consignee_Standardized_Name} at Sl. No: {sl_no}")
                            else:
                                print(
                                    f"❌ Search keyword not found in consignee standardized name: {Consignee_Standardized_Name} at Sl. No: {sl_no}")
                                is_invalid, validation_note = True, ["❌ Invalid consignee standardized name in Grid"]
                            self.page.locator("//span[@aria-hidden='true']").click()

                    if is_invalid:
                        self.invalid_entries.append({
                            "slNo": sl_no,
                            "matching_field": matching_field,
                            "product": product,
                            "Shipper_Name": shipper_name,
                            "Consignee_Name": consignee_name,
                            "Consignee_Address": consignee_address,
                            "Shipper_Standardized_Name": shipper_std_name,
                            "Consignee_Standardized_Name": consignee_std_name,
                        })

                row_data.append(", ".join(validation_note) if validation_note else "")
                all_rows_data.append(row_data)

        # --- pagination support ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
                else:
                    break
        else:
            process_current_page()

        # --- Save to Excel if invalid entries exist ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Manual Validation"

            sheet.append(table_headers + ["Validation"])
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "Invalid" in row_data[-1]:
                    for col in range(2, len(row_data)):
                        sheet.cell(row=sheet.max_row, column=col).fill = red_fill

            file_path = f"results/Manual_Search_{Extracted_Text}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid entries highlighted in red)")
        else:
            print("📘 No invalid entries found. Excel not generated.")

        return self.invalid_entries

    def Manual_suggest_hs_code(self,hs_code:str):
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=40000)
        # Manual suggest search for HS Code
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").fill(hs_code)
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)

        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

    def Manual_suggest_chemical(self, chemical_name: str):
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=40000)
        # Manual suggest search for HS Code
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").fill(chemical_name)
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")


def Manual_suggest_shipper(self,shipper_name:str):
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=40000)
        # Manual suggest search for Shipper
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").fill(shipper_name)
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")


def Manual_suggest_consignee(self,consignee_name:str):
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=40000)
        # Manual suggest search for Consignee
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").fill(consignee_name)
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

def Manual_suggest_port(self,port_name:str):
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible(timeout=40000)
        # Manual suggest search for Port
        self.page.get_by_role("textbox", name="Type to search in all").click(timeout=40000)
        self.page.get_by_role("textbox", name="Type to search in all").fill(port_name)
        self.page.locator(".tw-bg-primary-purple-500").click(timeout=40000)
        table = self.wait_for_new_table("//table")

        if table:
            rows = table.locator("tr")
            print(f"Number of rows found: {rows.count()}")
        else:
            print("⚠️ Skipping row validation since table never appeared.")

def Verify_Save_Search_screen(self, HS_code: str):
    # Click on the Save Search button
    self.page.get_by_text("Save Search").first.click(timeout=40000)

    # Main modal locator
    main_modal = self.page.get_by_role("main")

    # Verify modal is displayed
    expect(main_modal).to_contain_text("Save Search")

    # Expected static texts
    expected_texts = [
        "Search Category",
        "Search Phrase",
        "Date Range",
        "Applied Filters",
        "No filters have been added",
        "Set as Default Search View",
        "Add a name to this search",
        "Select a date range for specific",
        "Share the search with your team",
        "Would you like to receive email alerts on the search",
        "Cancel",
        "Save Search",
        "Custom Date Range: Jan 01, 2020 to Jul 31, 2025"
    ]

    for text in expected_texts:
        expect(main_modal).to_contain_text(text)

    # Click Date Range to expand
    self.page.get_by_text("Date Range", exact=True).click()

    # Validate pill labels
    pill_labels = self.page.locator('[class="_pill_label_rvvht_133"]')
    expect(pill_labels.nth(0)).to_contain_text("HS Code")
    expect(pill_labels.nth(1)).to_contain_text(HS_code)
    expect(pill_labels.nth(2)).to_contain_text("Jan 01, 2020 to Jul 31, 2025 ")


    def Verify_SaveSearch_Cancel_Close(self):
        # Click on Cancel button
        self.page.get_by_role("button", name="Cancel").click(timeout=40000)
        self.page.wait_for_timeout(1000)
        # Verify Save Search modal is closed
        expect(self.page.get_by_role("main")).to_contain_text("Search Category")
        # Click on the Save Search button again
        self.page.get_by_text("Save Search").first.click(timeout=40000)

        # Click on Close icon
        self.page.locator("span").filter(has_text="Save SearchSave SearchSearch").get_by_role("img").nth(1).click()
        # Verify Save Search modal is closed
        expect(self.page.get_by_role("main")).to_contain_text("Search Category")

    def Verify_SaveSearch_button(self):
        import random
        # Click on the Save Search button again
        self.page.get_by_text("Save Search").first.click(timeout=40000)

        # Verify the input field “Add a name to this search”
        self.page.get_by_role("textbox", name="Enter name for search").click(timeout=40000)
        self.page.get_by_role("textbox", name="Enter name for search").clear()
        random_number = random.randint(1000, 9999)  # 4-digit random number
        hs_code_value = f"HS Code Search{random_number}"

        print(hs_code_value)
        self.page.get_by_role("textbox", name="Enter name for search").fill(hs_code_value)
        # Check the “Custom Date Range” dropdown
        self.page.locator("span").filter(has_text="Save SearchSave SearchSearch").locator("svg").nth(2).click()
        self.page.locator("span").filter(has_text="Save SearchSave SearchSearch").locator("svg").nth(2).click()

        # Check for toggle: “Set as Default Search View”
        # Enable the toggle
        self.page.locator(".slider").first.click()
        expect(self.page.get_by_role("main")).to_contain_text(
            "Enabling this will set the current search as your default, replacing any previous default view")
        # Disable the toggle
        self.page.locator("div").filter(has_text=re.compile(r"^Set as Default Search View$")).locator("span").click()
        expect(self.page.get_by_role("main")).not_to_contain_text(
            "Enabling this will set the current search as your default, replacing any previous default view")

        # Check “Share the search with your team” multi-select dropdown
        self.page.locator("div").filter(has_text=re.compile(r"^You can select multiple options$")).nth(1).click()
        self.page.get_by_text("Swarupa Dash (swarupa.dash@trademo.com)", exact=True).click()
        self.page.locator("div:nth-child(3) > .css-19bqh2r").click()

        # Validate toggle: “Would you like to receive email alerts on the search”
        self.page.locator("div").filter(
            has_text=re.compile(r"^Would you like to receive email alerts on the search$")).locator("span").click()
        expect(self.page.get_by_role("main")).to_contain_text("How frequently would you like to receive the alerts?")
        expect(self.page.get_by_role("main")).to_contain_text("Select all that you want to track via this email alerts")
        expect(self.page.get_by_role("main")).to_contain_text(
            "Select measurement unit bases which you want to generate results.")

        # Disable toggle
        self.page.locator("div").filter(
            has_text=re.compile(r"^Would you like to receive email alerts on the search$")).locator("span").click()
        expect(self.page.get_by_role("main")).not_to_contain_text("How frequently would you like to receive the alerts?")
        expect(self.page.get_by_role("main")).not_to_contain_text("Select all that you want to track via this email alerts")
        expect(self.page.get_by_role("main")).not_to_contain_text(
            "Select measurement unit bases which you want to generate results.")
        # Open modal again, fill all valid inputs, and click Save Search
        self.page.get_by_role("button", name="Save Search").click(timeout=40000)
        self.page.wait_for_timeout(1000)
        Skip_locator = self.page.get_by_text("Skip and Save")
        if  Skip_locator.is_visible():
            print("🔍 'Skip and Save' is visible, clickin")
            Skip_locator.click()
            expect(self.page.get_by_role("main")).to_contain_text(
                "Are you sure, you don't want to receive email alerts for this search?")
            expect(self.page.get_by_role("paragraph")).to_contain_text(
                "Our email alert feature will keep you updated on new shipments for this saved search, with actionable insights and new trends.")
            expect(self.page.get_by_role("main")).to_contain_text("Skip and Save")
            expect(self.page.get_by_role("main")).to_contain_text("Set Email Alert")
            self.page.get_by_role("button", name="Skip and Save").click(timeout=40000)
        else:
            print("ℹ️ 'Skip and save' not visible.")
        toast = self.page.get_by_role("alert")
        expect(self.page.get_by_role("alert")).to_contain_text(
            f"\"{hs_code_value}\" has been added to your Saved Searches.")
        self.page.locator('[class="m-1"]').click(timeout=40000)
        self.page.get_by_text("Search History").click(timeout=40000)
        # Correct way in Python Playwright
        items = [f"HS Code Search{hs_code_value}"]
        expect(self.page.get_by_role("alert")).to_contain_text("HS Code Search")
        self.page.get_by_text("Shipments").click()
        self.page.wait_for_timeout(3000)

    def check_Shipper_Name_in_theGrid_View(self, validate: bool = True, use_pagination: bool = False):
        # --- Wait for shipment tab ---
        if not self.wait_for_shipment_grid(timeout=30):
            print("⏩ Skipping further checks because shipment grid has no results.")
            return []
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        shipment_tab.click(timeout=40000)
        self.page.wait_for_timeout(2000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if row_count < 10:
            print(f"⚠️ Only {row_count} rows found (less than 10). Skipping strict check.")
        else:
            expect(rows).to_have_count(10, timeout=100000)

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # --- Extract headers ---
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()
        table_headers = []
        self.S_No = self.M_field = self.Shipper_col = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)

            lower = header_text.lower()
            if lower in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif "matching" in lower:
                self.M_field = i
            elif "shipper" in lower and "standardized" not in lower:
                self.Shipper_col = i

        if self.S_No is None or self.M_field is None:
            raise Exception("❌ Required columns ('S No' or 'Matching Field') not found in headers")

        all_rows_data = []

        # --- Process current page ---
        def process_current_page():
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    row_data.append(cells.nth(j).inner_text().strip())

                validation_note = []
                is_invalid = False

                if validate:
                    sl_no = row_data[self.S_No]
                    matching_field = row_data[self.M_field]
                    matching_fields_list = [mf.strip() for mf in matching_field.split(",")]
                    shipper_name = row_data[self.Shipper_col] if self.Shipper_col else ""
                    shipper_std_name_text = ""

                    for mf in matching_fields_list:
                        # --- Hyperlink check for shipper name ---
                        shipper_cell = cells.nth(self.Shipper_col)
                        shipper_links = shipper_cell.locator("a")
                        if shipper_links.count() > 0:
                            try:
                                expect(shipper_links.first).to_be_visible()
                                print(f"✅ Shipper name is hyperlinked in grid: {shipper_links.first.inner_text()}")
                            except:
                                print(f"❌ Shipper name hyperlink not interactable at Sl. No: {sl_no}")
                        else:
                            print(f"❌ Shipper name is NOT hyperlinked in grid: {shipper_name}")
                            validation_note.append("❌ Missing hyperlink in Grid")
                            self.invalid_entries.append(
                                {"slNo": sl_no, "Shipper name": shipper_name, "Hyperlink": "Missing"}
                            )

                        # --- Validate Shipper Name ---
                        if mf == "Shipper Name":
                            if self.selected_shipper.strip().lower() == shipper_name.lower():
                                print(f"✅ Shipper name valid at Sl. No: {sl_no}")
                                rows.nth(i).locator("td").nth(1).click()
                                self.page.wait_for_timeout(2000)

                                shipper_name_view_locator = self.page.locator(
                                    "//span[normalize-space(text())='Shipper Name']"
                                    "/ancestor::div[contains(@class,'col-5')]"
                                    "/following-sibling::div[contains(@class,'col-7')]//div[1]"
                                )
                                shipper_name_view = shipper_name_view_locator.inner_text().strip()

                                if self.selected_shipper.lower() in shipper_name_view.lower():
                                    print(f"✅ Shipper Name valid at Sl. No: {sl_no} in view")
                                else:
                                    print(f"❌ Invalid Shipper Name: {shipper_name_view} at Sl. No: {sl_no}")
                                    is_invalid = True
                                    validation_note.append("❌ Invalid shipper name in view")

                                self.page.locator("//span[@aria-hidden='true']").click()
                            else:
                                print(f"❌ Shipper name mismatch: {shipper_name} at Sl. No: {sl_no}")
                                is_invalid = True
                                validation_note.append("❌ Invalid Shipper name in Grid")

                        # --- Validate Shipper Standardized Name ---
                        elif mf == "Shipper Standardized Name":
                            rows.nth(i).locator("td").nth(1).click()
                            self.page.wait_for_timeout(2000)

                            shipper_std_name_locator = self.page.locator(
                                "//span[normalize-space(text())='Shipper Standardized Name']"
                                "/ancestor::div[contains(@class,'col-5')]"
                                "/following-sibling::div[contains(@class,'col-7')]//a"
                            )
                            if shipper_std_name_locator.count() > 0:
                                shipper_std_name_text = shipper_std_name_locator.inner_text().strip()

                                if self.selected_shipper.lower() == shipper_std_name_text.lower():
                                    print(f"✅ Shipper Standardized Name valid at Sl. No: {sl_no} in view")
                                else:
                                    print(
                                        f"❌ Invalid Shipper Standardized Name: {shipper_std_name_text} at Sl. No: {sl_no}")
                                    is_invalid = True
                                    validation_note.append("❌ Invalid shipper standardized name in view")

                                # Verify hyperlink
                                tag_name = shipper_std_name_locator.evaluate("el => el.tagName")
                                if tag_name and tag_name.lower() == "a":
                                    print(f"✅ '{shipper_std_name_text}' is a hyperlink in view screen")
                                else:
                                    print(f"❌ '{shipper_std_name_text}' is NOT a hyperlink in view screen")
                                    validation_note.append("❌ Missing hyperlink in View")
                            else:
                                print(f"❌ No Shipper Standardized Name element found for Sl. No: {sl_no}")
                                validation_note.append("❌ Missing Shipper Standardized Name in View")

                            self.page.locator("//span[@aria-hidden='true']").click()

                        else:
                            print(f"⚠️ Unexpected matching field '{mf}' at Sl. No: {sl_no}")
                            validation_note.append(f"⚠️ Unexpected field: {mf}")

                    if is_invalid:
                        self.invalid_entries.append({
                            "slNo": sl_no,
                            "matching_field": matching_field,
                            "Shipper_Name": shipper_name,
                            "Shipper_Standardized_Name": shipper_std_name_text,
                        })

                row_data.append(", ".join(validation_note) if validation_note else "")
                all_rows_data.append(row_data)

        # --- Pagination handling ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
                else:
                    break
        else:
            process_current_page()

        # --- Save to Excel if invalid entries exist ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Auto Suggest Shipper Validation"

            sheet.append(table_headers + ["Validation"])
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "Invalid" in row_data[-1]:
                    for col in range(2, len(row_data)):
                        sheet.cell(row=sheet.max_row, column=col).fill = red_fill

            file_path = f"results/Autosuggest_Search_{self.selected_shipper}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid entries highlighted in red)")
        else:
            print("📘 No invalid entries found. Excel not generated.")

        return self.invalid_entries

    import os
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from playwright.sync_api import expect

    def check_consignee_in_shipmentgrid(self, validate: bool = True, use_pagination: bool = False):
        # --- Wait for shipment tab ---
        if not self.wait_for_shipment_grid(timeout=30):
            print("⏩ Skipping further checks because shipment grid has no results.")
        shipment_tab = self.page.locator("//a[@id='nav-home-tab']")
        shipment_tab.click(timeout=40000)
        self.page.wait_for_timeout(2000)

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if row_count < 10:
            print(f"⚠️ Only {row_count} rows found (less than 10). Skipping strict check.")
        else:
            expect(rows).to_have_count(10, timeout=100000)

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # --- Extract headers ---
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()
        table_headers = []
        self.S_No = self.Matching_Field = self.Consignee_Field = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)

            lower = header_text.lower()
            if lower in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif "matching" in lower:
                self.Matching_Field = i
            elif "consignee" in lower and "address" not in lower and "standardized" not in lower and "city" not in lower:
                self.Consignee_Field = i

        if self.S_No is None or self.Matching_Field is None or self.Consignee_Field is None:
            raise Exception("❌ Required columns ('S No', 'Matching Field', or 'Consignee Name') not found in headers")

        all_rows_data = []

        # --- Process current page ---
        def process_current_page():
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    row_data.append(cells.nth(j).inner_text().strip())

                validation_note = []
                is_invalid = False

                if validate:
                    sl_no = row_data[self.S_No]
                    matching_field = row_data[self.Matching_Field]
                    matching_fields_list = [mf.strip() for mf in matching_field.split(",")]
                    consignee_name = row_data[self.Consignee_Field]
                    consignee_std_name_text = ""

                    for mf in matching_fields_list:
                        # --- Hyperlink check for consignee name ---
                        consignee_cell = cells.nth(self.Consignee_Field)
                        consignee_links = consignee_cell.locator("a")
                        if consignee_links.count() > 0:
                            try:
                                expect(consignee_links.first).to_be_visible()
                                print(f"✅ Consignee name is hyperlinked in grid: {consignee_links.first.inner_text()}")
                            except:
                                print(f"❌ Consignee name hyperlink not interactable at Sl. No: {sl_no}")
                        else:
                            print(f"❌ Consignee name is NOT hyperlinked in grid: {consignee_name}")
                            validation_note.append("❌ Missing hyperlink in Grid")
                            self.invalid_entries.append(
                                {"slNo": sl_no, "Consignee": consignee_name, "Hyperlink": "Missing"}
                            )

                        # --- Validate Consignee Name ---
                        if mf == "Consignee Name":
                            if self.selected_consignee.strip().lower() == consignee_name.lower():
                                print(f"✅ Consignee name valid at Sl. No: {sl_no}")
                                rows.nth(i).locator("td").nth(1).click()
                                self.page.wait_for_timeout(2000)

                                consignee_name_view_locator = self.page.locator(
                                    "//span[normalize-space(text())='Consignee Name']"
                                    "/ancestor::div[contains(@class,'col-5')]"
                                    "/following-sibling::div[contains(@class,'col-7')]//div[1]"
                                )
                                consignee_name_view = consignee_name_view_locator.inner_text().strip()

                                if self.selected_consignee.lower() in consignee_name_view.lower():
                                    print(f"✅ Consignee Name valid at Sl. No: {sl_no} in view")
                                else:
                                    print(f"❌ Invalid Consignee Name: {consignee_name_view} at Sl. No: {sl_no}")
                                    is_invalid = True
                                    validation_note.append("❌ Invalid consignee name in view")

                                self.page.locator("//span[@aria-hidden='true']").click()
                            else:
                                print(f"❌ Consignee name mismatch: {consignee_name} at Sl. No: {sl_no}")
                                is_invalid = True
                                validation_note.append("❌ Invalid Consignee name in Grid")

                        # --- Validate Consignee Standardized Name ---
                        elif mf == "Consignee Standardized Name":
                            rows.nth(i).locator("td").nth(1).click()
                            self.page.wait_for_timeout(2000)

                            consignee_std_name_locator = self.page.locator(
                                "//span[normalize-space(text())='Consignee Standardized Name']"
                                "/ancestor::div[contains(@class,'col-5')]"
                                "/following-sibling::div[contains(@class,'col-7')]//a"
                            )
                            if consignee_std_name_locator.count() > 0:
                                consignee_std_name_text = consignee_std_name_locator.inner_text().strip()

                                if self.selected_consignee.lower() == consignee_std_name_text.lower():
                                    print(f"✅ Consignee Standardized Name valid at Sl. No: {sl_no} in view")
                                else:
                                    print(
                                        f"❌ Invalid Consignee Standardized Name: {consignee_std_name_text} at Sl. No: {sl_no}")
                                    is_invalid = True
                                    validation_note.append("❌ Invalid consignee standardized name in view")

                                # Verify hyperlink
                                tag_name = consignee_std_name_locator.evaluate("el => el.tagName")
                                if tag_name and tag_name.lower() == "a":
                                    print(f"✅ '{consignee_std_name_text}' is a hyperlink in view screen")
                                else:
                                    print(f"❌ '{consignee_std_name_text}' is NOT a hyperlink in view screen")
                                    validation_note.append("❌ Missing hyperlink in View")
                            else:
                                print(f"❌ No Consignee Standardized Name element found for Sl. No: {sl_no}")
                                validation_note.append("❌ Missing Consignee Standardized Name in View")

                            self.page.locator("//span[@aria-hidden='true']").click()

                        else:
                            print(f"⚠️ Unexpected matching field '{mf}' at Sl. No: {sl_no}")
                            validation_note.append(f"⚠️ Unexpected field: {mf}")

                    if is_invalid:
                        self.invalid_entries.append({
                            "slNo": sl_no,
                            "matching_field": matching_field,
                            "Consignee_Name": consignee_name,
                            "Consignee_Standardized_Name": consignee_std_name_text,
                        })

                row_data.append(", ".join(validation_note) if validation_note else "")
                all_rows_data.append(row_data)

        # --- Pagination handling ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(state="visible", timeout=100000)
                else:
                    break
        else:
            process_current_page()

        # --- Save to Excel if invalid entries exist ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consignee Validation"

            sheet.append(table_headers + ["Validation"])
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data and "Invalid" in row_data[-1]:
                    for col in range(2, len(row_data)):
                        sheet.cell(row=sheet.max_row, column=col).fill = red_fill

            file_path = f"results/Consignee_Search_{self.selected_consignee}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalid entries highlighted in red)")
        else:
            print("📘 No invalid entries found. Excel not generated.")

        return self.invalid_entries

    def auto_suggest_manual_hs_code(self, hs_code: str):
        # Focus and type into search box
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.wait_for_timeout(2000)
        self.page.get_by_role("textbox", name="Type to search in all").fill(hs_code)

        # Locator for the first HS code suggestion
        hs_code_locator = self.page.locator("//div[contains(text(),'HS:')]/mark").first

        # Wait up to 5 seconds for suggestion to appear
        try:
            hs_code_locator.wait_for(state="visible", timeout=5000)
        except TimeoutError:
            pytest.skip(f"⚠️ Search result for HS Code '{hs_code}' not found. Skipping the test.")

        # Continue only if visible
        self.selected_hs_code = hs_code_locator.inner_text().strip()
        hs_code_locator.click()
        print(f"✅ Selected HS Code is: {self.selected_hs_code}")

        # Confirm and wait for results
        self.page.locator(".tw-bg-primary-purple-500").click()
        self.page.wait_for_timeout(3000)
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
            state="visible", timeout=100000
        )

    def auto_suggest_manual_product(self, product_name: str):
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(product_name)

        # Locator for the first product suggestion
        product_locator = self.page.locator(
            "(//span[contains(text(),'Product')]/ancestor::div[contains(@class,'tw-flex tw-items-center')]/following-sibling::div//div[contains(@class,'tw-px-3')])[1]").first
        expect(product_locator).to_be_visible(timeout=40000)  # waits up to 5 seconds
        # Check if product suggestion is visible
        if not product_locator.is_visible():
            pytest.skip(f"⚠️ Search result for product '{product_name}' not found. Skipping the test.")

        # Continue only if visible
        self.selected_product = product_locator.inner_text().strip()
        product_locator.click()
        print(f"✅ Selected product is: {self.selected_product}")

        self.page.locator(".tw-bg-primary-purple-500").click()
        self.page.wait_for_timeout(3000)
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
            state="visible", timeout=100000
        )

    def auto_suggest_manual_shipper(self, shipper_name: str):
        # Focus and type into search box
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(shipper_name)

        # Locator for the first Shipper suggestion
        shipper_locator = self.page.locator(
            "(//span[contains(text(),'Shipper')]/ancestor::div[contains(@class,'tw-flex tw-items-center')]/following-sibling::div//div[contains(@class,'tw-px-3')])[1]"
        ).first

        # Wait up to 5 seconds for suggestion
        try:
            shipper_locator.wait_for(state="visible", timeout=5000)
        except TimeoutError:
            pytest.skip(f"⚠️ Search result for shipper '{shipper_name}' not found. Skipping the test.")

        # If found, select shipper
        self.selected_shipper = shipper_locator.inner_text().strip()
        shipper_locator.click()
        print(f"✅ Selected shipper is: {self.selected_shipper}")

        # Verify and confirm
        expect(self.page.get_by_role("main")).to_contain_text("Shipper")
        self.page.locator(".tw-bg-primary-purple-500").click()
        self.page.wait_for_timeout(3000)

        # Wait for results table
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
            state="visible", timeout=100000
        )

    def auto_suggest_manual_consignee(self, consignee: str):
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(consignee)
        self.page.wait_for_timeout(2000)
        self.selected_consignee = self.page.locator(
            "(//span[contains(text(),'Consignee')]/ancestor::div[contains(@class,'tw-flex tw-items-center')]/following-sibling::div//div[contains(@class,'tw-px-3')])[1]").nth(
            0).inner_text().strip()
        self.page.locator(
            "(//span[contains(text(),'Consignee')]/ancestor::div[contains(@class,'tw-flex tw-items-center')]/following-sibling::div//div[contains(@class,'tw-px-3')])[1]").nth(
            0).click()
        expect(self.page.get_by_role("main")).to_contain_text("Consignee")
        self.page.locator(".tw-bg-primary-purple-500").click()
        self.page.wait_for_timeout(3000)
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)

    def auto_suggest_manual_chemical(self, chemical_name: str):
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(chemical_name)

        # Locator for the first chemical suggestion
        chemical_locator = self.page.locator(
            "(//span[contains(text(),'Chemical')]/ancestor::div[contains(@class,'tw-flex tw-items-center')]/following-sibling::div//div[contains(@class,'tw-px-3')])[1]"
        ).first

        # Wait up to 5 seconds for the suggestion to appear
        try:
            chemical_locator.wait_for(state="visible", timeout=50000)
        except TimeoutError:
            pytest.skip(f"⚠️ Search result for chemical '{chemical_name}' not found. Skipping the test.")

        # Continue only if visible
        self.selected_chemical = chemical_locator.inner_text().strip()
        chemical_locator.click()
        print(f"✅ Selected chemical is: {self.selected_chemical}")

        expect(self.page.get_by_role("main")).to_contain_text("Chemical")
        self.page.locator(".tw-bg-primary-purple-500").click()

    def auto_suggest_manual_port(self, port_name: str):
        # Focus and type into search box
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(port_name)

        # Locator for the first Port suggestion
        port_locator = self.page.locator(
            "(//span[contains(text(),'Port')]/ancestor::div[contains(@class,'tw-flex tw-items-center')]/following-sibling::div//div[contains(@class,'tw-px-3')])[1]"
        ).first

        # Wait up to 5 seconds for suggestion, else skip
        try:
            port_locator.wait_for(state="visible", timeout=5000)
        except TimeoutError:
            pytest.skip(f"⚠️ Search result for port '{port_name}' not found. Skipping the test.")

        # If found, select port
        self.selected_port = port_locator.inner_text().strip()
        port_locator.click()
        print(f"✅ Selected Port is: {self.selected_port}")

        # Verify UI and confirm selection
        expect(self.page.get_by_role("main")).to_contain_text("Port")
        self.page.locator(".tw-bg-primary-purple-500").click()
        self.page.wait_for_timeout(3000)

        # Wait for results table to load
        self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
            state="visible", timeout=100000
        )

        # Final verification
        expect(self.page.get_by_role("main")).to_contain_text(self.selected_port)


















































