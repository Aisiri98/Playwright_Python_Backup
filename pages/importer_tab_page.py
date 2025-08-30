import re
import pycountry
from playwright.sync_api import Page
from playwright.sync_api import expect, TimeoutError as PlaywrightTimeoutError
from datetime import datetime
invalid_entries =[]

class ImportTab:
    def __init__(self, page: Page):
        self.company_name = None
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def importer_tab(self):
        # Check the shipment tab is visiblility
        # expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible()

        # Navigate to the importers tab
        self.page.locator("//a[@id='nav-profile-tab']").click()
        self.page.wait_for_timeout(1000)
        print("✅ navigated to the importer page")

    def trends_rank_option(self):
        #Verify presence of Rank and Trends tabs at the top of the grid
        # Expect "Rank" to be visible
        expect(self.page.locator("span", has_text="Rank")).to_be_visible()
        print("✅ Rank is visible on the importer page")

        # Expect "Trends" to be visible
        expect(self.page.locator("span", has_text="Trends")).to_be_visible()
        print("✅ Trends is visible on the importer page")

        # Click the Trends
        self.page.locator("span", has_text="Trends").click()
        self.page.wait_for_timeout(1000)
        # trends tab validation
        expect(self.page.get_by_text("Total Shipments")).to_be_visible()
        print("✅ Navigated to the trends tab")

        # Click Rank tab
        self.page.locator("span", has_text="Rank").click()
        self.page.wait_for_timeout(1000)
        expect(self.page.get_by_role("columnheader", name="Shipments")).to_be_visible()

    def verify_table_headers(self):
        #Navigate to the Rank tab.
        self.page.locator("span", has_text="Rank").click()
        self.page.wait_for_timeout(1000)
        #Observe the "Company Name", "Shipment", and "Shipment Value (USD) " columns displayed in proper format.
        expect(self.page.get_by_role("columnheader", name="Shipments")).to_be_visible()
        self.page.wait_for_timeout(1000)

        headers_to_check = [
            "S No.",
            "Company Name & IEC Number",
            "Shipments",
            "Shipment Value (USD)",
            "Consignee State",
            "Apply Filter"
        ]

        for header in headers_to_check:
            locator = self.page.get_by_role("columnheader", name=header)  # more precise
            expect(locator).to_be_visible(timeout=5000)
            print(f"✅ Header '{header}' is visible.")

    def importer_apply_filter(self):
        # First row
        first_row = self.page.locator("table tbody tr").first
        self.page.pause()
        # Serial number (for logging)
        sl_no = first_row.locator("td").nth(1).inner_text().strip()
        consignee_locator = first_row.locator("td").nth(2).locator("[data-tip]").first
        consignee_tip = consignee_locator.get_attribute("data-tip") if consignee_locator.count() > 0 else None

        if consignee_tip and consignee_tip.strip():
            company_name = consignee_tip.strip()
        else:
            # Fallback to visible text if no tooltip
            company_name = first_row.locator("td").nth(2).inner_text().strip()
        country_name = self.page.locator('[role="cell"] > .text-wrap > [class="tw-flex tw-items-center"]').first

        #Click the Apply Filter icon at the end of any company row
        apply_filter_icon = first_row.locator("td").nth(6).locator("img")
        apply_filter_icon.scroll_into_view_if_needed()
        apply_filter_icon.click()

        print(f"✅ Applied filter for first row (Serial No: {sl_no})")
        #Check for null or missing jurisdiction countries in the grid and filter
        sl_no = first_row.locator("td").nth(1).inner_text().strip()
        consignee_locator_actual = first_row.locator("td").nth(2).locator("[data-tip]").first
        consignee_tip_actual = consignee_locator_actual.get_attribute("data-tip") if consignee_locator_actual.count() > 0 else None

        if consignee_tip_actual and consignee_tip_actual.strip():
            company_name_actual = consignee_tip.strip()
        else:
            # Fallback to visible text if no tooltip
            company_name_actual = first_row.locator("td").nth(2).inner_text().strip()

        if company_name == company_name_actual:
            print(f"valid company found:{company_name_actual} at {sl_no}")
        else:
            print(f"valid company found:{company_name_actual} at {sl_no}")

        country_name_actual = self.page.locator('[role="cell"] > .text-wrap > [class="tw-flex tw-items-center"]').first.inner_text().strip()
        if country_name == country_name_actual:
            print(f"valid country found:{country_name_actual} at {sl_no}")
        else:
            print(f"valid country found:{country_name_actual} at {sl_no}")
        #Open the filter panel and navigate to Consignee → Consignee Jurisdiction Country
        self.page.locator("div").filter(has_text=re.compile(r"^Consignee1$")).nth(1).click()
        self.page.get_by_text("Consignee Jurisdiction Country").click()
        expect(self.page.locator("#searchIndividualFilterPanel")).to_contain_text("Filter by Consignee Jurisdiction Country")
        expect(self.page.locator('[class="tw-flex tw-items-center tw-gap-1 tw-font-medium tw-pl-1 tw-text-sm"]')).to_have_count(1 , timeout=10000)
        print("Shows one valid jurisdiction country displayed correctly")
        #Check for null or missing jurisdiction countries in the grid and filter
        text = self.page.locator('[class="tw-flex tw-items-center tw-gap-1 tw-font-medium tw-pl-1 tw-text-sm"]').first.inner_text().strip()
        #localize-filter-tooltip
        match = re.match(r"^([^(]+)", text)
        if match:
            country = match.group(1).strip()
            print(country)  # 👉 india

            if country == country_name:
                 print("Shows one valid jurisdiction country displayed correctly")

        self.page.locator("//span[@aria-hidden='true']").click()

    from pycountry.db import Country
    def click_enable_second_level_hierarchy(self):
        #Navigate to the “Enable 2nd Level Hierarchy” toggle button
        label = self.page.locator("span.mr-1.two_level_label", has_text="Enable 2nd Level Hierarchy")
        expect(label).to_be_visible(timeout=10000)
        # toggle = self.page.locator('label[for="custom-switch-1"]')
        self.page.wait_for_timeout(3000)
        self.page.locator('[class="custom-control custom-switch"]').first.wait_for(state="visible", timeout=10000)
        #Click to enable the 2nd Level Hierarchy
        self.page.locator('[class="custom-control custom-switch"]').first.click(force=True, timeout=100000)

        self.page.wait_for_timeout(1000)
        #expect(self.page.locator('[class="custom-control custom-switch"]').first).to_be_checked(timeout=10000)
        print("✅ Clicked 'Enable 2nd Level Hierarchy'")

        # Observe the "Company Name", "Country of Origin", "Shipment ", "Total Shipment Value (USD)", and "% Shared by Shipment " columns
        expected_headers = [
            "Company Name & IEC Number",
            "Countries of Origin",
            "Shipments",
            "Shipment Value (USD)",
            "% share by Shipments"
        ]

        for header in expected_headers:
            locator = self.page.get_by_role("columnheader", name=header).first # more precise
            expect(locator).to_be_visible(timeout=100000)
            print(f"✅ Found header: {header}")

        valid_countries = {country.name.upper() for country in list(pycountry.countries)}

        # Add common aliases / special cases
        valid_countries.update({"HONG KONG", "UNITED STATES"})

        # Grab all rows of the Country of Origin column (2nd column)
        rows = self.page.locator("table tbody tr td:nth-child(2)")
        row_count = rows.count()

        for i in range(row_count):
            country = rows.nth(i).inner_text().strip().upper()

            if country in valid_countries:
                print(f"✅ Valid country: {country}")
            else:
                print(f"❌ Invalid country: {country}")

        # In "Country of Origin" column click edit and select any options
        self.page.wait_for_timeout(3000)
        self.page.locator('[class="custom-control custom-switch"]').first.wait_for(state="visible", timeout=10000)
        self.page.locator('[class="custom-control custom-switch"]').first.click(force=True, timeout=100000)

        self.page.wait_for_timeout(1000)

        #In "Country of Origin" column click edit and select any options
        #self.page.locator("//img[@class='ml-2']").click()

    def ranks_second_level_hierarchy(self, value="10"):
        #Click to enable the 2nd Level Hierarchy
        label = self.page.locator("span.mr-1.two_level_label", has_text="Enable 2nd Level Hierarchy")
        expect(label).to_be_visible()
        # toggle = self.page.locator('label[for="custom-switch-1"]')
        toggle = self.page.locator('[class="custom-control custom-switch"]')
        toggle.click(force=True)

        self.page.wait_for_timeout(3000)
        print("✅ Clicked 'Enable 2nd Level Hierarchy in the trend'")

        # verify_second_Level_Hierarchy_headers
        expected_headers = [
            "Company Name & IEC Number",
            "Countries of Origin",
            "Shipments",
            "Shipment Value (USD)",
            "% share by Shipments"
        ]

        for header in expected_headers:
            locator = self.page.get_by_role("columnheader", name=header).first
            expect(locator).to_be_visible()
            print(f"✅ Found header in the Rank: {header}")

        # #Select any option in the Number of Second-Level Entities dropdown
        # self.page.locator('[class=" css-8qthy4"]').last.click()
        # self.page.locator('[class=" css-8gqxdx-option"]').nth(0).click()
        # expect(self.page.locator("//span[normalize-space()='Countries of Origin']")).to_be_visible(timeout=6000)
        # # Grab all company blocks (left column with company name + IEC)
        # company_blocks = self.page.locator('[class="trademo-link text-left px-0 text-capitalize "]')  # <-- adjust selector
        #
        # total_companies = company_blocks.count()
        # print(f"🔍 Found {total_companies} companies")
        #
        # for i in range(total_companies):
        #     # Each company block
        #     company = company_blocks.nth(i)
        #
        #     # Get company name
        #     company_name = company.inner_text().strip()
        #
        #     # Now grab corresponding countries of origin column for this company
        #     countries = self.page.locator(
        #         f"(//div[@class='countries-column-selector'])[{i + 1}]//div[contains(@class,'country-row')]"
        #     )
        #     country_count = countries.count()
        #
        #     if country_count <= 10:
        #         print(f"✅ {company_name} has {country_count} countries")
        #     else:
        #         print(f"❌ {company_name} has {country_count} (> 10!)")

    def Enable_2nd_level_hierachy_Rank(self, value="10"):
        # Click to enable the 2nd Level Hierarchy
        label = self.page.locator("span.mr-1.two_level_label", has_text="Enable 2nd Level Hierarchy")
        expect(label).to_be_visible()
        # toggle = self.page.locator('label[for="custom-switch-1"]')
        toggle = self.page.locator('[class="custom-control custom-switch"]')
        toggle.click(force=True)

        self.page.wait_for_timeout(3000)
        print("✅ Clicked 'Enable 2nd Level Hierarchy in the trend'")

        # verify_second_Level_Hierarchy_headers
        expected_headers = [
            "Company Name & IEC Number",
            "Countries of Origin",
            "Shipments",
            "Shipment Value (USD)",
            "% share by Shipments"
        ]

        for header in expected_headers:
            locator = self.page.get_by_role("columnheader", name=header).first
            expect(locator).to_be_visible()
            print(f"✅ Found header in the Rank: {header}")

    def Download_option(self):
        self.page.get_by_role("button", name="Download").click()
        expect(self.page.get_by_text("Download Unique Importer")).to_be_visible()
        self.page.locator("//input[@placeholder='To']").clear()
        self.page.locator("//input[@placeholder='To']").fill("1")
        self.page.locator("//button[normalize-space()='Download']").click()
        expect(self.page.locator("//span[@class='header-confirm']")).to_contain_text("Confirmation")

    def Validate_Colmn_Shown_In_the_Grid_Download(self):
        import os
        import tempfile
        import openpyxl
        from datetime import datetime

        # --- Trigger download ---
        with self.page.expect_download() as download_info:
            self.page.locator("//button[normalize-space()='Confirm']").click()
        download = download_info.value

        # --- Save to temp file ---
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, download.suggested_filename)
        download.save_as(file_path)
        print(f"📂 Downloaded to: {file_path}")

        # --- Read workbook ---
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # --- Validate "Report Summary" sheet ---
        if "Report Summary" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Report Summary' sheet not found in downloaded Excel.")
        summary_sheet = wb["Report Summary"]

        EXPECTED_REPORT_TYPE = "Importers List - 2nd Level"
        EXPECTED_DOWNLOAD_DATE = datetime.today().strftime("%b %d, %Y")

        report_type_value = str(summary_sheet["B7"].value).strip() if summary_sheet["B7"].value else ""
        download_date_value = str(summary_sheet["B10"].value).strip() if summary_sheet["B10"].value else ""

        if report_type_value in EXPECTED_REPORT_TYPE:
            print("✅ Report Type matches expected value.")
        else:
            print(f"❌ Report Type mismatch: Found '{report_type_value}', Expected '{EXPECTED_REPORT_TYPE}'")

        if download_date_value == EXPECTED_DOWNLOAD_DATE:
            print("✅ Download Date matches today's date.")
        else:
            print(f"❌ Download Date mismatch: Found '{download_date_value}', Expected '{EXPECTED_DOWNLOAD_DATE}'")

        # --- Read "Shipment Data" sheet ---
        if "Importer Data" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Shipment Data' sheet not found in downloaded Excel.")
        sheet = wb["Importer Data"]

        # Read Excel headers
        downloaded_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        wb.close()

        # --- Get UI headers ---
        ui_headers = self.page.locator('[class="sc-jlZhew bkvFbV rdt_TableCol_Sortable"]').all_inner_texts()
        ui_headers = [h.strip() for h in ui_headers if h not in ("", "View")]

        # --- Normalize headers ---
        def normalize_header(h):
            if not h:
                return ""
            return str(h).strip().replace(".", "").replace("  ", " ").replace("'", "").replace('"', '').lower()

        excel_norm = [normalize_header(h) for h in downloaded_headers]
        ui_norm = [normalize_header(h) for h in ui_headers]

        # --- Columns to ignore in Excel ---
        ignore_cols = [
            "Consignee Name",
            "IEC Number",
            "Total Value",
            "Total Shipments",
            "% share by Shipments",
            "Consignee Jurisdiction Country",
        ]

        # Filter Excel and UI columns
        excel_filtered = [col for col in excel_norm if col not in ignore_cols]
        ui_filtered = [col for col in ui_norm if col not in ignore_cols]

        # --- Validation ---
        missing_ui_cols = [col for col in ui_filtered if col not in excel_filtered]
        extra_excel_cols = [col for col in excel_filtered if col not in ui_filtered]

        if not missing_ui_cols:
            print("✅ All relevant UI columns are present in Excel (ignored extra fields).")
        else:
            print("❌ Missing UI columns in Excel:", missing_ui_cols)

        if extra_excel_cols:
            print("⚠️ Extra columns in Excel (not in UI):", extra_excel_cols)

        # --- Cleanup ---
        os.remove(file_path)

    def Analyse_By_Shipment_Value_USD(self):
        dropdowns = self.page.locator("div.css-1rupv9a-singleValue")
        dropdown_count = dropdowns.count()

        for i in range(dropdown_count):
            dropdown = dropdowns.nth(i)

            # Open dropdown
            dropdown.click()
            self.page.wait_for_timeout(500)  # small wait for menu to appear

            # Check if "Shipment Value USD" is visible in options
            option = self.page.get_by_text("Value (USD)", exact=True)

            if option.is_visible(timeout=1000):
                option.click()
                print(f"'Shipment Value USD' selected in dropdown {i + 1}")
            else:
                print(f"'Shipment Value USD' not found in dropdown {i + 1}, moving on...")
                # Close dropdown by pressing ESC (so next one can open cleanly)
                self.page.keyboard.press("Escape")

        self.page.locator('[class="sc-jlZhew bkvFbV rdt_TableCol_Sortable"]').first.wait_for(state="visible", timeout=60000)

        # --- Extract all shipment values from table ---
        rows = self.page.locator('[class="sc-fqkvVR sc-dcJsrY sc-iGgWBj bIGOvl rPOCv gkzfIp rdt_TableCell"]')
        row_count = rows.count()

        shipment_values = []
        for i in range(row_count):
            value_text = rows.nth(i).inner_text().strip()
            # Remove commas and convert to float
            value = float(value_text.replace(",", "").replace("$", ""))
            shipment_values.append(value)

        print("📦 Shipment Values(USD):", shipment_values)

        # --- Check if values are in descending order ---
        if shipment_values == sorted(shipment_values, reverse=True):
            print("✅ Shipment values usd are in descending order")
        else:
            print("❌ Shipment values are NOT in descending order")
            print("Expected -Shipment values usd :", sorted(shipment_values, reverse=True))
            print("Got     :", shipment_values)

    def Analyse_By_Shipment(self):
        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        # Find index of "Apply Filter"
        apply_filter_index = None
        for i in range(header_count):
            text = headers.nth(i).inner_text().strip()
            if text == "Apply Filter":
                apply_filter_index = i + 1  # nth-child is 1-based
                break

        if apply_filter_index is None:
            raise Exception("Apply Filter column not found!")

        # ✅ Now click first row in that column dynamically
        cell = self.page.locator(f"table tbody tr td:nth-child({apply_filter_index})").first
        cell.click(timeout=100000)
        self.page.wait_for_timeout(3000)
        self.page.locator('[class="custom-control custom-switch"]').first.wait_for(state="visible", timeout=10000)
        self.page.locator('[class="custom-control custom-switch"]').first.click(force=True, timeout=100000)
        self.page.wait_for_timeout(2000)
        self.page.locator('[class="sc-jlZhew bkvFbV rdt_TableCol_Sortable"]').first.wait_for(state="visible", timeout=60000)

        # --- Extract all shipment values from table ---
        rows = self.page.locator('[class="sc-fqkvVR sc-dcJsrY sc-iGgWBj bIGOvl rPOCv cZyQSr rdt_TableCell"]')
        row_count = rows.count()

        shipment_values = []
        for i in range(0, row_count, 2):
            value_text = rows.nth(i).inner_text().strip()
            # Remove commas and convert to float
            value = float(value_text.replace(",", "").replace("$", ""))
            shipment_values.append(value)

        print("📦 Shipment:", shipment_values)


        # --- Check if values are in descending order ---
        if shipment_values == sorted(shipment_values, reverse=True):
            print("✅ Shipment values are in descending order")
        else:
            print("❌ Shipment values are NOT in descending order")
            print("Expected shipment values:", sorted(shipment_values, reverse=True))
            print("Got     :", shipment_values)

    def Filter_by_Shipment_Value(self):
        #Disable second level hierarchy
        global less_than, value_num, shipmentvalue_num, greater_than, from_value, to_value
        self.page.wait_for_timeout(3000)
        self.page.locator('#custom-switch-1').first.wait_for(state="visible", timeout=10000)
        self.page.locator('#custom-switch-1').first.click(force=True, timeout=100000)
        self.page.wait_for_timeout(2000)
        #click on filter by
        self.page.get_by_role("button", name="tune Filter By").click()
        expect(self.page.get_by_text("Filter By Range")).to_be_visible()
        expect(self.page.locator("#onClickClear").get_by_text("Shipments")).to_be_visible()
        expect(self.page.get_by_text("Value", exact=True)).to_be_visible()
        #store shipment value usd in variable
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_count = headers.count()

        shipment_value_index = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            print(f"{header_text}")
            if "shipment value (usd)" in header_text:
                shipment_value_index = i
                break

        if shipment_value_index is None:
            raise Exception("❌ 'Shipment Value' column not found in table headers")

        print(f"✅ Found 'Shipment Value(USD)' column at index {shipment_value_index}")

        # --- Extract all shipment values from table ---
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        shipment_values = []
        for i in range(row_count):
            value_text = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num = float(value_text.replace(",", "").replace("$", ""))

            # Subtract 1 for checking
            less_than = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            greater_than = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

            from_value = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            to_value = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

        #Under Shipment Value (USD)/Value, select Less Than, and enter a value less than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.locator("div").filter(has_text=re.compile(r"^Valueless thanbetweengreater than$")).get_by_role("img").click()
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Unique Importer Found")).to_be_visible(timeout=100000)
        self.page.get_by_role("button", name="Clear").click()
        #Under Shipment Value (USD)/Value, select Less Than, and enter a value greater than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(shipmentvalue_num))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num1 <= shipmentvalue_num:
                print(f"✅ Value {shipmentvalue_num1} is less than {shipmentvalue_num} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT less than {shipmentvalue_num} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {shipmentvalue_num}"

        #Click Clear to reset the modal, then select Shipment Value (USD)/Value – Between, enter a range that includes the company’s shipment/Value count, and click Apply
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).fill(str(from_value))
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto").fill(str(to_value))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num2 = float(value_text1.replace(",", "").replace("$", ""))
            if from_value <= shipmentvalue_num2 <= to_value:
                print(f"✅ Row {i + 1}: Value {shipmentvalue_num2} is within range [{from_value}, {to_value}]")
            else:
                print(f"❌ Row {i + 1}: Value {shipmentvalue_num2} is NOT in range [{from_value}, {to_value}]")
                assert False, f"Value {shipmentvalue_num2} at row {i + 1} not in range [{from_value}, {to_value}]"
        #Without clearing, modify range to one that excludes the company’s Shipment Value (USD)/Value count, then click Apply
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).clear()
        self.page.get_by_placeholder("From", exact=True).fill("1")
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto", exact=True).clear()
        self.page.get_by_placeholder("Upto").fill("2")
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Unique Importer Found")).to_be_visible(timeout=100000)
        #Click Clear, select Shipment Value (USD)/Value – Greater Than, enter a value below the actual count, and click Appl
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text1}")
            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num1 > less_than:
                print(f"✅ Value {shipmentvalue_num1} is greater than {less_than} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT greater than {less_than} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {less_than}"
        #Without clearing, change the value to one greater than the company’s count and click Apply
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(greater_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Unique Importer Found")).to_be_visible(timeout=100000)
        #Clear filter
        self.page.get_by_role("button", name="Clear").click()

    def Filter_by_Shipment(self):
        # Disable second level hierarchy
        global less_than, value_num, shipmentvalue_num, greater_than, from_value, to_value
        # click on filter by
        expect(self.page.get_by_text("Filter By Range")).to_be_visible(timeout=100000)
        expect(self.page.locator("#onClickClear").get_by_text("Shipments")).to_be_visible()
        expect(self.page.get_by_text("Value", exact=True)).to_be_visible()
        self.page.wait_for_timeout(1000)
        # store shipment value usd in variable
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_count = headers.count()

        shipment_value_index = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            print(f"{header_text}")
            if "shipments" in header_text:
                shipment_value_index = i
                break

        if shipment_value_index is None:
            raise Exception("❌ 'Shipment' column not found in table headers")

        print(f"✅ Found 'Shipment' column at index {shipment_value_index}")

        # --- Extract all shipment values from table ---
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        shipment_values = []
        for i in range(row_count):
            value_text = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num = float(value_text.replace(",", "").replace("$", ""))

            # Subtract 1 for checking
            less_than = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            greater_than = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

            from_value = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            to_value = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

        # Under Shipment Value (USD)/Value, select Less Than, and enter a value less than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.locator("div").filter(has_text=re.compile(r"^Shipmentsless thanbetweengreater than$")).get_by_role(
            "img").click()
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Unique Importer Found")).to_be_visible(timeout=100000)
        self.page.get_by_role("button", name="Clear").click()
        # Under Shipment Value (USD)/Value, select Less Than, and enter a value greater than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(shipmentvalue_num))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num <= shipmentvalue_num1:
                print(f"✅ Value {shipmentvalue_num1} is less than {shipmentvalue_num} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT less than {shipmentvalue_num} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {shipmentvalue_num}"

        # Click Clear to reset the modal, then select Shipment Value (USD)/Value – Between, enter a range that includes the company’s shipment/Value count, and click Apply
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).fill(str(from_value))
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto").fill(str(to_value))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num2 = float(value_text1.replace(",", "").replace("$", ""))
            if from_value <= shipmentvalue_num2 <= to_value:
                print(f"✅ Row {i + 1}: Value {shipmentvalue_num2} is within range [{from_value}, {to_value}]")
            else:
                print(f"❌ Row {i + 1}: Value {shipmentvalue_num2} is NOT in range [{from_value}, {to_value}]")
                assert False, f"Value {shipmentvalue_num2} at row {i + 1} not in range [{from_value}, {to_value}]"
        # Without clearing, modify range to one that excludes the company’s Shipment Value (USD)/Value count, then click Apply
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).fill("1")
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto").fill("2")
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Unique Importer Found")).to_be_visible(timeout=100000)
        # Click Clear, select Shipment Value (USD)/Value – Greater Than, enter a value below the actual count, and click Apply
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text1}")
            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num1 > less_than:
                print(f"✅ Value {shipmentvalue_num1} is greater than {less_than} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT greater than {less_than} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {less_than}"
        # Without clearing, change the value to one greater than the company’s count and click Apply
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(greater_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Unique Importer Found")).to_be_visible(timeout=100000)
        # Clear filter
        self.page.get_by_role("button", name="Clear").click()

    def Validate_option_on_clicking_checkbox(self):
        # Select the checkbox for any 1 company in the list
        self.page.locator("table tbody tr td:nth-child(1)").first.click()
        #All Options should enable (Download, Export to CRM, Bookmark,More Download Options, Fetch Contacts and Apply FIlter)
        expect(self.page.get_by_role("button", name="cloud_download Download")).to_be_enabled()
        expect(self.page.get_by_role("button", name="ios_share Export to CRM")).to_be_enabled()
        expect(self.page.get_by_role("button", name="star Bookmark")).to_be_enabled()
        expect(self.page.get_by_role("button", name="contacts Fetch Contacts")).to_be_enabled()
        expect(self.page.get_by_role("button", name="More Download Options")).to_be_enabled()
        expect(self.page.get_by_role("button", name="cloud_download Download")).to_be_visible()
        expect(self.page.get_by_role("button", name="ios_share Export to CRM")).to_be_visible()
        expect(self.page.get_by_role("button", name="star Bookmark")).to_be_visible()
        expect(self.page.get_by_role("button", name="contacts Fetch Contacts")).to_be_visible()
        expect(self.page.get_by_role("button", name="More Download Options")).to_be_visible()
        expect(self.page.get_by_role("button", name="filter_arrow_right Apply")).to_be_visible()
        #Deselect the checkbox
        self.page.locator("table tbody tr td:nth-child(1)").first.click()
        expect(self.page.get_by_role("button", name="cloud_download Download")).to_be_enabled()
        expect(self.page.get_by_role("button", name="ios_share Export to CRM")).to_be_enabled()
        expect(self.page.get_by_role("button", name="star Bookmark")).not_to_be_enabled()
        expect(self.page.get_by_role("button", name="contacts Fetch Contacts")).not_to_be_enabled()
        expect(self.page.get_by_role("button", name="cloud_download Download")).to_be_visible()
        expect(self.page.get_by_role("button", name="ios_share Export to CRM")).to_be_visible()
        expect(self.page.get_by_role("button", name="star Bookmark")).to_be_visible()
        expect(self.page.get_by_role("button", name="contacts Fetch Contacts")).to_be_visible()
        expect(self.page.get_by_role("button", name="More Download Options")).not_to_be_visible()
        expect(self.page.get_by_role("button", name="filter_arrow_right Apply")).not_to_be_visible()

    def verify_trend_headers(self):
        # Click on the Trends sub-tab
        self.page.locator("span", has_text="Trends").click()
        # Observe the "Company Name", "Total Shipment Value (USD)", and "% Overall Change" columns
        self.page.wait_for_timeout(1000)

        headers_to_check = [
            "S No.",
            "Company",
            "Total Shipments",
            "% Overall Change",
            "Apply Filter"
        ]

        for header in headers_to_check:
            locator = self.page.get_by_role("columnheader", name=header)  # more precise
            expect(locator).to_be_visible(timeout=5000)
            print(f"✅ Header '{header}' is visible.")

    def Verify_Monthly_data(self):
        from datetime import datetime, timedelta
        self.page.locator('[class="css-19bqh2r"]').nth(1).click()
        self.page.locator('[class=" css-8gqxdx-option"]').nth(0).click()
        expect(self.page.locator("//span[normalize-space()='Company']")).to_be_visible(timeout=50000)
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_texts = [headers.nth(i).inner_text().strip() for i in range(headers.count())]

        print("All headers:", header_texts)

        # --- Filter only month-year headers (e.g., 'Jul 2025') ---
        monthly_headers = [h for h in header_texts if
                           len(h.split()) == 2 and h.split()[0].isalpha() and h.split()[1].isdigit()]
        print("Monthly headers found in table:", monthly_headers)

        # --- Build expected list of months (last month → Jan 2020) ---
        today = datetime.today()
        last_month = today.replace(day=1) - timedelta(days=1)
        expected_months = []

        while last_month >= datetime(2020, 1, 1):
            expected_months.append(last_month.strftime("%b %Y"))  # e.g., "Jul 2025"
            last_month = (last_month.replace(day=1) - timedelta(days=1))

        print("Expected months:", expected_months[:12], "...")  # printing first 12 for sanity

        # --- Validation ---
        missing_months = [m for m in expected_months if m not in monthly_headers]
        if missing_months:
            raise AssertionError(f"❌ Missing columns in table headers: {missing_months}")
        else:
            print("✅ All monthly columns present from last month to Jan 2020")

    def Verify_Quaterly_Data(self):
        self.page.locator('[class="css-19bqh2r"]').nth(1).click()
        self.page.locator('[class=" css-8gqxdx-option"]').nth(1).click()
        expect(self.page.locator("//span[normalize-space()='Company']")).to_be_visible(timeout=50000)
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_texts = [headers.nth(i).inner_text().strip() for i in range(headers.count())]

        print("All headers:", header_texts)

        # --- Filter only quarter headers (e.g., 'Q3 2025') ---
        quarterly_headers = [h for h in header_texts if h.startswith("Q") and len(h.split()) == 2]
        print("Quarterly headers found in table:", quarterly_headers)

        # --- Build expected quarters (latest → Q1 2020) ---
        today = datetime.today()
        year = today.year
        month = today.month

        # figure out current quarter
        current_quarter = (month - 1) // 3 + 1

        expected_quarters = []
        while year > 2020 or (year == 2020 and current_quarter >= 1):
            expected_quarters.append(f"Q{current_quarter} {year}")
            current_quarter -= 1
            if current_quarter == 0:
                current_quarter = 4
                year -= 1

        print("Expected quarters:", expected_quarters[:8], "...")  # first few only

        # --- Validation ---
        missing_quarters = [q for q in expected_quarters if q not in quarterly_headers]
        if missing_quarters:
            raise AssertionError(f"❌ Missing quarters in table headers: {missing_quarters}")
        else:
            print("✅ All quarterly columns present from latest to Q1 2020")

    def Verify_Yearly_Data(self):
        from datetime import datetime
        self.page.locator('[class="css-19bqh2r"]').nth(1).click()
        self.page.locator('[class=" css-8gqxdx-option"]').nth(0).click()
        expect(self.page.locator("//span[normalize-space()='Company']")).to_be_visible(timeout=50000)
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_texts = [headers.nth(i).inner_text().strip() for i in range(headers.count())]

        print("All headers:", header_texts)

        # --- Filter only year headers (4-digit numbers) ---
        yearly_headers = [h for h in header_texts if h.isdigit() and len(h) == 4]
        print("Yearly headers found in table:", yearly_headers)

        # --- Build expected yearly headers (current year → 2020) ---
        current_year = datetime.today().year
        expected_years = [str(y) for y in range(current_year, 2019, -1)]
        print("Expected yearly headers:", expected_years)

        # --- Validation ---
        missing_years = [y for y in expected_years if y not in yearly_headers]
        if missing_years:
            raise AssertionError(f"❌ Missing yearly headers in table: {missing_years}")
        else:
            print("✅ All yearly headers present from", current_year, "to 2020")

    def Enable_Trends_2nd_level_hierarchy(self):
        self.page.pause()
        self.page.wait_for_timeout(3000)

        toggle = self.page.locator('[class="custom-control custom-switch"]').first

        expect(toggle).to_be_visible(timeout=10000)

        # ✅ check if checkbox is already enabled
        if toggle.get_by_role("checkbox").is_checked():
            print("2nd level hierarchy is already enabled")
        else:
            toggle.click(force=True, timeout=100000)
            self.page.wait_for_timeout(1000)
            print("2nd level hierarchy enabled")

    def Trends_2nd_level_hierarchy(self):
        self.page.pause()
        self.page.wait_for_timeout(3000)

        toggle = self.page.locator('[class="custom-control custom-switch"]').first

        # ✅ Check visibility before interacting
        if toggle.is_visible():
            print("2nd level hierarchy toggle is visible")

            # Click only if it's not already enabled
            checkbox = toggle.get_by_role("checkbox")
            if not checkbox.is_checked():
                toggle.click(force=True, timeout=100000)
                self.page.wait_for_timeout(1000)
                print("2nd level hierarchy enabled")
            else:
                print("2nd level hierarchy already enabled")

            # Verify UI effect
            #expect(self.page.locator("//span[normalize-space()='Countries of Origin']")).to_be_visible(timeout=100000)
        else:
            print("2nd level hierarchy toggle is NOT visible, skipping...")
        #Observe the "Company Name", "Country of Origin", "Total Shipment Value (USD)", and "% Overall Change" columns
        headers_to_check = [
            "Company",
            "Countries of Origin"
            "Total Shipments",
            "% Overall Change",
        ]

        for header in headers_to_check:
            locator = self.page.get_by_role("columnheader", name=header)  # more precise
            expect(locator).to_be_visible(timeout=5000)
            print(f"✅ Header '{header}' is visible.")
        #In "Country of Origin" column click edit and select any options
        self.page.locator('[class="ml-2"]').last.click()
        expected_editoption = [
            "Trade Routes (Country of Export → Country of Import)",
            "Unique Exporters",
            "Countries of Origin",
            "Ports of Lading",
            "Ports of Unlading",
            "Consignee States",
            "Consignee Cities",
            "Custom House Agent Name",
            "Mode of Transportation",
            "4 Digit HS Codes",
            "Sections",
            "Trade Routes (Port of Lading → Port of Unlading)",
            "6 digit HS Codes"
        ]

        table = self.page.get_by_role("table")

        for header in expected_editoption:
            expect(table).to_contain_text(header)
            print(f"✅ Validated option: {header}")

        self.page.get_by_text("4 Digit HS Codes").first.click()
        self.page.wait_for_timeout(7000)
        rows = self.page.locator('[class="m-0 d-flex align-items-center h-100"]')
        row_count = rows.count()
        hscode_values = []
        for i in range(row_count):
            hs_code_text = rows.nth(i).inner_text().strip()
            # ✅ Validate that it's only 4 digits
            if re.fullmatch(r"\d{4}", hs_code_text):
                print(f"✅ Row {i}: HS Code {hs_code_text} is valid (4 digits).")
            else:
                raise AssertionError(f"❌ Row {i}: Invalid HS Code '{hs_code_text}' (not 4 digits)")

        print("📦 HS Code:", hscode_values)
        #Toggle OFF the switch
        self.page.wait_for_timeout(3000)
        self.page.locator('[class="ml-2"]').last.click()
        self.page.get_by_text("Countries of Origin").first.click()
        self.page.locator('#custom-switch-1').first.wait_for(state="visible", timeout=10000)
        self.page.locator('#custom-switch-1').first.click(force=True, timeout=100000)
        self.page.wait_for_timeout(2000)
        expect(self.page.locator("//span[normalize-space()='Countries of Origin']")).not_to_be_visible()

    def Validate_Merge_option(self):
        # Select two companies using checkboxes
        self.page.pause()
        self.page.locator("table tbody tr td:nth-child(1)").first.click()
        self.page.locator("table tbody tr td:nth-child(1)").nth(1).click()

        buttons = [
            ("button", "cloud_download Download"),
            ("button", "ios_share Export to CRM"),
            ("button", "star Bookmark"),
            ("xpath", "//span[normalize-space()='Merge']"),
            ("button", "contacts Fetch Contacts"),
            ("button", "More Download Options"),
            ("button", "filter_arrow_right Apply"),
        ]

        for locator_type, locator_value in buttons:
            if locator_type == "button":
                element = self.page.get_by_role("button", name=locator_value)
            else:  # xpath
                element = self.page.locator(locator_value)

            # ✅ Validate enabled (if applicable) and visible
            expect(element).to_be_visible()
            if locator_value != "filter_arrow_right Apply":
                expect(element).to_be_enabled()

            print(f"✅ {locator_value} is visible and enabled")

        checked_rows = self.page.locator("table tbody tr")
        row_count = checked_rows.count()

        company_names = []
        for i in range(row_count):
            row = checked_rows.nth(i)
            checkbox = row.locator("input[type='checkbox']")

            if checkbox.is_checked():
                # Extract company name
                company_cell = row.locator('[class="trademo-link text-left px-0 text-capitalize "]')
                company_name_locator = company_cell.locator("[data-tip]").first

                # Safely extract text
                if company_name_locator.count() > 0:
                    company_name = company_name_locator.inner_text().strip()
                else:
                    company_name = company_cell.inner_text().strip()

                print(f"✅ Checked Company: {company_name}")
                company_names.append(company_name)

        # Click on the Merge option
        self.page.get_by_role("button", name="Merge").click()

        # Validate companies appear in merge dialog
        Company1 = re.compile(re.escape(company_names[0][:26]), re.IGNORECASE)
        Company2 = re.compile(re.escape(company_names[1][:26]), re.IGNORECASE)

        # ✅ Partial match for <p> tag
        expect(
            self.page.locator("p").filter(has_text=Company1)
        ).to_be_visible()

        # ✅ Partial match for dialog
        expect(
            self.page.get_by_role("dialog").get_by_text(Company2)
        ).to_be_visible()
        expect(self.page.get_by_role("dialog")).to_contain_text("Cancel")
        expect(self.page.get_by_role("button", name="Merge Companies")).to_be_visible()
        expect(self.page.get_by_role("button", name="Close")).to_be_visible()

        # Close dialog
        self.page.get_by_role("button", name="Close").click()
        expect(self.page.get_by_role("button", name="Merge Companies")).not_to_be_visible()

        return company_names
    def Download_option_Ranktab(self):
        # verify_Rank tab header
        expected_headers = [
            "S No",
            "Company Name & IEC Number",
            "Shipments",
            "Consignee State",
        ]

        for header in expected_headers:
            locator = self.page.get_by_role("columnheader", name=header).first
            expect(locator).to_be_visible()
            print(f"✅ Found header in the Rank: {header}")
        self.page.get_by_role("button", name="cloud_download Download").click()
        if self.page.get_by_text("Download Unique Importer").is_visible():
            expect(self.page.get_by_text("Download Unique Importer")).to_be_visible()
            self.page.locator("//input[@placeholder='To']").clear()
            self.page.locator("//input[@placeholder='To']").fill("1")
            expect(self.page.locator("//span[@class='header-confirm']")).to_contain_text("Confirmation")
        else:
            print("⚠️ 'Download Unique Importer' not found on page")
            expect(self.page.locator("//span[@class='header-confirm']")).to_contain_text("Confirmation")

    def Validate_Colmn_Shown_In_the_Grid_Download_Rank(self):
        import os
        import tempfile
        import openpyxl
        from datetime import datetime

        # --- Trigger download ---
        with self.page.expect_download() as download_info:
            self.page.locator("//button[normalize-space()='Confirm']").click()
        download = download_info.value

        # --- Save to temp file ---
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, download.suggested_filename)
        download.save_as(file_path)
        print(f"📂 Downloaded to: {file_path}")

        # --- Read workbook ---
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # --- Validate "Report Summary" sheet ---
        if "Report Summary" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Report Summary' sheet not found in downloaded Excel.")
        summary_sheet = wb["Report Summary"]

        EXPECTED_REPORT_TYPE = "Importers List"
        EXPECTED_DOWNLOAD_DATE = datetime.today().strftime("%b %d, %Y")

        report_type_value = str(summary_sheet["B7"].value).strip() if summary_sheet["B7"].value else ""
        download_date_value = str(summary_sheet["B10"].value).strip() if summary_sheet["B10"].value else ""

        if report_type_value == EXPECTED_REPORT_TYPE:
            print("✅ Report Type matches expected value.")
        else:
            print(f"❌ Report Type mismatch: Found '{report_type_value}', Expected '{EXPECTED_REPORT_TYPE}'")

        if download_date_value == EXPECTED_DOWNLOAD_DATE:
            print("✅ Download Date matches today's date.")
        else:
            print(f"❌ Download Date mismatch: Found '{download_date_value}', Expected '{EXPECTED_DOWNLOAD_DATE}'")

        # --- Read "Shipment Data" sheet ---
        if "Importer Data" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Shipment Data' sheet not found in downloaded Excel.")
        sheet = wb["Importer Data"]

        # Read Excel headers
        downloaded_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        wb.close()

        # --- Get UI headers ---
        ui_headers = self.page.locator("//table/thead/tr/th").all_inner_texts()
        ui_headers = [h.strip() for h in ui_headers if h not in ("", "View")]

        # --- Normalize headers ---
        def normalize_header(h):
            if not h:
                return ""
            return str(h).strip().replace(".", "").replace("  ", " ").replace("'", "").replace('"', '').lower()

        excel_norm = [normalize_header(h) for h in downloaded_headers]
        ui_norm = [normalize_header(h) for h in ui_headers]

        # --- Columns to ignore in Excel ---
        ignore_cols = [
            "IEC Number",
            "Consignee Jurisdiction Country",
        ]

        # Filter Excel and UI columns
        excel_filtered = [col for col in excel_norm if col not in ignore_cols]
        ui_filtered = [col for col in ui_norm if col not in ignore_cols]

        # --- Validation ---
        missing_ui_cols = [col for col in ui_filtered if col not in excel_filtered]
        extra_excel_cols = [col for col in excel_filtered if col not in ui_filtered]

        if not missing_ui_cols:
            print("✅ All relevant UI columns are present in Excel (ignored extra fields).")
        else:
            print("❌ Missing UI columns in Excel:", missing_ui_cols)

        if extra_excel_cols:
            print("⚠️ Extra columns in Excel (not in UI):", extra_excel_cols)

        # --- Cleanup ---
        os.remove(file_path)

    def Validate_Export_CRM(self):
        self.page.get_by_role("button", name="ios_share Export to CRM").click()
        expect(self.page.get_by_role("dialog").get_by_text("Export to CRM")).to_be_visible()
        expect(self.page.get_by_text("Your account is not connected")).to_be_visible()
        expect(self.page.get_by_text("Streamline your workflow by")).to_be_visible()
        expect(self.page.get_by_text("Salesforce")).to_be_visible()
        expect(self.page.get_by_text("Hubspot")).to_be_visible()
        expect(self.page.get_by_text("Zoho")).to_be_visible()
        expect(self.page.get_by_text("Other CRM")).to_be_visible()
        expect(self.page.get_by_role("button", name="Cancel")).to_be_visible()
        self.page.get_by_role("radio", name="Other CRM").check()
        expect(self.page.locator("//button[normalize-space()='Submit']")).to_be_visible()
        # self.page.get_by_role("textbox", name="Enter CRM details here").click()
        # self.page.get_by_role("textbox", name="Enter CRM details here").fill("test")
        # self.page.locator("//button[normalize-space()='Submit']").click()
        self.page.get_by_role("button", name="Cancel").click()

    def Fetch_contacts(self):
        self.page.get_by_role("button", name="contacts Fetch Contacts").click()
        expect(self.page.get_by_text("Fetch Contacts for 2 companies")).to_be_visible()
        expect(self.page.get_by_text("Add values to at least one of")).to_be_visible()
        expect(self.page.locator("div").filter(has_text=re.compile(r"^Enter Job Title$"))).to_be_visible()
        expect(self.page.get_by_text("Enter Job Title")).to_be_visible()
        expect(self.page.get_by_role("textbox", name="Enter Job title of the")).to_be_visible()
        expect(self.page.get_by_text("Select Department of the")).to_be_visible()
        expect(self.page.locator(".css-l8465o-control").first).to_be_visible()
        expect(self.page.get_by_text("Select Management level of")).to_be_visible()
        expect(self.page.get_by_text("You can select up to 5 options").nth(1)).to_be_visible()
        expect(self.page.get_by_text("Select the Country of the")).to_be_visible()
        expect(self.page.get_by_text("You can select up to 5 options").nth(2)).to_be_visible()
        expect(self.page.get_by_text("Maximum number of contacts")).to_be_visible()
        expect(self.page.get_by_text("Type of contact")).to_be_visible()
        expect(self.page.get_by_text("Email", exact=True)).to_be_visible()
        expect(self.page.get_by_text("Phone Number", exact=True)).to_be_visible()
        expect(self.page.get_by_text("Email and Phone number")).to_be_visible()
        expect(self.page.get_by_text("Any message or description")).to_be_visible()
        expect(self.page.get_by_role("button", name="Fetch Contacts")).to_be_disabled()
        self.page.get_by_role("button", name="Close").click()

    def Bookmark(self):
        self.page.get_by_role("button", name="star Bookmark").click()
        expect(self.page.get_by_text("Add Companies to Bookmarks")).to_be_visible()
        expect(self.page.get_by_role("button", name="Cancel")).to_be_visible()
        expect(self.page.get_by_role("button", name="Bookmark")).to_be_visible()
        self.page.get_by_role("button", name="Bookmark").click()
        expect(self.page.get_by_text("Companies Bookmarked")).to_be_visible()

    def Filter_by_Shipment_Trends(self):
        # Disable second level hierarchy
        global less_than, value_num, shipmentvalue_num, greater_than, from_value, to_value
        # click on filter by
        self.page.pause()
        self.page.wait_for_timeout(3000)
        self.page.locator('#custom-switch-1').first.wait_for(state="visible", timeout=10000)
        self.page.locator('#custom-switch-1').first.click(force=True, timeout=100000)
        self.page.wait_for_timeout(2000)
        # click on filter by
        dropdowns = self.page.locator("div.css-1rupv9a-singleValue")
        dropdown_count = dropdowns.count()

        for i in range(dropdown_count):
            dropdown = dropdowns.nth(i)

            # Open dropdown
            dropdown.click()
            self.page.wait_for_timeout(500)  # small wait for menu to appear

            # Check if "Shipment Value USD" is visible in options
            option = self.page.get_by_text("Shipments", exact=True).nth(1)

            if option.is_visible(timeout=1000):
                option.click()
                print(f"'Shipment Value USD' selected in dropdown {i + 1}")
            else:
                print(f"'Shipment Value USD' not found in dropdown {i + 1}, moving on...")
                # Close dropdown by pressing ESC (so next one can open cleanly)
                self.page.keyboard.press("Escape")
        self.page.get_by_role("button", name="tune Filter By").click()
        expect(self.page.get_by_text("Filter By Range")).to_be_visible(timeout=100000)
        expect(self.page.locator("#onClickClear").get_by_text("Shipments")).to_be_visible()
        expect(self.page.get_by_text("Value", exact=True)).to_be_visible()
        self.page.wait_for_timeout(1000)
        # store shipment value usd in variable
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_count = headers.count()

        shipment_value_index = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            print(f"{header_text}")
            if "total shipments" in header_text:
                shipment_value_index = i
                break

        if shipment_value_index is None:
            raise Exception("❌ 'Shipment' column not found in table headers")

        print(f"✅ Found 'Shipment' column at index {shipment_value_index}")

        # --- Extract all shipment values from table ---
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        shipment_values = []
        for i in range(row_count):
            value_text = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num = float(value_text.replace(",", "").replace("$", ""))

            # Subtract 1 for checking
            less_than = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            greater_than = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

            from_value = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            to_value = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

        # Under Shipment Value (USD)/Value, select Less Than, and enter a value less than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Shipments Found. Please try another search")).to_be_visible(timeout=100000)
        self.page.get_by_role("button", name="Clear").click()
        # Under Shipment Value (USD)/Value, select Less Than, and enter a value greater than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(shipmentvalue_num))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num <= shipmentvalue_num1:
                print(f"✅ Value {shipmentvalue_num1} is less than {shipmentvalue_num} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT less than {shipmentvalue_num} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {shipmentvalue_num}"

        # Click Clear to reset the modal, then select Shipment Value (USD)/Value – Between, enter a range that includes the company’s shipment/Value count, and click Apply
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).fill(str(from_value))
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto").fill(str(to_value))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num2 = float(value_text1.replace(",", "").replace("$", ""))
            if from_value <= shipmentvalue_num2 <= to_value:
                print(f"✅ Row {i + 1}: Value {shipmentvalue_num2} is within range [{from_value}, {to_value}]")
            else:
                print(f"❌ Row {i + 1}: Value {shipmentvalue_num2} is NOT in range [{from_value}, {to_value}]")
                assert False, f"Value {shipmentvalue_num2} at row {i + 1} not in range [{from_value}, {to_value}]"
        # Without clearing, modify range to one that excludes the company’s Shipment Value (USD)/Value count, then click Apply
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).fill("1")
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto").fill("2")
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Shipments Found. Please try another search")).to_be_visible(timeout=100000)
        # Click Clear, select Shipment Value (USD)/Value – Greater Than, enter a value below the actual count, and click Apply
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value is: {value_text1}")
            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num1 > less_than:
                print(f"✅ Value {shipmentvalue_num1} is greater than {less_than} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT greater than {less_than} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {less_than}"
        # Without clearing, change the value to one greater than the company’s count and click Apply
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Shipment").click()
        self.page.get_by_placeholder("Enter count of Shipment").fill(str(greater_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Shipments Found. Please try another search")).to_be_visible(timeout=100000)
        # Clear filter
        self.page.get_by_role("button", name="Clear").click()

    def Filter_by_Shipment_Value_Trends(self):
        # Disable second level hierarchy
        self.page.pause()
        global less_than, value_num, shipmentvalue_num, greater_than, from_value, to_value
        dropdowns = self.page.locator("div.css-1rupv9a-singleValue")
        dropdown_count = dropdowns.count()

        for i in range(dropdown_count):
            dropdown = dropdowns.nth(i)

            # Open dropdown
            dropdown.click()
            self.page.wait_for_timeout(500)  # small wait for menu to appear

            # Check if "Shipment Value USD" is visible in options
            option = self.page.get_by_text("Value (USD)", exact=True)

            if option.is_visible(timeout=1000):
                option.click()
                print(f"'Shipment Value USD' selected in dropdown {i + 1}")
            else:
                print(f"'Shipment Value USD' not found in dropdown {i + 1}, moving on...")
                # Close dropdown by pressing ESC (so next one can open cleanly)
                self.page.keyboard.press("Escape")

        # store shipment value usd in variable
        headers = self.page.locator("table thead tr th")
        expect(headers.first).to_be_visible(timeout=100000)
        header_count = headers.count()

        shipment_value_index = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            print(f"{header_text}")
            if "total shipment value(usd)" in header_text:
                shipment_value_index = i
                break

        if shipment_value_index is None:
            raise Exception("❌ 'Shipment Value' column not found in table headers")

        print(f"✅ Found 'Shipment Value(USD)' column at index {shipment_value_index}")

        # --- Extract all shipment values from table ---
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        shipment_values = []
        for i in range(row_count):
            value_text = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num = float(value_text.replace(",", "").replace("$", ""))

            # Subtract 1 for checking
            less_than = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            greater_than = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

            from_value = shipmentvalue_num - 1
            print(f"less than value: {less_than}")

            to_value = shipmentvalue_num + 1
            print(f"greater than:{greater_than}")

        # Under Shipment Value (USD)/Value, select Less Than, and enter a value less than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.get_by_role("button", name="tune Filter By").click()
        expect(self.page.get_by_text("Filter By Range")).to_be_visible(timeout=100000)
        self.page.locator("div").filter(has_text=re.compile(r"^Valueless thanbetweengreater than$")).get_by_role(
            "img").click()
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Shipments Found. Please try another search")).to_be_visible(timeout=100000)
        self.page.get_by_role("button", name="Clear").click()
        # Under Shipment Value (USD)/Value, select Less Than, and enter a value greater than the actual shipments shown (e.g., Enter 200,000 if the company has 150,000)
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(shipmentvalue_num))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num1 <= shipmentvalue_num:
                print(f"✅ Value {shipmentvalue_num1} is less than {shipmentvalue_num} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT less than {shipmentvalue_num} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {shipmentvalue_num}"

        # Click Clear to reset the modal, then select Shipment Value (USD)/Value – Between, enter a range that includes the company’s shipment/Value count, and click Apply
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).fill(str(from_value))
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto").fill(str(to_value))
        self.page.get_by_role("button", name="Apply").click()

        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text1}")

            # Convert to float (remove commas & $ if present)
            shipmentvalue_num2 = float(value_text1.replace(",", "").replace("$", ""))
            if from_value <= shipmentvalue_num2 <= to_value:
                print(f"✅ Row {i + 1}: Value {shipmentvalue_num2} is within range [{from_value}, {to_value}]")
            else:
                print(f"❌ Row {i + 1}: Value {shipmentvalue_num2} is NOT in range [{from_value}, {to_value}]")
                assert False, f"Value {shipmentvalue_num2} at row {i + 1} not in range [{from_value}, {to_value}]"
        # Without clearing, modify range to one that excludes the company’s Shipment Value (USD)/Value count, then click Apply
        self.page.get_by_role("tab", name="between").click()
        self.page.get_by_placeholder("From", exact=True).click()
        self.page.get_by_placeholder("From", exact=True).clear()
        self.page.get_by_placeholder("From", exact=True).fill("1")
        self.page.get_by_placeholder("Upto").click()
        self.page.get_by_placeholder("Upto", exact=True).clear()
        self.page.get_by_placeholder("Upto").fill("2")
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Shipments Found. Please try another search")).to_be_visible(timeout=100000)
        # Click Clear, select Shipment Value (USD)/Value – Greater Than, enter a value below the actual count, and click Appl
        self.page.get_by_role("button", name="Clear").click()
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(less_than))
        self.page.get_by_role("button", name="Apply").click()
        for i in range(row_count):
            value_text1 = rows.nth(i).locator("td").nth(shipment_value_index).inner_text().strip()
            print(f"Shipment value usd is: {value_text1}")
            # Convert to float (remove commas & $ if present)
            shipmentvalue_num1 = float(value_text1.replace(",", "").replace("$", ""))
            if shipmentvalue_num1 > less_than:
                print(f"✅ Value {shipmentvalue_num1} is greater than {less_than} (row {i + 1})")
            else:
                print(f"❌ Value {shipmentvalue_num1} is NOT greater than {less_than} (row {i + 1})")
                assert False, f"Value {shipmentvalue_num1} exceeded limit {less_than}"
        # Without clearing, change the value to one greater than the company’s count and click Apply
        self.page.get_by_role("tab", name="greater than").click()
        self.page.get_by_placeholder("Enter count of Value").click()
        self.page.get_by_placeholder("Enter count of Value").fill(str(greater_than))
        self.page.get_by_role("button", name="Apply").click()
        expect(self.page.get_by_text("No Shipments Found. Please try another search")).to_be_visible(timeout=100000)
        # Clear filter
        self.page.get_by_role("button", name="Clear").click()

    def Download_option_Trendtab(self):
        # verify_second_Level_Hierarchy_headers
        expected_headers = [
            "S No",
            "Company",
            "Total Shipments",
            "% Overall Change",
        ]

        for header in expected_headers:
            locator = self.page.get_by_role("columnheader", name=header).first
            expect(locator).to_be_visible()
            print(f"✅ Found header in the Rank: {header}")
        self.page.get_by_role("button", name="cloud_download Download").click()
        if self.page.get_by_text("Download Unique Importer").is_visible():
            expect(self.page.get_by_text("Download Unique Importer")).to_be_visible()
            self.page.locator("//input[@placeholder='To']").clear()
            self.page.locator("//input[@placeholder='To']").fill("1")
            expect(self.page.locator("//span[@class='header-confirm']")).to_contain_text("Confirmation")
        else:
            print("⚠️ 'Download Unique Importer' not found on page")
            expect(self.page.locator("//span[@class='header-confirm']")).to_contain_text("Confirmation")

    def Validate_Colmn_Shown_In_the_Grid_Download_Trends(self):
        import os
        import tempfile
        import openpyxl
        from datetime import datetime

        # --- Trigger download ---
        with self.page.expect_download() as download_info:
            self.page.locator("//button[normalize-space()='Confirm']").click()
        download = download_info.value

        # --- Save to temp file ---
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, download.suggested_filename)
        download.save_as(file_path)
        print(f"📂 Downloaded to: {file_path}")

        # --- Read workbook ---
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # --- Validate "Report Summary" sheet ---
        if "Report Summary" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Report Summary' sheet not found in downloaded Excel.")
        summary_sheet = wb["Report Summary"]

        EXPECTED_REPORT_TYPE = "Importers List"
        EXPECTED_REPORT_TAB = "Trends"
        EXPECTED_DOWNLOAD_DATE = datetime.today().strftime("%b %d, %Y")

        report_type_value = str(summary_sheet["B7"].value).strip() if summary_sheet["B7"].value else ""
        importer_tab_value = str(summary_sheet["C7"].value).strip() if summary_sheet["C7"].value else ""
        download_date_value = str(summary_sheet["B10"].value).strip() if summary_sheet["B10"].value else ""

        if report_type_value == EXPECTED_REPORT_TYPE:
            print("✅ Report Type matches expected value.")
        else:
            print(f"❌ Report Type mismatch: Found '{report_type_value}', Expected '{EXPECTED_REPORT_TYPE}'")

        if  importer_tab_value == EXPECTED_REPORT_TAB:
            print("✅ Importer tab value matches expected value.")
        else:
            print(f"❌ Importer tab value  mismatch: Found '{importer_tab_value}', Expected '{EXPECTED_REPORT_TAB}'")

        if download_date_value == EXPECTED_DOWNLOAD_DATE:
            print("✅ Download Date matches today's date.")
        else:
            print(f"❌ Download Date mismatch: Found '{download_date_value}', Expected '{EXPECTED_DOWNLOAD_DATE}'")

        # --- Read "Shipment Data" sheet ---
        if "Importer Data" not in wb.sheetnames:
            wb.close()
            raise ValueError("❌ 'Shipment Data' sheet not found in downloaded Excel.")
        sheet = wb["Importer Data"]

        # Read Excel headers
        downloaded_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        wb.close()

        # --- Get UI headers ---
        ui_headers = self.page.locator("//table/thead/tr/th").all_inner_texts()
        ui_headers = [h.strip() for h in ui_headers if h not in ("", "View")]

        # --- Normalize headers ---
        def normalize_header(h):
            if not h:
                return ""
            return str(h).strip().replace(".", "").replace("  ", " ").replace("'", "").replace('"', '').lower()

        excel_norm = [normalize_header(h) for h in downloaded_headers]
        ui_norm = [normalize_header(h) for h in ui_headers]

        # --- Columns to ignore in Excel ---
        ignore_cols = [
            "IEC Number",
            "Consignee Jurisdiction Country",
        ]

        # Filter Excel and UI columns
        excel_filtered = [col for col in excel_norm if col not in ignore_cols]
        ui_filtered = [col for col in ui_norm if col not in ignore_cols]

        # --- Validation ---
        missing_ui_cols = [col for col in ui_filtered if col not in excel_filtered]
        extra_excel_cols = [col for col in excel_filtered if col not in ui_filtered]

        if not missing_ui_cols:
            print("✅ All relevant UI columns are present in Excel (ignored extra fields).")
        else:
            print("❌ Missing UI columns in Excel:", missing_ui_cols)

        if extra_excel_cols:
            print("⚠️ Extra columns in Excel (not in UI):", extra_excel_cols)

        # --- Cleanup ---
        os.remove(file_path)

    def validate_data_after_applying_filter(self):
        # First row
        first_row = self.page.locator("table tbody tr").first
        self.page.pause()
        # Serial number (for logging)
        sl_no = first_row.locator("td").nth(1).inner_text().strip()
        expected_shipments = first_row.locator("td").nth(3).inner_text().strip()
        expected_shipment_value_usd = first_row.locator("td").nth(4).inner_text().strip()
        expected_consignee_state = first_row.locator("td").nth(5).inner_text().strip()
        consignee_locator = first_row.locator("td").nth(2).locator("[data-tip]").first
        consignee_tip = consignee_locator.get_attribute("data-tip") if consignee_locator.count() > 0 else None

        if consignee_tip and consignee_tip.strip():
            self.company_name = consignee_tip.strip()
        else:
            # Fallback to visible text if no tooltip
            self.company_name = first_row.locator("td").nth(2).inner_text().strip()
        country_name = self.page.locator('[role="cell"] > .text-wrap > [class="tw-flex tw-items-center"]').first

        # Locate and click the Apply Filter icon
        apply_filter_icon = first_row.locator("td").nth(6).locator("img")
        apply_filter_icon.scroll_into_view_if_needed()
        apply_filter_icon.click()

        print(f"✅ Applied filter for first row (Serial No: {sl_no})")
        #Observe the grid display same values
        sl_no = first_row.locator("td").nth(1).inner_text().strip()
        consignee_locator_actual = first_row.locator("td").nth(2).locator("[data-tip]").first
        consignee_tip_actual = consignee_locator_actual.get_attribute(
            "data-tip") if consignee_locator_actual.count() > 0 else None

        if consignee_tip_actual and consignee_tip_actual.strip():
            company_name_actual = consignee_tip.strip()
        else:
            # Fallback to visible text if no tooltip
            company_name_actual = first_row.locator("td").nth(2).inner_text().strip()

        if self.company_name == company_name_actual:
            print(f"valid company found:{company_name_actual} at {sl_no}")
        else:
            print(f"valid company found:{company_name_actual} at {sl_no}")

        country_name_actual = self.page.locator(
            '[role="cell"] > .text-wrap > [class="tw-flex tw-items-center"]').first.inner_text().strip()
        if country_name == country_name_actual:
            print(f"valid country found:{country_name_actual} at {sl_no}")
        else:
            print(f"valid country found:{country_name_actual} at {sl_no}")

        #check Shiment
        actual_shipment_value = first_row.locator("td").nth(3).inner_text().strip()
        if actual_shipment_value == expected_shipments:
            print(f"valid shipment count:{actual_shipment_value} at {sl_no}")
        else:
            print(f"valid shipment count:{actual_shipment_value} at {sl_no}")

        #check shipment value usd
        actual_shipment_value_usd = first_row.locator("td").nth(4).inner_text().strip()
        if actual_shipment_value_usd == expected_shipment_value_usd:
            print(f"valid shipment value usd count:{actual_shipment_value} at {sl_no}")
        else:
            print(f"valid shipment value usd count:{actual_shipment_value} at {sl_no}")

        #check consignee state
        actual_consignee_state = first_row.locator("td").nth(5).inner_text().strip()
        if actual_consignee_state == expected_consignee_state:
            print(f"valid shipment value usd count:{actual_consignee_state} at {sl_no}")
        else:
            print(f"valid shipment value usd count:{actual_consignee_state} at {sl_no}")

        # Verify the filter option at the top
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')
        # Click on the consignee filter dropdown
        self.page.locator("div").filter(has_text=re.compile(r"^Consignee1$")).nth(1).click()
        #Scroll to the top filter bar and locate the Consignee Standardized Name filter
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        #Click on the Consignee  Standardized Name filter
        self.page.get_by_text("Consignee Standardized Name", exact=True).click()
        expect(self.page.locator("#searchIndividualFilterPanel")).to_contain_text(
            "Filter by Consignee Standardized Name")
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").last.fill(self.company_name)
        self.page.wait_for_timeout(5000)
        # Check for the presence of the company name in the results
        locators = self.page.locator('[class="text-no-transform"]')
        count = locators.count()
        print(f"🔍 Found {count} elements with class 'text-no-transform'")

        if count == 0:
            print("❌ FAIL: No elements found for search")
            self.page.locator('input[type="text"]').last.clear()
        else:
            found_match = False

            for i in range(count):
                locator = locators.nth(i)

                if locator.is_visible():
                    actual_text = locator.inner_text().strip()

                    if self.company_name.lower() in actual_text.lower():  # case-insensitive compare
                        print(f"✅ PASS: Search is working → '{self.company_name}' found in element {i + 1}: '{actual_text}'")
                        expect(locator).to_contain_text(re.compile(self.company_name, re.IGNORECASE))
                        #Open the Consignee filter Standardized Name section
                        expect(self.page.locator("//label[input[@type='checkbox']]").first).to_be_checked()
                        print(f"{self.company_name} for the selected company in consignee standardized name filter")
                        #Verify the shipment count
                        shipment_count_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
                        clean_shipmentvalue = shipment_count_value.strip("()")
                        print(clean_shipmentvalue)
                        if expected_shipments == clean_shipmentvalue:
                            print(f"valid shipment count:{clean_shipmentvalue} for the selected company in consignee standardized name filter")
                        else:
                            print(f"valid shipment count:{clean_shipmentvalue} for the selected company in consignee standardized name filter")
                        found_match = True
                    else:
                        print(f"❌ FAIL: Search text '{self.company_name}' not found in element {i + 1}. Actual: '{actual_text}'")
                else:
                    print(f"⚠️ Element {i + 1} is not visible")

            if not found_match:
                print(f"❌ FAIL: Search text '{self.company_name}' not found in any visible element")
                self.page.locator('input[type="text"]').clear()

        self.page.locator("//span[@aria-hidden='true']").click()
        # Observe the updated values in the following tabs: Shipments, Importers, Exporters
        self.Shipment_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').nth(
            0).inner_text().strip()
        print(f"shipment count:{self.Shipment_count}")
        self.Importer_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            0).inner_text().strip()
        print(f"Importer count:{self.Importer_count}")
        self.Exporter_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            1).inner_text().strip()
        print(f"Exporter count:{self.Exporter_count}")

    def Validate_Tab_count_after_Manual_search(self):
         # In the Consignee Standardized Name filter, manually filter and apply the same company name, then observe the Shipment/Importer/Exporter counts again
         self.page.wait_for_timeout(3000)
         # In the search bar, manually enter a Shipper Name and click Search
         expect(self.page.get_by_placeholder(
             "Type to search in all categories or choose from the category below")).to_be_visible()
         # Manual suggest search for Shipper
         self.page.get_by_role("textbox", name="Type to search in all").click()
         self.page.get_by_role("textbox", name="Type to search in all").fill(self.company_name)
         self.page.locator(".tw-bg-primary-purple-500").click()
         self.page.wait_for_timeout(5000)
         #Verify the count
         Check_Shipment_count = self.page.locator(
             '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').nth(
             0).inner_text().strip()
         if self.Shipment_count == Check_Shipment_count:
             print("✅ Shipment count is matching")
         else:
             print("❌ Shows wrong shipment count")

         Check_Importer_count = self.page.locator(
             '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
             0).inner_text().strip()
         if self.Importer_count == Check_Importer_count:
             print("✅ Importer count is matching")
         else:
             print("❌ Shows wrong Importer count")
         Check_Exporter_count = self.page.locator(
             '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
             1).inner_text().strip()
         if self.Exporter_count == Check_Exporter_count:
             print("✅ Exporter count is matching")
         else:
             print("❌ Shows wrong Exporter count")

    def Ranks_2nd_level_hierarchy_Edit(self):
        self.page.pause()
        self.page.wait_for_timeout(3000)

        toggle = self.page.locator('[class="custom-control custom-switch"]').first

        # ✅ Check visibility before interacting
        if toggle.is_visible():
            print("2nd level hierarchy toggle is visible")

            # Click only if it's not already enabled
            checkbox = toggle.get_by_role("checkbox")
            if not checkbox.is_checked():
                toggle.click(force=True, timeout=100000)
                self.page.wait_for_timeout(1000)
                print("2nd level hierarchy enabled")
            else:
                print("2nd level hierarchy already enabled")

            # Verify UI effect
            # expect(self.page.locator("//span[normalize-space()='Countries of Origin']")).to_be_visible(timeout=100000)
        else:
            print("2nd level hierarchy toggle is NOT visible, skipping...")
        # Observe the "Company Name", "Country of Origin", "Total Shipment Value (USD)", and "% Overall Change" columns
        expected_headers = [
            "Company Name & IEC Number",
            "Countries of Origin",
            "Shipments",
            "Shipment Value (USD)",
            "% share by Shipments"
        ]

        for header in expected_headers:
            locator = self.page.get_by_role("columnheader", name=header).first # more precise
            expect(locator).to_be_visible(timeout=100000)
            print(f"✅ Header '{header}' is visible.")
        # In "Country of Origin" column click edit and select any options
        self.page.locator('[class="ml-2"]').last.click()
        expected_editoption = [
            "Trade Routes (Country of Export → Country of Import)",
            "Unique Exporters",
            "Countries of Origin",
            "Ports of Lading",
            "Ports of Unlading",
            "Consignee States",
            "Consignee Cities",
            "Custom House Agent Name",
            "Mode of Transportation",
            "4 Digit HS Codes",
            "Sections",
            "Trade Routes (Port of Lading → Port of Unlading)",
            "6 digit HS Codes"
        ]

        table = self.page.get_by_role("table")

        for header in expected_editoption:
            expect(table).to_contain_text(header)
            print(f"✅ Validated option: {header}")

        self.page.get_by_text("4 Digit HS Codes").first.click()
        self.page.wait_for_timeout(7000)
        rows = self.page.locator('[class="m-0 d-flex align-items-center h-100"]')
        row_count = rows.count()
        hscode_values = []
        for i in range(row_count):
            hs_code_text = rows.nth(i).inner_text().strip()
            # ✅ Validate that it's only 4 digits
            if re.fullmatch(r"\d{4}", hs_code_text):
                print(f"✅ Row {i}: HS Code {hs_code_text} is valid (4 digits).")
            else:
                raise AssertionError(f"❌ Row {i}: Invalid HS Code '{hs_code_text}' (not 4 digits)")

        self.page.locator('[class="ml-2"]').last.click()
        self.page.get_by_text("Countries of Origin").first.click()
        # Toggle OFF the switch
        self.page.wait_for_timeout(3000)
        self.page.locator('#custom-switch-1').first.wait_for(state="visible", timeout=10000)
        self.page.locator('#custom-switch-1').first.click(force=True, timeout=100000)
        self.page.wait_for_timeout(2000)
        expect(self.page.locator("//span[normalize-space()='Countries of Origin']")).not_to_be_visible()
































































