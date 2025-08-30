import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page
from playwright.sync_api import expect

invalid_entries =[]
import difflib
from decimal import Decimal, ROUND_HALF_UP
def format_number(num_str: str) -> str:
    """Convert number string like '149,760' → '149.76K', '268,000' → '268K'.
       Uses ROUND_HALF_UP and trims unnecessary trailing zeros."""
    num_str = num_str.replace(",", "").strip()

    try:
        num = int(num_str)
    except ValueError:
        raise ValueError(f"Invalid number string: {num_str}")

    def format_decimal(val: Decimal, suffix: str) -> str:
        # Round to 2 decimals
        val = val.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # Convert to string and strip trailing zeros/decimal if not needed
        val_str = f"{val:.2f}".rstrip("0").rstrip(".")
        return f"{val_str}{suffix}"

    if num >= 1_000_000:
        return format_decimal(Decimal(num) / Decimal(1_000_000), "M")
    elif num >= 1_000:
        return format_decimal(Decimal(num) / Decimal(1_000), "K")
    else:
        return str(num)
class PortFilter:
    def __init__(self, page: Page):
        self.C_O_field = None
        self.C_E_field = None
        self.S_JC_field = None
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def check_consignee_filter(self):
        # Wait for the Consignee Standardized Name label
        consignee_standardized_locator = self.page.locator(
            "//div[contains(text(),'Consignee Standardized Name')]"
        )
        expect(consignee_standardized_locator).to_be_visible(timeout=100000)

        # Locate Consignee filter options
        elements = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count = elements.count()
        print(f"Found {count} Consignee filter options (before toggle)")

        # Loop through each element
        for i in range(count):
            element = elements.nth(i)
            if element.is_visible():
                text = element.inner_text()
                print(f"✅ Consignee option {i + 1} is visible: {text}")
            else:
                print(f"❌ Consignee option {i + 1} is not visible")

        # Click on the Consignee filter again (collapse/close)
        self.page.locator("//span[normalize-space()='Consignee']").click()
        self.page.wait_for_timeout(2000)  # give time for collapse

        # Re-locate elements again after DOM change
        elements_after = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count_after = elements_after.count()
        print(f"Checking {count_after} Consignee filter options")

        # Expect no options should be visible
        for i in range(count_after):
            element = elements_after.nth(i)
            expect(element).not_to_be_visible(timeout=5000)

    def open_port_dropdown(self):
        port_locator = self.page.locator("//div[contains(text(),'Port Of Lading')]")

        # Check if "port lading" is visible
        if not port_locator.is_visible():
            print("📌 port dropdown is closed. Clicking to open...")
            self.page.locator("//span[normalize-space()='Ports']").click()
        else:
            print("✅ port dropdown is already open.")

    def Click_Port_Lading(self):
        #Ports filter modal opens with sub-filters such as:
        expect(self.page.locator("//div[contains(text(),'Port Of Lading')]")).to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Port Of Unlading')]")).to_be_visible()
        self.page.locator("//div[contains(text(),'Port Of Lading')]").click()
        self.page.wait_for_timeout(2000)
        expect(self.page.get_by_text("Filter by Port Of Lading").first).to_be_visible()

    def Port_Filter_Search(self, country: str):
        #Use the search bar in the modal to search for a port (e.g., "Shanghai")
        self.page.wait_for_timeout(1000)
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").fill(country, timeout=100000)
        self.page.locator("//span[@data-for='localize-filter-tooltip']").first.wait_for(state="visible", timeout=10000)
        # Get search results
        elements = self.page.locator("//span[@data-for='localize-filter-tooltip']")
        count = elements.count()

        found = False
        for i in range(count):
            text = elements.nth(i).text_content()  # safer than inner_text()
            if text:
                text = text.strip().lower()
                print(f"🔎 Found element text: {text}")  # debug log

                if country.lower() in text:
                    print(f"✅ Searched value '{country}' is shown in the filter → {text}")
                    found = True
                    break

        if not found:
            print(f"❌ Searched value '{country}' not found in filter list")

        self.page.locator('input[type="text"]').last.clear()
        self.page.locator("//span[@data-for='localize-filter-tooltip']").first.wait_for(state="visible", timeout=10000)
        # Click on the checkbox of the first result
        all_checkboxes = self.page.locator("//span[@data-for='localize-filter-tooltip']")
        all_count = all_checkboxes.count()

        # Find at least one checkbox that is NOT "Maharashtra"
        found_other = False
        for i in range(all_count):
            name = all_checkboxes.nth(i).inner_text().strip()
            if country not in name:
                found_other = True
                break

        if found_other:
            print("✅ Clearing Port filter input box shows all the checkboxes (not just searched values).")
        else:
            print(f"❌ Only {country} is still visible after clearing — bug in filter.")

    def Port_Apply_Filter(self):
        Portladingwithcount = self.page.locator('[data-for = "localize-filter-tooltip"]').nth(
            0).inner_text().strip()
        match = re.match(r"(.+?)\(([\d,]+)\)", Portladingwithcount)
        if match:
            raw_port = match.group(1).strip()
            self.Total_Record_Count = match.group(2)  # keep commas as-is (string)

            # Save full port name
            self.full_port_name = raw_port

            # Extract only main port name (before comma)
            if "," in raw_port:
                self.selected_port_name = raw_port.split(",")[0].strip()
            else:
                self.selected_port_name = raw_port

            print("✅ Full Port Name:", self.full_port_name)
            print("✅ Main Port Name:", self.selected_port_name)
            print("✅ Value:", self.Total_Record_Count)

        else:
            print("❌ No match found")

        # Convert to float
        port_record_clean = re.sub(r"[()\sA-Za-z,]", "", self.Total_Record_Count)
        port_record_num = float(float(port_record_clean))
        if 0 <= port_record_num < 100_000:
            self.total_in_millions = f"{port_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{port_record_num / 1_000_000:.2f}M"

        # Convert to millions
        total_in_millions = f"{port_record_num / 1_000_000:.2f}"
        print(total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(1).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()

    def Validate_ShipmentRecord_Count(self):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=10000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(3000)
        # Validate total shipment count in 'Shipment' tab
        record_count = self.page.locator(
            "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']"
        )

        # Get actual text
        actual_text = record_count.inner_text()

        # If–else validation
        if self.total_in_millions == actual_text:
            print(f"✅ Record count matches: {actual_text}")
        else:
            print(f"❌ Record count mismatch. Expected {self.total_in_millions}, Found {actual_text}")

    def Validate_Total_RecordCount_In_the_Table(self):
        total_record_text = self.page.locator('[class="trademo-table-count-text"]').inner_text()
        total_from_summary = total_record_text.split("of")[1].strip()
        print(f"🔢 Value after 'of': {total_from_summary}")
        print(f"total record count:{self.Total_Record_Count}")
        if total_from_summary == self.Total_Record_Count:
            print(f"✅ Record count matched: {total_from_summary}")
        else:
            print(f"❌ Record count mismatch! Summary: {total_from_summary}, Table: {self.Total_Record_Count}")

    def Validate_C_S_Name_label(self):
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')

        # Click on the consignee filter dropdown
        self.page.locator("//span[normalize-space()='Ports']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        self.page.locator("//span[normalize-space()='Ports']").click()
        self.page.wait_for_timeout(1000)

    def check_port_lading_inthe_grid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.P_L_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "port of lading":
                self.P_L_field = i

        # --- Handle missing "Consignee JC" column ---
        if self.P_L_field is None:
            raise Exception("❌ 'port of lading' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Port lading: {self.P_L_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                Port_lading = row_data[self.P_L_field]
                Port_lading_norm = re.sub(r"\s+", " ", Port_lading).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.full_port_name.lower() in Port_lading_norm:
                        print(f"✅ Grid valid: {Port_lading} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {Port_lading} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": Port_lading, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    expect(self.page.locator("//span[normalize-space()='Shipment Origin Details']")).to_be_visible(timeout=100000)
                    popup_port = self.page.locator(
                        "//span[normalize-space(text())='Port of Lading'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a").inner_text().strip().lower()

                    if self.full_port_name.lower() in popup_port:
                        print(f"✅ View details valid: {popup_port} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ View details invalid: {popup_port} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in view details"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_port, "Source": "Popup"})

                    Port_lading_link = self.page.locator(
                        "//span[normalize-space(text())='Port of Lading'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a")
                    Port_page1 = Port_lading_link.inner_text().strip()
                    with self.page.expect_popup() as page1_info:
                        self.page.get_by_role("dialog").get_by_text(Port_page1).click()
                    page1 = page1_info.value
                    self.page.wait_for_timeout(2000)
                    expect(page1.locator('[class="_portName_9369m_19"]')).to_contain_text(self.selected_port_name)
                    page1.close()
                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Port lading Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.P_L_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.P_L_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/port_lading_{self.selected_port_name}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid port lading found. Excel not generated.")

        return self.invalid_entries

    def Click_Port_UnLading(self):
        # Ports filter modal opens with sub-filters such as:
        expect(self.page.locator("//div[contains(text(),'Port Of Lading')]")).to_be_visible()
        expect(self.page.locator("//div[contains(text(),'Port Of Unlading')]")).to_be_visible()
        self.page.locator("//div[contains(text(),'Port Of Unlading')]").click()
        self.page.wait_for_timeout(2000)
        expect(self.page.get_by_text("Filter by Port Of Unlading").first).to_be_visible()

    def check_port_unlading_inthe_grid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.P_U_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() == "port of unlading":
                self.P_U_field = i

        # --- Handle missing "Consignee JC" column ---
        if self.P_U_field is None:
            raise Exception("❌ 'port of unlading' column not found ")

        print(f"✅ Found column indexes → S No: {self.S_No}, Port unlading: {self.P_U_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                Port_unlading = row_data[self.P_U_field]
                Port_unlading_norm = re.sub(r"\s+", " ", Port_unlading).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.full_port_name.lower() in Port_unlading_norm:
                        print(f"✅ Grid valid: {Port_unlading} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {Port_unlading} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": Port_unlading, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    expect(self.page.locator("//span[normalize-space()='Shipment Origin Details']")).to_be_visible(
                        timeout=100000)
                    popup_port = self.page.locator(
                        "//span[normalize-space(text())='Port of Unlading'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a").inner_text().strip().lower()

                    if self.full_port_name.lower() in popup_port:
                        print(f"✅ View details valid: {popup_port} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ View details invalid: {popup_port} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid data in view details"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_port, "Source": "Popup"})

                    Port_lading_link = self.page.locator(
                        "//span[normalize-space(text())='Port of Unlading'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a")
                    Port_page1 = Port_lading_link.inner_text().strip()
                    with self.page.expect_popup() as page1_info:
                        self.page.get_by_role("dialog").get_by_text(Port_page1).click()
                    page1 = page1_info.value
                    self.page.wait_for_timeout(2000)
                    expect(page1.locator('[class="_portName_9369m_19"]')).to_contain_text(self.selected_port_name)
                    page1.close()
                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Port lading Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.P_U_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.P_U_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/port_unlading_{self.selected_port_name}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid port unlading found. Excel not generated.")

        return self.invalid_entries



