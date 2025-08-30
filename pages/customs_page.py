import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page
from playwright.sync_api import expect

invalid_entries =[]
from decimal import Decimal, ROUND_HALF_UP
def format_number(num_str: str) -> str:
    """Convert number string like '149,760' → '149.76K', '268,000' → '268K'.
       Uses ROUND_HALF_UP and trims unnecessary trailing zeros."""
    num_str = num_str.replace(",", "").strip()

    try:
        num = int(num_str)
    except ValueError:
        raise ValueError(f"Invalid number string: {num_str}")

    def format_decimal(val: Decimal, suffix: str) -> str:
        # Round to 2 decimals
        val = val.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # Convert to string and strip trailing zeros/decimal if not needed
        val_str = f"{val:.2f}".rstrip("0").rstrip(".")
        return f"{val_str}{suffix}"

    if num >= 1_000_000:
        return format_decimal(Decimal(num) / Decimal(1_000_000), "M")
    elif num >= 1_000:
        return format_decimal(Decimal(num) / Decimal(1_000), "K")
    else:
        return str(num)
class CustomsFilter:
    def __init__(self, page: Page):
        self.C_O_field = None
        self.C_E_field = None
        self.S_JC_field = None
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def check_customs_filter(self):
        # Wait for the customs details label
        Customs_House_Agent_locator = self.page.locator('[class="_filteritemtext_a6k5f_88"]').first
        expect(Customs_House_Agent_locator).to_be_visible(timeout=100000)

        # Locate freight filter options
        elements = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count = elements.count()
        print(f"Found {count} customs filter options")

        for i in range(count):
            element = elements.nth(i)
            if element.is_visible():
                text = element.inner_text()
                print(f"✅ customs house agent option {i + 1} is visible: {text}")
            else:
                print(f"❌ customs house agent option {i + 1} is not visible")

        # Click on the customs filter again (collapse/close)
        self.page.locator("//span[normalize-space()='Customs Details']").click()
        self.page.wait_for_timeout(2000)  # give time for collapse

        # Re-locate elements again after DOM change
        elements_after = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count_after = elements_after.count()
        print(f"Checking {count_after} customs details filter options")

        # Expect no options should be visible
        for i in range(count_after):
            element = elements_after.nth(i)
            expect(element).not_to_be_visible(timeout=5000)

    def open_customs_dropdown(self):
        customs_locator = self.page.locator('[class="_filteritemtext_a6k5f_88"]').first
        self.page.pause()

        # Check if "customs house agent name" is visible
        if not customs_locator.is_visible():
            print("📌 Customs details dropdown is closed. Clicking to open...")
            self.page.locator("//span[normalize-space()='Customs Details']").click()
        else:
            print("✅ freight dropdown is already open.")

    def Click_Customs_House_Agent_Name(self):
        #Mode of trasportation filter modal opens with sub-filters such as:
        expect(self.page.locator('[class="_filteritemtext_a6k5f_88"]').first).to_be_visible()
        self. page.get_by_text("Customs House Agent Name").first.click()
        self.page.wait_for_timeout(2000)
        expect(self.page.get_by_text("Filter by Customs House Agent Name").first).to_be_visible()

    def Customs_Filter_Search(self, C_H_A_N: str):
        # Use the search bar in the modal to search for a consignee (e.g., "Walmart")
        self.page.locator("//input[@type='text' and contains(@placeholder, 'Search in')]").last.fill(C_H_A_N)
        self.page.wait_for_timeout(2000)
        # Check for the presence of the company name in the results
        locators = self.page.locator('[class="text-no-transform"]')
        count = locators.count()
        print(f"🔍 Found {count} checkbox")

        if count == 0:
            print("❌ FAIL: No elements found for search")
            self.page.locator('input[type="text"]').last.clear()
        else:
            found_match = False

            for i in range(count):
                locator = locators.nth(i)

                if locator.is_visible():
                    actual_text = locator.inner_text().strip()

                    if C_H_A_N.lower() in actual_text.lower():  # case-insensitive compare
                        print(f"✅ PASS: Search is working → '{C_H_A_N}' found in element {i + 1}: '{actual_text}'")
                        expect(locator).to_contain_text(C_H_A_N)
                        found_match = True
                    else:
                        print(f"❌ FAIL: Search text '{C_H_A_N}' not found in element {i + 1}. Actual: '{actual_text}'")
                else:
                    print(f"⚠️ Element {i + 1} is not visible")

            if not found_match:
                print(f"❌ FAIL: Search text '{C_H_A_N}' not found in any visible element")
                self.page.locator('input[type="text"]').clear()

    from playwright.sync_api import expect

    def validate_customs_type_checkboxes(self):
        # Locate all checkboxes with their labels
        checkboxes = self.page.locator("//label[input[@type='checkbox']]")

        count = checkboxes.count()
        print(f"Total checkboxes found: {count}")

        for i in range(count):
            checkbox_label = checkboxes.nth(i)

            # Ensure checkbox is visible
            expect(checkbox_label).to_be_visible(timeout=5000)

            # Extract the label text (like "Importer (245,542,386)")
            text_value = checkbox_label.inner_text().strip()
            print(f"Checkbox {i + 1} with label: {text_value} found in customs house agent name filter")

            # Validate checkbox input exists inside label
            input_box = checkbox_label.locator("input[type='checkbox']")
            expect(input_box).to_be_visible()

            # You can also assert that text is not empty
            assert text_value != "", f"Checkbox {i + 1} has no text!"

    def Apply_Filter_Functionality(self):
        last_input = self.page.locator('input[type="text"]').last
        value = last_input.input_value().strip()

        if value == "":
            print("ℹ️  input box is already empty. Skipping clear()...")
        else:
            print(f"🧹 Clearing input box (current value: '{value}')")
            last_input.clear()
        self.page.wait_for_timeout(2000)
        # Click on the checkbox of the first result
        expect(self.page.locator('[class="text-no-transform"]').nth(1)).to_be_visible()
        self.customs_house_agent_selected = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Selected value is : {self.customs_house_agent_selected}")

        import re
        # Calculate total from all .grey-text elements
        total = 0

        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        shipper_record_clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        shipper_record_num = float(shipper_record_clean)
        if 0 <= shipper_record_num < 100_000:
            self.total_in_millions = f"{shipper_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{shipper_record_num / 1_000_000:.2f}M"

        print(self.total_in_millions)
        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        return False

    def Validate_ShipmentRecord_Count(self):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=10000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(3000)
        # Validate total shipment count in 'Shipment' tab
        record_count = self.page.locator(
            "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']"
        )

        # Get actual text
        actual_text = record_count.inner_text()

        # If–else validation
        if self.total_in_millions == actual_text:
            print(f"✅ Record count matches: {actual_text}")
        else:
            print(f"❌ Record count mismatch. Expected {self.total_in_millions}, Found {actual_text}")

    def Validate_Total_RecordCount_In_the_Table(self):
        total_record_text = self.page.locator('[class="trademo-table-count-text"]').inner_text()
        total_from_summary = total_record_text.split("of")[1].strip()
        print(f"🔢 Value after 'of': {total_from_summary}")
        print(f"total record count:{self.Total_Record_Count}")
        if total_from_summary == self.Total_Record_Count:
            print(f"✅ Record count matched: {total_from_summary}")
        else:
            print(f"❌ Record count mismatch! Summary: {total_from_summary}, Table: {self.Total_Record_Count}")

    def Validate_Customs_Name_label(self):
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')
        # In the Details section, the " Custom House Agent Name" label shows the count as (1)
        self.page.locator("//span[normalize-space()='Customs Details']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        self.page.locator("//span[normalize-space()='Customs Details']").click()
        self.page.wait_for_timeout(1000)

    def check_customs_house_agent_inthegrid_view(self, validate: bool = True, use_pagination: bool = False):
        self.page.locator("//a[@id='nav-home-tab']").click()

        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.C_H_field = None

        # Extract headers
        table_headers = []
        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip()
            table_headers.append(header_text)
            if header_text.lower() in ["s no", "sl no", "serial no"]:
                self.S_No = i
            elif header_text.lower() in ["country of import", "consignee of import"]:
                self.C_H_field = i

        # --- Handle missing "Country of Import" column ---
        if self.C_H_field is None:
            self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
            self.page.locator("//span[normalize-space()='custom details']").click()
            self.page.locator("//span[@data-for='Custom House Agent Nameelipsis']").click()
            self.page.locator("//button[normalize-space()='Save Configuration']").click()
            expect(self.page.get_by_text("View").last).to_be_visible(timeout=10000)

            # 🔄 Re-fetch headers after update
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()
            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text == "custom house agent name":
                    self.C_H_field = i
                    print("✅ 'Custom House Agent Name' column added successfully!")

        if self.C_H_field is None:
            raise Exception("❌ 'Custom House Agent Name' column not found even after customization.")

        print(f"✅ Found column indexes → S No: {self.S_No}, Custom House Agent Name: {self.C_H_field}")

        all_rows_data = []  # store all table rows (with full column values)

        # --- function to process current page ---
        def process_current_page():
            rows = self.page.locator("table tbody tr")
            row_count = rows.count()

            for i in range(row_count):
                row_data = []
                cells = rows.nth(i).locator("td")
                cell_count = cells.count()

                for j in range(cell_count):
                    cell_text = cells.nth(j).inner_text().strip()
                    row_data.append(cell_text)

                # Fallback: use row index if "S No" column missing
                sl_no = row_data[self.S_No] if self.S_No is not None else str(i + 1)

                consignee_country = row_data[self.C_H_field]
                consignee_country_norm = re.sub(r"\s+", " ", consignee_country).strip().lower()

                validation_status = "✅ Valid"
                error_type = None  # will hold "Grid" or "Popup"

                if validate:
                    # --- Grid check ---
                    if self.customs_house_agent_selected.lower() == consignee_country_norm:
                        print(f"✅ Grid valid: {consignee_country} at Sl. No: {sl_no}")
                    else:
                        print(f"❌ Grid invalid: {consignee_country} at Sl. No: {sl_no}")
                        validation_status = "❌ Invalid (Grid)"
                        error_type = "Grid"
                        self.invalid_entries.append({"slNo": sl_no, "Value": consignee_country, "Source": "Grid"})

                    # --- Popup check ---
                    view = rows.nth(i).locator("td").nth(1)
                    view.click()
                    popup_house_agent_name = self.page.locator(
                        "//span[normalize-space(text())='Custom House Agent Name']/parent::div/following-sibling::div").inner_text().strip().lower()

                    if self.customs_house_agent_selected.lower() == popup_house_agent_name:
                        print(f"✅ Shows valid: {popup_house_agent_name} for Sl. No: {sl_no} in view details")
                    else:
                        print(f"❌ Shows invalid: {popup_house_agent_name} for Sl. No: {sl_no} in view details")
                        validation_status = "❌ Invalid data in View details modal"
                        error_type = "Popup"
                        self.invalid_entries.append({"slNo": sl_no, "Value": popup_house_agent_name, "Source": "Popup"})

                    # Close popup
                    self.page.locator("//span[@aria-hidden='true']").click()

                row_data.append(validation_status)
                row_data.append(error_type)  # helper column to highlight later
                all_rows_data.append(row_data)

        # --- pagination loop ---
        if use_pagination:
            while True:
                process_current_page()
                next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
                if next_button.is_enabled():
                    next_button.click()
                    self.page.locator("table tbody tr td:nth-child(1)").first.wait_for(
                        state="visible", timeout=100000
                    )
                else:
                    break
        else:
            process_current_page()

        # --- save results ---
        if self.invalid_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Country of Import Data"

            # Headers (+Validation column)
            sheet.append(table_headers + ["Validation", "ErrorType"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            for row_data in all_rows_data:
                sheet.append(row_data)
                if row_data[-1] == "Grid":  # error type
                    col_idx = self.C_H_field + 1  # excel index (1-based)
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = red_fill
                elif row_data[-1] == "Popup":
                    col_idx = self.C_H_field + 1
                    sheet.cell(row=sheet.max_row, column=col_idx).fill = orange_fill

            file_path = f"results/customs_house_agent_Name_{self.customs_house_agent_selected}.xlsx"
            workbook.save(file_path)
            print(f"📁 All rows saved to {file_path} (invalids highlighted: red=Grid, orange=Popup)")
        else:
            print("📘 No invalid customs house agent name found. Excel not generated.")

        return self.invalid_entries




