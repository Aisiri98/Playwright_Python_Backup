import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page
from playwright.sync_api import expect

invalid_entries =[]
import difflib
def format_number(num_str: str) -> str:
    """Convert number string like '149,760' → '149.76k' or '2' → '2'."""
    num_str = num_str.replace(",", "").strip()

    try:
        num = int(num_str)
    except ValueError:
        raise ValueError(f"Invalid number string: {num_str}")

    if num >= 1_000_000:
        return f"{num / 1_000_000:.2f}M"
    elif num >= 1_000:
        return f"{num / 1_000:.2f}K"
    else:
        return str(num)
class ShipperFilter:
    def __init__(self, page: Page):
        self.C_O_field = None
        self.C_E_field = None
        self.S_JC_field = None
        self.page = page
        self.invalid_date_entries = []
        self.date_mismatches = []

    def Shipment_Filter_Option(self):
        # Shipper filter opens with sub-filter options
        shipper_standardized_locator = self.page.locator("//div[contains(text(),'Shipper Standardized Name')]")
        expect(shipper_standardized_locator).to_be_visible(timeout=100000)
        # Locate all elements with the given class
        elements = self.page.locator('[class="_filteritemtext_a6k5f_88"]')
        count = elements.count()

        print(f"Found {count} shipment options")

        # Loop through each element
        for i in range(count):
            element = elements.nth(i)
            if element.is_visible():
                text = element.inner_text()
                print(f"✅ Shipment option {i + 1} is visible: {text}")
            else:
                print(f"❌ Shipment option {i + 1} is not visible")

        # Click on the shipper filter again
        self.page.locator("//span[normalize-space()='Shipper']").click()

        # Check if each option is not visible
        # Locate all elements with the given class
        # Loop through each element
        for i in range(count):
            element = elements.nth(i)
            if element.is_visible():
                text = element.inner_text()
                print(f"✅ Shipment option {i + 1} is not visible: {text}")
            else:
                print(f"❌ Shipment option {i + 1} is visible")

    def open_shipper_dropdown(self):
        shipper_standardized_locator = self.page.locator("//div[contains(text(),'Shipper Standardized Name')]")

        # Check if "Consignee Standardized Name" is visible
        if not shipper_standardized_locator.is_visible():
            print("📌 Shipper dropdown is closed. Clicking to open...")
            self.page.locator("//span[normalize-space()='Shipper']").click()
        else:
            print("✅ Shipper dropdown is already open.")

    def Click_S_SName(self):
        self.page.locator("//div[contains(text(),'Shipper Standardized Name')]").click()
        self.page.wait_for_timeout(2000)
        expect(self.page.locator("//h5[normalize-space()='Filter by Shipper Standardized Name']")).to_be_visible()

    def Shipper_Filter_Search(self, S_S_Name:str):
        # Use the search bar in the modal to search for a Shipper (e.g., "Walmart")
        self.page.locator('input[type="text"]').last.fill(S_S_Name)
        self.page.wait_for_timeout(2000)
        # Check for the presence of the company name in the results
        locator = self.page.locator('[class="text-no-transform"]').first

        # Check if the element is visible
        if locator.is_visible():
            actual_text = locator.inner_text().strip()
            if S_S_Name.lower() in actual_text.lower():  # case-insensitive compare
                print(f"✅ PASS: Search is working → '{S_S_Name}' found in '{actual_text}'")
                expect(locator).to_contain_text(S_S_Name)  # Playwright assertion
            else:
                self.page.locator('input[type="text"]').clear()
                print(f"❌ FAIL: Search text '{S_S_Name}' not found. Actual: '{actual_text}'")
        else:
            # Element not visible → search not working
            self.page.locator('input[type="text"]').clear()
            print("❌ FAIL: Search is not working (element not visible)")

    def Apply_Filter_Functionality(self):
        last_input = self.page.locator('input[type="text"]').last
        value = last_input.input_value().strip()

        if value == "":
            print("ℹ️ Last input box is already empty. Skipping clear()...")
        else:
            print(f"🧹 Clearing last input box (current value: '{value}')")
            last_input.clear()
        self.page.wait_for_timeout(2000)
        # Click on the checkbox of the first result
        expect(self.page.locator('[class="text-no-transform"]').nth(1)).to_be_visible()
        self.Company_name = self.page.locator('[class="text-no-transform"]').nth(0).inner_text().strip()
        print(f"Selected Shipper Standardized Name: {self.Company_name}")

        import re
        # Calculate total from all .grey-text elements
        total = 0

        grey_text_value = self.page.locator('.grey-text').nth(0).inner_text().strip()
        self.Total_Record_Count = re.sub(r"[()]", "", grey_text_value)
        print(self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        shipper_record_clean = re.sub(r"[()\sA-Za-z,]", "", grey_text_value)

        # Convert to float
        shipper_record_num = float(shipper_record_clean)
        if 0 <= shipper_record_num < 100_000:
            self.total_in_millions = f"{shipper_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{shipper_record_num / 1_000_000:.2f}M"

        print(self.total_in_millions)
        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(0).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()
        return False

    def Validate_ShipmentRecord_Count(self):
        self.page.locator("//a[@id='nav-home-tab']").click(timeout=10000)
        expect(self.page.locator("//a[@id='nav-home-tab']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(3000)
        # Validate total shipment count in 'Shipment' tab
        expect(self.page.locator(
            "//span[@class=' tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium']")).to_contain_text(
            f"{self.total_in_millions}")

    def Validate_Total_RecordCount_In_the_Table(self):
        total_record_text = self.page.locator('[class="trademo-table-count-text"]').inner_text()
        total_from_summary = total_record_text.split("of")[1].strip()
        print(f"🔢 Value after 'of': {total_from_summary}")
        print(f"total record count:{self.Total_Record_Count}")
        if total_from_summary == self.Total_Record_Count:
            print(f"✅ Record count matched: {total_from_summary}")
        else:
            print(f"❌ Record count mismatch! Summary: {total_from_summary}, Table: {self.Total_Record_Count}")

    def Validate_S_S_Name_label(self):
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]')).to_contain_text('1')

        # Click on the shipper filter dropdown
        self.page.locator("//span[normalize-space()='Shipper']").click()
        expect(self.page.locator('[class="_badgefilterno_a6k5f_121"]').nth(1)).to_contain_text('1')
        self.page.locator("//span[normalize-space()='Shipper']").click()
        self.page.wait_for_timeout(1000)

    from collections import defaultdict
    from playwright.sync_api import expect

    def format_number(num_str: str) -> str:
        """Convert number string like '149,760' → '149.76k' or '2' → '2'."""
        num_str = num_str.replace(",", "").strip()

        try:
            num = int(num_str)
        except ValueError:
            raise ValueError(f"Invalid number string: {num_str}")

        if num >= 1_000_000:
            return f"{num / 1_000_000:.2f}M"
        elif num >= 1_000:
            return f"{num / 1_000:.2f}k"
        else:
            return str(num)

    def Validate_Exporter_Tab(self):
        self.page.locator("//a[@id='nav-contact-tab']").click()

        # Get record text like "Showing 1-10 of 149,760"
        total_record_text = self.page.locator(".trademo-table-count-text").inner_text().strip()

        if "of" in total_record_text:
            total_count_raw = total_record_text.split("of")[-1].strip()
        else:
            total_count_raw = total_record_text

        # Format total count (e.g. 149.76k)
        total_count = format_number(total_count_raw)
        print(f"✅ Total Exporter count: {total_count}")

        # Verify UI badge matches the formatted count
        badge_locator = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 '
            'tw-text-primary-purple-600 tw-font-medium"]'
        )
        expect(badge_locator).to_have_text(total_count)

        # Convert back to int for logical checks
        total_count_int = int(total_count_raw.replace(",", ""))
        return total_count

    def check_Shipper_Standardized_Name_Filter(self, validate: bool = True):
        """Validate shipper standardized name across pages and save invalid entries to Excel"""
        self.page.locator("//a[@id='nav-home-tab']").click()
        # Keep invalid entries consistent across recursive calls
        if not hasattr(self, "invalid_shipper_entries"):
            self.invalid_shipper_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.S_Name = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            if header_text in ["s no", "sl no", "serial no"]:  # flexible match
                self.S_No = i
            elif header_text == "shipper name":
                self.S_Name_field = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.S_Name_field is None:
            raise Exception("❌ 'Shipper name' column not found in table headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Shipper Name: {self.S_Name_field}")

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(self.S_No).inner_text().strip()
                view = rows.nth(i).locator("td").nth(1)

                # --- Shipper Name Validation ---
                shipper_name = rows.nth(i).locator("td").nth(self.S_Name_field).inner_text().strip()
                compare_shipper_Name = self.Company_name.lower()
                Shipper_name = re.sub(r"\s+", " ", shipper_name).strip().lower()

                if Shipper_name == compare_shipper_Name:
                    print(f"✅ Valid Shipper name found: {Shipper_name} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid Shipper name found: {Shipper_name} at Sl. No: {sl_no}")
                    self.invalid_shipper_entries.append({"slNo": sl_no, "Company": Shipper_name})

                # --- View Click + Read More ---
                view.click()
                read_more_locator = self.page.get_by_text("Read more").first
                self.page.wait_for_timeout(1000)

                if read_more_locator.is_visible():
                    print("🔍 'Read More' is visible, clicking...")
                    read_more_locator.click()
                else:
                    print("ℹ️ 'Read More' not visible.")

                # --- Shipper Standardized Name check ---
                Shipper_Standardized_Name = self.page.locator("//span[normalize-space(text())='Shipper Standardized Name'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a").inner_text().strip()
                if self.Company_name.strip().lower() == Shipper_Standardized_Name.strip().lower():
                    print(f"✅ Correct Shipper standardized name {Shipper_Standardized_Name} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Incorrect Shipper standardized name {Shipper_Standardized_Name} at Sl. No: {sl_no}")
                    self.invalid_shipper_entries.append({"slNo": sl_no, "Company": Shipper_Standardized_Name})

                # --- Popup validation ---
                Shipper_link = self.page.locator("//span[normalize-space(text())='Shipper Standardized Name'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//a")
                Shipper_page1 = Shipper_link.inner_text().strip()
                with self.page.expect_popup() as page1_info:
                    self.page.get_by_role("dialog").get_by_text(Shipper_page1).click()
                page1 = page1_info.value
                self.page.wait_for_timeout(2000)
                expect(page1.get_by_role("main")).to_contain_text(self.Company_name)
                page1.close()
                self.page.locator("//span[@aria-hidden='true']").click()

        # --- Pagination handling ---
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(state="visible", timeout=100000)
        #     self.check_Shipper_Standardized_Name_Filter(validate=validate)  # recursive call
        #     return self.invalid_shipper_entries  # recursive return
        # else:
        #     print("✅ All pages validated")

        # --- Save results only once at the end ---
        if not self.invalid_shipper_entries:
            print("📘 No invalid shipper name found.")
            return []

        # Save invalid entries to Excel
        os.makedirs("results", exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invalid Shipper Name"

        # Write headers
        sheet.append(["Sl. No", "Company Name"])

        # Red fill style for invalid cells
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for item in self.invalid_shipper_entries:
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1, value=item["slNo"])
            shipper_name_cell = sheet.cell(row=row, column=2, value=item["Company"])
            shipper_name_cell.fill = red_fill

        filename = f"results/invalid_Shipper_{self.Company_name}.xlsx"
        workbook.save(filename)
        print(f"📁 Invalid shipment data saved to '{filename}'")

        return self.invalid_shipper_entries

    def Manual_Search(self, S_S_Name:str,search_value):
        self.page.wait_for_timeout(3000)
        # In the search bar, manually enter a Shipper Name and click Search
        expect(self.page.get_by_placeholder(
            "Type to search in all categories or choose from the category below")).to_be_visible()
        # Manual suggest search for Shipper
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.get_by_role("textbox", name="Type to search in all").fill(S_S_Name)
        self.page.locator(".tw-bg-primary-purple-500").click()
        expect(self.page.locator("table tbody tr")).to_have_count(10, timeout=100000)
        self.page.wait_for_timeout(3000)
        Shipment_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').nth(
            0).inner_text().strip()
        print(f"shipment count:{Shipment_count}")
        Importer_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            0).inner_text().strip()
        print(f"Importer count:{Importer_count}")
        Exporter_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            1).inner_text().strip()
        print(f"Exporter count:{Exporter_count}")
        # Clear the search bar → Select “Shipper ” from category → Type partial name
        self.page.locator('input[type="text"]').clear()
        self.page.get_by_role("textbox", name="Type to search in all").click()
        self.page.locator("(//div[contains(text(),'Shipper')])[1]").click()
        self.page.get_by_role("textbox", name="Search upto 20 Shipper's name or address").fill(search_value)
        self.page.wait_for_timeout(1000)
        elements = self.page.locator('[class="tw-font-medium tw-bg-transparent tw-p-0"]')
        count = elements.count()

        print(f"🔍 Total elements found: {count}")

        for i in range(count):
            text = elements.nth(i).inner_text().strip()
            print(f"Checking element {i + 1}: {text}")

            if text.lower() == search_value.lower():  # case-insensitive match
                print(f"✅ Found '{search_value}' at index {i + 1}")
                elements.nth(i).click()
                return True  # ✅ exit function immediately after finding

        print(f"❌ '{search_value}' not found in {count} elements")
        self.page.wait_for_timeout(3000)
        self.page.locator("//a[@id='nav-home-tab']").click()
        self.page.wait_for_timeout(3000)
        Check_Shipment_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-primary-purple-100 tw-text-primary-purple-600 tw-font-medium"]').nth(0).inner_text().strip()
        if Shipment_count == Check_Shipment_count:
            print("✅ Shipment count is matching")
        else:
            print("❌ Shows wrong shipment count")

        Check_Importer_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            0).inner_text().strip()
        if Importer_count == Check_Importer_count:
            print("✅ Importer count is matching")
        else:
            print("❌ Shows wrong Importer count")
        Check_Exporter_count = self.page.locator(
            '[class=" tw-text-xs tw-p-1 tw-rounded tw-bg-gray-300 tw-text-gray-900 tw-font-normal"]').nth(
            1).inner_text().strip()
        if Exporter_count == Check_Exporter_count:
            print("✅ Exporter count is matching")
        else:
            print("❌ Shows wrong Exporter count")
        return False

    def Click_ShipperName_Option(self):
        self.page.get_by_text('Shipper Name').first.click()
        expect(self.page.locator("//h5[normalize-space()='Filter by Shipper Name']")).to_be_visible(timeout=10000)
        self.page.wait_for_timeout(2000)

    def Click_ShipperType_Option(self):
        self.page.locator("//div[contains(text(),'Shipper Type')]").first.click()
        expect(self.page.locator("//h5[normalize-space()='Filter by Shipper Type']")).to_be_visible(timeout=10000)
        expect(self.page.locator("//span[normalize-space()='Exporter']")).to_be_visible()
        expect(self.page.locator("//span[normalize-space()='Logistics Provider']")).to_be_visible()
        expect(self.page.locator("//span[normalize-space()='Potential Individual']")).to_be_visible()

    def normalize_company_name(self, name: str) -> str:
        """
        Normalize company names by shortening common terms.
        Example: "Limited" → "Ltd", "Company" → "Co".
        """
        # Lowercase and strip spaces
        name = name.lower().strip()

        # Replace common full forms with abbreviations
        replacements = {
            r"\blimited\b": "ltd",
        }

        for pattern, replacement in replacements.items():
            name = re.sub(pattern, replacement, name)

        # Collapse multiple spaces
        name = re.sub(r"\s+", " ", name)

        return name

    def check_Shipper_Name_intheGrid_View(self, validate: bool = True, threshold: float = 0.8):
        self.page.locator("//a[@id='nav-home-tab']").click()
        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.S_Name_field = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            if header_text in ["s no", "sl no", "serial no"]:  # flexible match
                self.S_No = i
            elif header_text == "shipper name":
                self.S_Name_field = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.S_Name_field is None:
            raise Exception("❌ 'Shipper name' column not found in table headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Shipper Name: {self.S_Name_field}")

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(self.S_No).inner_text().strip()
                view = rows.nth(i).locator("td").nth(1)

                # Read shipper name from table
                shipper_name = rows.nth(i).locator("td").nth(self.S_Name_field).inner_text().strip()

                expected = self.normalize_company_name(self.Company_name)
                actual = self.normalize_company_name(shipper_name)

                if expected == actual:
                    print(f"✅ PASS: Correct Shipper name '{actual}' at Sl. No: {sl_no}")
                else:
                    print(f"❌ FAIL: Invalid Shipper name found '{shipper_name}' at Sl. No: {sl_no} "
                          f"(Expected: {self.Company_name})")
                    self.invalid_entries.append({"slNo": sl_no, "Company": shipper_name})

                # --- open detail popup ---
                view.click()
                read_more_locator = self.page.get_by_text("Read more").first
                self.page.wait_for_timeout(1000)  # optional wait

                if read_more_locator.is_visible():
                    print("🔍 'Read More' is visible, clicking...")
                    read_more_locator.click()
                else:
                    print("ℹ️ 'Read More' not visible.")

                shipper_name_popup = self.page.locator("//span[normalize-space(text())='Shipper Name'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip()

                if self.Company_name.strip().lower() == shipper_name_popup.strip().lower():
                    print(f"✅ Popup shows correct Shipper name '{shipper_name_popup}' at Sl. No: {sl_no}")
                else:
                    print(f"❌ Popup shows incorrect Shipper name '{shipper_name_popup}' at Sl. No: {sl_no}")
                    self.invalid_entries.append({"slNo": sl_no, "Company": shipper_name_popup})

                # Close popup
                self.page.locator("//span[@aria-hidden='true']").click()

            # ✅ Handle pagination
            next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
            # if next_button.is_enabled():
            #     next_button.click()
            #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
            #         state="visible", timeout=100000
            #     )
            #     # Recursive call (do not reinitialize invalid_entries)
            #     self.check_Shipper_Name_intheGrid_View(validate=True)
            # else:
            #     print("✅ All pages validated")

            # Save results at the end
            if not next_button.is_enabled():
                if not self.invalid_entries:
                    print("📘 No invalid shipper name found.")
                    return self.invalid_entries

                os.makedirs("results", exist_ok=True)
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Invalid Shipper Name"

                # Write headers
                sheet.append(["Sl. No", "Company Name"])

                # Red fill style
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                for item in self.invalid_entries:
                    row = sheet.max_row + 1
                    sheet.cell(row=row, column=1, value=item["slNo"])
                    company_cell = sheet.cell(row=row, column=2, value=item["Company"])
                    company_cell.fill = red_fill

                workbook.save(f"results/invalid_Shipper_name_{self.Company_name}.xlsx")
                print("📁 Invalid shipment data saved to results/invalid_Shipper_name.xlsx")

        return self.invalid_entries

    def Shipper_Juridiction_option(self):
        self.page.wait_for_timeout(1000)
        self.page.locator("//div[contains(text(),'Shipper Jurisdiction Country')]").first.click()
        expect(self.page.get_by_text("Filter by Shipper Jurisdiction Country")).to_be_visible(timeout=10000)


    def Shipper_Country_Filter_Search(self,country:str):
        self.page.wait_for_timeout(1000)
        self.page.locator('input[type="text"]').last.fill(country)
        self.page.wait_for_timeout(2000)
        elements = self.page.locator("//span[@data-for='localize-filter-tooltip']")
        count = elements.count()
        found = False
        for i in range(count):
            if "india" in elements.nth(i).inner_text().lower():
                print(f"Searched country{country} is shown in the filter")
                break

        self.page.locator('input[type="text"]').last.clear()
        # Click on the checkbox of the first result
        expect(self.page.locator("//span[@data-for='localize-filter-tooltip']").nth(5)).to_be_visible()
        self.page.wait_for_timeout(3000)

    def Shipper_Country_Apply_Filter(self):
        Juridiction_countrywithcount = self.page.locator('[data-for = "localize-filter-tooltip"]').nth(
            0).inner_text().strip()
        match = re.match(r"([a-zA-Z\s]+)\(([\d,]+)\)", Juridiction_countrywithcount)
        if match:
            self.country_selelcted = match.group(1).strip()
            self.Total_Record_Count = match.group(2).strip()
            print("Country:", self.country_selelcted)
            print("Value:",  self.Total_Record_Count)
        # Remove parentheses, spaces, letters like 'M', and commas
        Country_record_clean = re.sub(r"[()\sA-Za-z,]", "",  self.Total_Record_Count)

        # Convert to float
        country_record_num = float(Country_record_clean)
        if 0 <= country_record_num < 100_000:
            self.total_in_millions = f"{country_record_num / 1_000:.2f}K"
        else:
            self.total_in_millions = f"{country_record_num / 1_000_000:.2f}M"

        # Convert to millions
        total_in_millions = f"{country_record_num / 1_000_000:.2f}"
        print(total_in_millions)

        # Click on the checkbox of the first result
        self.page.locator('input[type="checkbox"]').nth(1).click()

        # Click on 'Apply Filter' button
        self.page.locator('[class="btn btn-primary btn-sm trademo-new-btn"]').click()

    def validate_duplicate_company_names(self):
        # Locate all company names in the "Company Name" column
        company_elements = self.page.locator('[class="trademo-link text-left px-0 text-capitalize "]')
        count = company_elements.count()

        print(f"🔍 Total companies found: {count}")

        seen = set()
        duplicates = []

        for i in range(count):
            name = company_elements.nth(i).inner_text().strip()

            # Normalize text (convert Limited -> Ltd, lowercase etc.)
            normalized_name = name.lower().replace("limited", "ltd").strip()

            if normalized_name in seen:
                print(f"❌ Duplicate found: '{name}' at row {i + 1}")
                duplicates.append(name)
            else:
                seen.add(normalized_name)

        if not duplicates:
            print("✅ No duplicate company names found!")
        else:
            print(f"⚠️ Total Duplicates Found: {len(duplicates)} -> {duplicates}")

    def check_Shipper_Juridiction_Filter(self, validate: bool = True):
        self.page.locator("//a[@id='nav-home-tab']").click()
        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # ----------------- Subfunction to validate and add missing columns -----------------
        def validate_and_add_columns():
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()

            self.S_No = None
            self.S_JC_field = None

            # 🔍 First pass - check headers
            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text in ["s no", "sl no", "serial no"]:  # flexible match
                    self.S_No = i
                elif header_text == "shipper jurisdiction country":
                    self.S_JC_field = i

            # 🚨 If S No not found, fail
            if self.S_No is None:
                raise Exception("❌ 'S No' column not found in table headers")

            # ⚠️ If Shipper Jurisdiction Country not found, add from grid
            if self.S_JC_field is None:
                print("⚠️ 'Shipper Jurisdiction Country' column not found. Adding it...")

                self.page.locator("//span[normalize-space()='Customise Shipment Grid']").click()
                self.page.locator("//span[normalize-space()='shipper details']").click()

                # Add required columns
                self.page.locator("//span[@data-for='Shipper Jurisdiction Countryelipsis']").click()
                self.page.locator("//span[@data-for='Country of Originelipsis']").click()
                self.page.locator("//span[@data-for='Shipper Typeelipsis']").click()

                self.page.locator("//button[normalize-space()='Save Configuration']").click()
                self.page.wait_for_timeout(2000)

                # 🔄 Re-fetch headers after update
                headers = self.page.locator("table thead tr th")
                header_count = headers.count()

                for i in range(header_count):
                    header_text = headers.nth(i).inner_text().strip().lower()
                    if header_text == "shipper jurisdiction country":
                        self.S_JC_field = i
                        print("✅ 'Shipper Jurisdiction Country' column added successfully!")

                if self.S_JC_field is None:
                    raise Exception("❌ Failed to add 'Shipper Jurisdiction Country' column")

        # ✅ Call column validator
        validate_and_add_columns()

        # ----------------- Row iteration -----------------
        invalid_shipper_entries = []
        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(self.S_No).inner_text().strip()
                view = rows.nth(i).locator("td").nth(1)

                # Exported country
                shipper_juridiction = rows.nth(i).locator("td").nth(self.S_JC_field).inner_text().strip()
                country = re.sub(r"\s+", " ", shipper_juridiction).strip().lower()
                compare_country_juridiction = country.lower()

                if self.country_selelcted.lower() == compare_country_juridiction:
                    print(f"✅ Valid Shipper juridicition country name found: {country} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid Shipperjuridicition country name found: {country} at Sl. No: {sl_no}")
                    invalid_shipper_entries.append({"slNo": sl_no, "Company": country})

                # Open details view
                view.click()
                juridiction_country = self.page.locator("//span[normalize-space(text())='Shipper Jurisdiction Country'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip()

                if country == juridiction_country.strip().lower():
                    print(f"✅ Shows correct Shipper Juridiction country {juridiction_country} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Shows incorrect Shipper Juridiction country {juridiction_country} at Sl. No: {sl_no}")
                    invalid_shipper_entries.append({"slNo": sl_no, "Company": juridiction_country})

                # Close details
                self.page.locator("//span[@aria-hidden='true']").click()

        # ----------------- Pagination -----------------
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
        #         state="visible", timeout=100000
        #     )
        #     more_invalids = self.check_Shipper_Juridiction_Filter()
        #     invalid_shipper_entries.extend(more_invalids)
        # else:
        #     print("✅ All pages validated")

        # ----------------- Save invalid entries -----------------
        if self.page.locator('[class="trademo-table-arrow-button"]').nth(1).is_disabled():
            if not invalid_shipper_entries:
                print("📘 No invalid shipper name found.")
                return invalid_shipper_entries

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Shipper Name"

            # Headers
            sheet.append(["Sl. No", "Company Name"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in invalid_shipper_entries:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=item["slNo"])
                cell = sheet.cell(row=row, column=2, value=item["Company"])
                cell.fill = red_fill

            workbook.save(f"results/invalid_Shipper_Juridictionname.xlsx")
            print("📁 Invalid shipment data saved")

        return invalid_shipper_entries

    def check_Shipper_Type_inthgrid_view(self, validate: bool = True, invalid_shipper_entries=None):
        self.page.locator("//a[@id='nav-home-tab']").click()
        if invalid_shipper_entries is None:
            invalid_shipper_entries = []

        headers = self.page.locator("table thead tr th")
        header_count = headers.count()

        self.S_No = None
        self.S_Type_field = None

        for i in range(header_count):
            header_text = headers.nth(i).inner_text().strip().lower()
            if header_text in ["s no", "sl no", "serial no"]:  # flexible match
                self.S_No = i
            elif header_text == "shipper type":
                self.S_Type_field = i

        if self.S_No is None:
            raise Exception("❌ 'S No' column not found in table headers")
        if self.S_Type_field is None:
            raise Exception("❌ 'Shipper type' column not found in table headers")

        print(f"✅ Found column indexes → S No: {self.S_No}, Shipper Type: {self.S_Type_field}")

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        for i in range(row_count):
            sl_no = rows.nth(i).locator("td").nth(self.S_No).inner_text().strip()
            view = rows.nth(i).locator("td").nth(1)
            shipper_type = rows.nth(i).locator("td").nth(self.S_Type_field).inner_text().strip()

            if self.Company_name.lower() == shipper_type.lower():
                print(f"✅ Valid Shipper type found: {shipper_type} at Sl. No: {sl_no}")
            else:
                print(f"❌ Invalid Shipper type found: {shipper_type} at Sl. No: {sl_no}")
                invalid_shipper_entries.append({"slNo": sl_no, "ShipperType": shipper_type})

            # Click and validate popup
            view.click()
            self.page.wait_for_timeout(1000)

            Shipper_type_popup = self.page.locator("//span[normalize-space(text())='Shipper Type'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip()

            if self.Company_name.strip().lower() == Shipper_type_popup.lower():
                print(f"✅ Popup shows correct Shipper type {Shipper_type_popup} at Sl. No: {sl_no}")
            else:
                print(f"❌ Popup shows incorrect Shipper type {Shipper_type_popup} at Sl. No: {sl_no}")
                invalid_shipper_entries.append({"slNo": sl_no, "ShipperType": Shipper_type_popup})

            self.page.locator("//span[@aria-hidden='true']").click()

        # --- Next Page Handling ---
        next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        if next_button.is_enabled():
            next_button.click()
            self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
                state="visible", timeout=100000
            )
            # Recursive call, pass the same list
            return self.check_Shipper_Type_inthgrid_view(validate=validate, invalid_shipper_entries=invalid_shipper_entries)

        print("✅ All pages validated")

        # --- Save results if invalids exist ---
        if invalid_shipper_entries:
            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Shipper Type"

            # Write headers
            sheet.append(["Sl. No", "Shipper Type"])

            # Red fill for invalid
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in invalid_shipper_entries:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=item["slNo"])
                shipper_cell = sheet.cell(row=row, column=2, value=item["ShipperType"])
                shipper_cell.fill = red_fill

            file_path = "results/invalid_Shipper_Type.xlsx"
            workbook.save(file_path)
            print(f"📁 Invalid shipper data saved to {file_path}")
        else:
            print("📘 No invalid shipper type found.")

        return invalid_shipper_entries

    def Country_of_Export_Option(self):
        self.page.get_by_text("Country of Export").first.click()
        expect(self.page.get_by_text("Filter by Country of Export")).to_be_visible(timeout=10000)

    def check_country_of_export_inthegrid_view(self, validate: bool = True):
        self.page.locator("//a[@id='nav-home-tab']").click()

        # Keep invalid_entries consistent across pages
        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # ----------------- Subfunction to validate and add missing columns -----------------
        def validate_and_add_columns():
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()

            self.S_No = None
            self.C_E_field = None

            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text in ["s no", "sl no", "serial no"]:
                    self.S_No = i
                elif header_text == "country of export":
                    self.C_E_field = i

            if self.S_No is None:
                raise Exception("❌ 'S No' column not found in table headers")

            if self.C_E_field is None:
                raise Exception("❌ 'Country of Export' column not found in table headers")

        # ✅ Call once before validation
        validate_and_add_columns()

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(self.S_No).inner_text().strip()
                view = rows.nth(i).locator("td").nth(1)

                # Exported country (grid)
                country_actual = rows.nth(i).locator("td").nth(self.C_E_field).inner_text().strip()
                country = re.sub(r"\s+", " ", country_actual).strip().lower()

                if self.country_selelcted.lower() == country:
                    print(f"✅ Valid country of export: {country} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid country of export: {country} at Sl. No: {sl_no}")
                    self.invalid_entries.append({"slNo": sl_no, "Company": country})

                # Open details view
                view.click()
                country_export = self.page.locator("//span[normalize-space(text())='Country of Export'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip()

                if country.lower() == country_export.strip().lower():
                    print(f"✅ Details view matches country: {country_export} (Sl. No: {sl_no})")
                else:
                    print(f"❌ Mismatch in details view → {country_export} (Sl. No: {sl_no})")
                    self.invalid_entries.append({"slNo": sl_no, "Company": country_export})

                self.page.locator("//span[@aria-hidden='true']").click()

        # ----------------- Next page handling -----------------
        # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
        # if next_button.is_enabled():
        #     next_button.click()
        #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
        #         state="visible", timeout=100000
        #     )
        #     self.check_country_of_export_inthegrid_view(validate=True)  # recursive call
        # else:
        #     print("✅ All pages validated")

            # ----------------- Save invalid entries -----------------
            if not self.invalid_entries:
                print("📘 No invalid country of export found.")
                return []

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Country of Export"

            # Headers
            sheet.append(["Sl. No", "Company Name"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in self.invalid_entries:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=item["slNo"])
                cell = sheet.cell(row=row, column=2, value=item["Company"])
                cell.fill = red_fill

            safe_country = re.sub(r"[^\w\-]", "_", self.country_selelcted)
            file_path = f"results/invalid_country_of_export_{safe_country}.xlsx"
            workbook.save(file_path)
            print(f"📁 Invalid country of export data saved → {file_path}")

        return self.invalid_entries

    def check_country_of_origin_inthegrid_view(self, validate: bool = True):
        self.page.locator("//a[@id='nav-home-tab']").click()

        # Keep invalid_entries consistent across pages
        if not hasattr(self, "invalid_entries"):
            self.invalid_entries = []

        # ----------------- Subfunction to validate and add missing columns -----------------
        def validate_and_add_columns():
            headers = self.page.locator("table thead tr th")
            header_count = headers.count()

            self.S_No = None
            self.C_O_field = None

            for i in range(header_count):
                header_text = headers.nth(i).inner_text().strip().lower()
                if header_text in ["s no", "sl no", "serial no"]:
                    self.S_No = i
                elif header_text == "country of origin":
                    self.C_O_field = i

            if self.S_No is None:
                raise Exception("❌ 'S No' column not found in table headers")

            if self.C_O_field is None:
                raise Exception("❌ 'Country of origin' column not found in table headers")

        # ✅ Call once before validation
        validate_and_add_columns()

        rows = self.page.locator("table tbody tr")
        row_count = rows.count()

        if validate:
            for i in range(row_count):
                sl_no = rows.nth(i).locator("td").nth(self.S_No).inner_text().strip()
                view = rows.nth(i).locator("td").nth(1)

                # Exported country (grid)
                country_actual = rows.nth(i).locator("td").nth(self.C_O_field).inner_text().strip()
                country = re.sub(r"\s+", " ", country_actual).strip().lower()

                if self.country_selelcted.lower() == country:
                    print(f"✅ Valid country of origin: {country} at Sl. No: {sl_no}")
                else:
                    print(f"❌ Invalid country of origin: {country} at Sl. No: {sl_no}")
                    self.invalid_entries.append({"slNo": sl_no, "Company": country})

                # Open details view
                view.click()
                Origin_country = self.page.locator("//span[normalize-space(text())='Country of Origin'] /ancestor::div[contains(@class,'col-5')] /following-sibling::div[contains(@class,'col-7')]//div[1]").inner_text().strip()

                if country.lower() == Origin_country.strip().lower():
                    print(f"✅ Details view matches country: {Origin_country} (Sl. No: {sl_no})")
                else:
                    print(f"❌ Mismatch in details view → {Origin_country} (Sl. No: {sl_no})")
                    self.invalid_entries.append({"slNo": sl_no, "Company": Origin_country})

                self.page.locator("//span[@aria-hidden='true']").click()

            # ----------------- Next page handling -----------------
            # next_button = self.page.locator('[class="trademo-table-arrow-button"]').nth(1)
            # if next_button.is_enabled():
            #     next_button.click()
            #     self.page.locator("table tbody tr td:nth-child(4)").first.wait_for(
            #         state="visible", timeout=100000
            #     )
            #     self.check_country_of_export_inthegrid_view(validate=True)  # recursive call
            # else:
            #     print("✅ All pages validated")

            # ----------------- Save invalid entries -----------------
            if not self.invalid_entries:
                print("📘 No invalid country of export found.")
                return []

            os.makedirs("results", exist_ok=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invalid Country of Export"

            # Headers
            sheet.append(["Sl. No", "Company Name"])

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for item in self.invalid_entries:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=item["slNo"])
                cell = sheet.cell(row=row, column=2, value=item["Company"])
                cell.fill = red_fill

            safe_country = re.sub(r"[^\w\-]", "_", self.country_selelcted)
            file_path = f"results/invalid_country_of_origin_{safe_country}.xlsx"
            workbook.save(file_path)
            print(f"📁 Invalid country of export data saved → {file_path}")

        return self.invalid_entries










